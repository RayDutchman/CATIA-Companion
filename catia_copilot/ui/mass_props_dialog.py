"""
重量、重心、惯量统计对话框模块。

提供：
- MassPropsDialog – 遍历产品树，展示每个零件实例的质量/重心/转动惯量，
                    支持：
                      • 手动编辑重量（等比缩放惯量，联动同型号零件）
                      • 层级BOM / 汇总BOM 切换
                      • 重量单位 g/kg 独立选择
                      • 长度单位 mm/m 独立选择
                      • 惯量单位 g·mm²/g·m²/kg·mm²/kg·m² 独立选择（4 种）
                      • 惯量包络体读取模式：只读.1 / 最大编号 / 全部汇总
                      • 文件名 / 零件编号 / 术语 / 版本列可隐藏
                      • 计算装配体总质量特性并导出 Excel
"""

import logging
import math
import subprocess
from pathlib import Path

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTreeWidgetItem, QHeaderView, QAbstractItemView,
    QCheckBox, QGroupBox, QMessageBox, QApplication,
    QFileDialog, QProgressDialog, QLineEdit, QGridLayout, QFrame,
    QRadioButton, QButtonGroup, QWidget, QComboBox,
    QStyledItemDelegate, QMenu,
)
from PySide6.QtGui import QBrush, QColor
from PySide6.QtCore import Qt, QSettings

from catia_copilot.constants import (
    MASS_PROPS_COLUMNS,
    MASS_PROPS_COLUMN_DISPLAY_NAMES,
    MASS_PROPS_HIDEABLE_COLUMNS,
    MASS_PROPS_READONLY_COLUMNS,
    FILENAME_NOT_FOUND,
    FILENAME_UNSAVED,
)
from catia_copilot.catia.mass_props_collect import (
    collect_mass_props_rows, _row_inertia_to_root, recompute_product_rows,
    save_rows, load_rows, MAX_INERTIA_INDEX, remeasure_part_mass_props,
)
from catia_copilot.catia.mass_props_calc import rollup_mass_properties
from catia_copilot.ui.bom_widgets import _BomTreeWidget

logger = logging.getLogger(__name__)

# UserRole：行索引（映射到 self._rows）
_ROW_IDX_ROLE = Qt.ItemDataRole.UserRole
# UserRole+1：锁定标志位（不可编辑行）
_ITEM_LOCKED_ROLE = Qt.ItemDataRole.UserRole + 1

# 惯量列名 → (行索引, 列索引)，对应 3×3 张量位置
_INERTIA_IDX: dict[str, tuple[int, int]] = {
    "Ixx": (0, 0), "Iyy": (1, 1), "Izz": (2, 2),
    "Ixy": (0, 1), "Ixz": (0, 2), "Iyz": (1, 2),
}

# 显示值随当前单位制变化的列名
_UNIT_SENSITIVE_COLUMNS: tuple[str, ...] = (
    "Weight",
) + tuple(_INERTIA_IDX.keys()) + ("CogX", "CogY", "CogZ")
_SUMMARY_SORT_COLUMNS: list[str] = [
    "Part Number", "Nomenclature", "Revision", "Filename", "Weight",
    "CogX", "CogY", "CogZ",
]

# 数值格式化：判断"接近整数"的绝对容差（用于 _fmt / _fmt_scaled）
_INTEGER_ABS_TOL: float = 1e-9


class _MassPropsDelegate(QStyledItemDelegate):
    """只允许对未锁定零件行的"Weight"列进行编辑，其余列一律只读。"""

    def __init__(self, cols_fn, tree) -> None:
        super().__init__(tree)
        self._cols_fn = cols_fn  # callable: () -> list[str]

    def createEditor(self, parent, option, index):
        tree = self.parent()
        item = tree.itemFromIndex(index)
        if item is None:
            return None
        if item.data(0, _ITEM_LOCKED_ROLE):
            return None
        cols = self._cols_fn()
        if index.column() >= len(cols):
            return None
        col_name = cols[index.column()]
        if col_name in MASS_PROPS_READONLY_COLUMNS:
            return None
        return super().createEditor(parent, option, index)


def _fmt(value) -> str:
    """数值 → 字符串，None → '—'（不含单位换算）。"""
    if value is None:
        return "—"
    try:
        v = float(value)
        if math.isclose(v, round(v), rel_tol=0.0, abs_tol=_INTEGER_ABS_TOL):
            return f"{v:.0f}"
        if abs(v) >= 1e5 or (v != 0.0 and abs(v) < 0.001):
            return f"{v:.3e}"
        return f"{v:.3f}"
    except (TypeError, ValueError):
        return str(value)


class MassPropsDialog(QDialog):
    """重量、重心、惯量统计对话框。

    - 遍历 CATProduct 树，每个节点（零件/产品/部件实例）单独显示一行（层级BOM模式）。
      Weight / CogX / CogY / CogZ / Ixx–Iyz 均在根产品坐标系下显示，与装配位置有关。
    - 汇总BOM模式：相同零件编号的零件实例合并为一行，并显示数量（Quantity）；
      仅列出零件（不含产品和部件）；Weight / CogX / CogY / CogZ / Ixx–Iyz
      在零件自身坐标系下显示，与装配位置无关。
    - 仅零件节点的"重量"列可编辑；修改后等比缩放该行惯量，
      并同步更新所有相同零件编号的行（及 _rows 中全部同PN数据）。
    - 单位可在 kg/g 间切换（影响重量列与转动惯量列的显示和导出）。
    - "计算"按钮汇总装配体总质量特性（考虑位姿变换）。
    - "导出表格"将当前数据（含汇总行）写入 Excel。
    """

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("重量、重心、惯量统计")
        self.setMinimumSize(1100, 650)
        self.resize(1300, 750)

        self._settings = QSettings("CATIACompanion", "MassPropsDialog")
        self._last_browse_dir: str = self._settings.value("last_browse_dir", "")

        # ── 持久化显示选项 ─────────────────────────────────────────────────
        saved_hid = self._settings.value("visible_hideable_cols", list(MASS_PROPS_HIDEABLE_COLUMNS))
        if saved_hid is None:
            saved_hid = list(MASS_PROPS_HIDEABLE_COLUMNS)
        elif isinstance(saved_hid, str):
            saved_hid = [saved_hid]
        else:
            saved_hid = list(saved_hid)
        self._visible_hideable_cols: set[str] = {
            c for c in saved_hid if c in MASS_PROPS_HIDEABLE_COLUMNS
        }

        self._summarize: bool = self._settings.value("summarize", False, type=bool)

        # ── 单位制 ────────────────────────────────────────────────────────────
        self._mass_unit: str = self._settings.value("mass_unit", "g")
        self._cog_unit: str = self._settings.value("cog_unit", "mm")
        if self._mass_unit not in ("g", "kg"):
            self._mass_unit = "g"
        if self._cog_unit not in ("mm", "m"):
            self._cog_unit = "mm"

        # 惯量单位独立选择（4 种）
        _valid_inertia_units = ("g\u00b7mm\u00b2", "g\u00b7m\u00b2", "kg\u00b7mm\u00b2", "kg\u00b7m\u00b2")
        self._inertia_unit: str = self._settings.value("inertia_unit", "g\u00b7mm\u00b2")
        if self._inertia_unit not in _valid_inertia_units:
            self._inertia_unit = "g\u00b7mm\u00b2"

        # 内部单位为 SI（kg / m / kg·m²）；根据所选显示单位制设置换算因子
        self._unit_factor, _, self._cog_unit_factor = (
            self._calc_unit_factors(self._mass_unit, self._cog_unit)
        )
        self._inertia_unit_factor = self._calc_inertia_factor(self._inertia_unit)

        # ── 读取模式 ─────────────────────────────────────────────────────────
        self._read_mode: str = self._settings.value("read_mode", "all")
        if self._read_mode not in ("first", "last", "all"):
            self._read_mode = "all"

        # ── 忽略隐藏节点 ──────────────────────────────────────────────────────
        self._skip_hidden: bool = self._settings.value("skip_hidden", False, type=bool)

        # ── 汇总BOM专用选项 ───────────────────────────────────────────────────
        self._summary_sort_column: str = self._settings.value(
            "summary_sort_column", ""
        )

        # ── 内部状态 ──────────────────────────────────────────────────────
        self._rows: list[dict] = []
        # display_row_idx → QTreeWidgetItem
        self._item_by_row: list[QTreeWidgetItem] = []
        # Part Number → list[QTreeWidgetItem] (all visible items with that PN)
        self._pn_to_items: dict[str, list[QTreeWidgetItem]] = {}
        self._is_updating: bool = False
        self._rollup_result: dict | None = None
        self._loaded: bool = False
        self._col_widths: dict[str, int] = {}

        # 列名列表在可见性或模式改变时重建
        self._columns: list[str] = self._build_columns()

        self._build_ui()

    # ── 列管理 ──────────────────────────────────────────────────────────────

    @staticmethod
    def _calc_unit_factors(mass_unit: str, cog_unit: str) -> tuple[float, float, float]:
        """根据重量单位和长度单位返回 (mass_factor, inertia_factor_derived, cog_factor)。

        内部存储单位为 SI：质量 kg、坐标 m、惯量 kg·m²。
        inertia_factor_derived = mass_factor × cog_factor²（用于向后兼容推导）。
        独立惯量单位换算请使用 _calc_inertia_factor()。
        """
        mf = 1e3 if mass_unit == "g" else 1.0
        cf = 1e3 if cog_unit == "mm" else 1.0
        return mf, mf * cf * cf, cf

    @staticmethod
    def _calc_inertia_factor(inertia_unit: str) -> float:
        """从 SI 内部单位 kg·m² 换算到 inertia_unit 字符串对应显示单位的换算因子。

        支持的惯量单位字符串（Unicode 上标²）：
          "g·mm²"  → 1e9   (kg→g=×1e3, m→mm=×1e3, m²→mm²=×1e6, 合计×1e9)
          "g·m²"   → 1e3   (kg→g=×1e3, m→m=×1)
          "kg·mm²" → 1e6   (kg→kg=×1, m→mm=×1e3, m²→mm²=×1e6)
          "kg·m²"  → 1.0   (SI，无需换算)
        """
        _map = {
            "g\u00b7mm\u00b2":  1e9,
            "g\u00b7m\u00b2":   1e3,
            "kg\u00b7mm\u00b2": 1e6,
            "kg\u00b7m\u00b2":  1.0,
        }
        return _map.get(inertia_unit, 1.0)

    def _weight_unit_label(self) -> str:
        """返回重量列的单位标签字符串。"""
        return self._mass_unit

    def _inertia_unit_label(self) -> str:
        """返回惯量列的单位标签字符串（当前独立选择的惯量单位）。"""
        return self._inertia_unit

    def _cog_unit_label(self) -> str:
        """返回重心坐标列的单位标签字符串。"""
        return self._cog_unit

    def _build_columns(self) -> list[str]:
        """根据当前可见性设置和 BOM 模式，构建列名列表。

        层级BOM：Level 在第 0 列（承载树形装饰线），# 在第 1 列。
        汇总BOM：无 Level 列，# 在第 0 列（无装饰线需求），增加 Quantity 列。
        """
        if self._summarize:
            base = ["#", "Type"]
            for c in MASS_PROPS_HIDEABLE_COLUMNS:
                if c in self._visible_hideable_cols:
                    base.append(c)
            base += ["Quantity", "Weight", "CogX", "CogY", "CogZ",
                     "Ixx", "Iyy", "Izz", "Ixy", "Ixz", "Iyz"]
        else:
            base = ["Level", "#", "Type"]
            for c in MASS_PROPS_HIDEABLE_COLUMNS:
                if c in self._visible_hideable_cols:
                    base.append(c)
            base += ["Weight", "CogX", "CogY", "CogZ",
                     "Ixx", "Iyy", "Izz", "Ixy", "Ixz", "Iyz"]
        return base

    def _column_header(self, col_name: str) -> str:
        """返回列名的中文显示名（含当前单位后缀）。"""
        if col_name == "Weight":
            return f"重量 ({self._weight_unit_label()})"
        if col_name in _INERTIA_IDX:
            return f"{col_name} ({self._inertia_unit_label()})"
        if col_name in ("CogX", "CogY", "CogZ"):
            return f"{col_name} ({self._cog_unit_label()})"
        return MASS_PROPS_COLUMN_DISPLAY_NAMES.get(col_name, col_name)

    def _display_headers(self) -> list[str]:
        return [self._column_header(c) for c in self._columns]

    @staticmethod
    def _fmt_scaled(value, factor: float) -> str:
        """将原始 SI 值乘以换算因子后格式化为字符串。

        None → '—'；整数值（误差 < _INTEGER_ABS_TOL）→ 无小数位；
        |v| ≥ 1e5 或绝对值极小（0 < |v| < 0.001）→ 科学计数法；
        其余 → 保留三位小数。
        """
        if value is None:
            return "—"
        try:
            v = float(value) * factor
            if math.isclose(v, round(v), rel_tol=0.0, abs_tol=_INTEGER_ABS_TOL):
                return f"{v:.0f}"
            if abs(v) >= 1e5 or (v != 0.0 and abs(v) < 0.001):
                return f"{v:.3e}"
            return f"{v:.3f}"
        except (TypeError, ValueError):
            return str(value)

    def _fmt_mass_val(self, value) -> str:
        """将质量原始值（kg，SI 内部单位）乘以 _unit_factor 并格式化为字符串（重量列专用）。"""
        return self._fmt_scaled(value, self._unit_factor)

    def _fmt_inertia_val(self, value) -> str:
        """将惯量原始值（kg·m²，SI 内部单位）乘以 _inertia_unit_factor 并格式化为字符串（惯量列专用）。"""
        return self._fmt_scaled(value, self._inertia_unit_factor)

    def _fmt_cog_val(self, value) -> str:
        """将重心坐标原始值（m，SI 内部单位）乘以 _cog_unit_factor 并格式化为字符串（CogX/Y/Z 列专用）。"""
        return self._fmt_scaled(value, self._cog_unit_factor)

    # ── UI 构建 ────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── 前提条件说明（窗口过窄时允许截断）──────────────────────────────
        prereq_lbl = QLabel(
            "⚠ 使用说明：本功能读取指定产品树下的每个零件的'测量惯量'结果、"
            "在根产品中的位置，计算出根产品的重量、重心、转动惯量。"
            "请在 CATIA 中 <b>单独打开</b> 每个零件,执行'测量惯量'并勾选 <b>保持测量</b>,"
            f"测量结果必须命名为 <b>惯量包络体.x</b>（x 为 1–{MAX_INERTIA_INDEX} 的整数）。"
            "在产品窗口中建立的惯量包络体的坐标系为根产品坐标系（即使当前工作对象是零件），"
            "这会导致坐标系与根产品不重合的零件的测量结果不正确。"
            "支持一个零件具有多个惯量包络体，产品的惯量包络体将不被读取。"
        )
        prereq_lbl.setWordWrap(True)
        prereq_lbl.setMaximumHeight(46)
        prereq_lbl.setStyleSheet(
            "QLabel { background-color: #FFF8E1; border: 1px solid #F9A825;"
            " border-radius: 4px; padding: 4px 8px; color: #5D4037; font-size: 11px; }"
        )
        layout.addWidget(prereq_lbl)

        # ── 数据来源选择 ────────────────────────────────────────────────────
        self._use_active_chk = QCheckBox("使用当前CATIA活动文档（不选择文件）")
        self._use_active_chk.toggled.connect(self._toggle_file_row)
        layout.addWidget(self._use_active_chk)

        file_row = QHBoxLayout()
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("选择一个 CATProduct 文件…")
        self._file_edit.setReadOnly(True)
        self._file_browse_btn = QPushButton("浏览…")
        self._file_browse_btn.clicked.connect(self._browse_file)
        self._load_btn = QPushButton("加载")
        self._load_btn.clicked.connect(self._load_data)
        self._load_json_btn = QPushButton("载入已保存数据…")
        self._load_json_btn.setToolTip("从之前保存的数据文件中载入质量特性（无需打开CATIA）")
        self._load_json_btn.clicked.connect(self._load_data_from_json)
        file_row.addWidget(self._file_edit)
        file_row.addWidget(self._file_browse_btn)
        file_row.addWidget(self._load_btn)
        file_row.addWidget(self._load_json_btn)
        layout.addLayout(file_row)

        # ── 选项面板（2 行）────────────────────────────────────────────────
        opts_group = QGroupBox("读取与显示选项")
        opts_main = QVBoxLayout(opts_group)
        opts_main.setSpacing(4)
        opts_main.setContentsMargins(8, 6, 8, 6)

        # ── 第一行：BOM类型 ｜ 读取模式 ｜ 显示列 ──────────────────────────
        row1 = QHBoxLayout()
        row1.setSpacing(6)

        # BOM 类型
        self._bom_type_group = QButtonGroup(self)
        self._radio_hier = QRadioButton("层级BOM")
        self._radio_summ = QRadioButton("汇总BOM")
        self._radio_summ.setMinimumHeight(24)
        self._radio_hier.setChecked(not self._summarize)
        self._radio_summ.setChecked(self._summarize)
        self._bom_type_group.addButton(self._radio_hier)
        self._bom_type_group.addButton(self._radio_summ)
        self._radio_summ.toggled.connect(self._on_bom_type_changed)
        row1.addWidget(QLabel("BOM:"))
        row1.addWidget(self._radio_hier)
        row1.addWidget(self._radio_summ)

        _sep1 = QFrame(); _sep1.setFrameShape(QFrame.Shape.VLine)
        _sep1.setFrameShadow(QFrame.Shadow.Sunken)
        row1.addSpacing(4); row1.addWidget(_sep1); row1.addSpacing(4)

        # 读取模式
        self._read_mode_group = QButtonGroup(self)
        self._radio_read_first = QRadioButton("只读.1")
        self._radio_read_last  = QRadioButton("最大编号")
        self._radio_read_all   = QRadioButton("全部汇总")
        self._radio_read_first.setToolTip('仅读取名为"惯量包络体.1"的保持测量结果')
        self._radio_read_last.setToolTip("扫描所有编号，使用编号最大的有效保持测量结果")
        self._radio_read_all.setToolTip("读取所有有效的惯量包络体测量，并按平行轴定理汇总为单一质量特性")
        self._radio_read_first.setChecked(self._read_mode == "first")
        self._radio_read_last.setChecked(self._read_mode == "last")
        self._radio_read_all.setChecked(self._read_mode == "all")
        self._read_mode_group.addButton(self._radio_read_first)
        self._read_mode_group.addButton(self._radio_read_last)
        self._read_mode_group.addButton(self._radio_read_all)
        self._radio_read_first.toggled.connect(self._on_read_mode_changed)
        self._radio_read_last.toggled.connect(self._on_read_mode_changed)
        self._radio_read_all.toggled.connect(self._on_read_mode_changed)
        row1.addWidget(QLabel("惯量包络体读取:"))
        row1.addWidget(self._radio_read_first)
        row1.addWidget(self._radio_read_last)
        row1.addWidget(self._radio_read_all)

        _sep2 = QFrame(); _sep2.setFrameShape(QFrame.Shape.VLine)
        _sep2.setFrameShadow(QFrame.Shadow.Sunken)
        row1.addSpacing(4); row1.addWidget(_sep2); row1.addSpacing(4)

        # 显示列
        row1.addWidget(QLabel("显示列:"))
        self._hid_col_checks: dict[str, QCheckBox] = {}
        for col_name in MASS_PROPS_HIDEABLE_COLUMNS:
            cb = QCheckBox(MASS_PROPS_COLUMN_DISPLAY_NAMES.get(col_name, col_name))
            cb.setChecked(col_name in self._visible_hideable_cols)
            cb.setProperty("col_name", col_name)
            cb.toggled.connect(self._on_col_visibility_changed)
            row1.addWidget(cb)
            self._hid_col_checks[col_name] = cb

        _sep3 = QFrame(); _sep3.setFrameShape(QFrame.Shape.VLine)
        _sep3.setFrameShadow(QFrame.Shadow.Sunken)
        row1.addSpacing(4); row1.addWidget(_sep3); row1.addSpacing(4)

        # 忽略隐藏节点
        self._skip_hidden_chk = QCheckBox("忽略隐藏的节点")
        self._skip_hidden_chk.setChecked(self._skip_hidden)
        self._skip_hidden_chk.setToolTip(
            "勾选时：零件处于隐藏状态则跳过；产品/部件处于隐藏状态则连同其子孙一并跳过"
        )
        self._skip_hidden_chk.toggled.connect(self._on_skip_hidden_changed)
        row1.addWidget(self._skip_hidden_chk)

        row1.addStretch()
        opts_main.addLayout(row1)

        # ── 第二行：重量单位 ｜ 长度单位 ｜ 惯量单位（4选1）｜ 汇总BOM排序列 ──
        row2 = QHBoxLayout()
        row2.setSpacing(6)

        # 重量单位
        self._mass_unit_group = QButtonGroup(self)
        self._radio_mass_g  = QRadioButton("g")
        self._radio_mass_kg = QRadioButton("kg")
        self._radio_mass_g.setMinimumHeight(24)
        self._radio_mass_g.setChecked(self._mass_unit == "g")
        self._radio_mass_kg.setChecked(self._mass_unit == "kg")
        self._mass_unit_group.addButton(self._radio_mass_g)
        self._mass_unit_group.addButton(self._radio_mass_kg)
        self._radio_mass_g.toggled.connect(self._on_unit_changed)
        self._radio_mass_kg.toggled.connect(self._on_unit_changed)
        row2.addWidget(QLabel("重量:"))
        row2.addWidget(self._radio_mass_g)
        row2.addWidget(self._radio_mass_kg)

        _sep3 = QFrame(); _sep3.setFrameShape(QFrame.Shape.VLine)
        _sep3.setFrameShadow(QFrame.Shadow.Sunken)
        row2.addSpacing(4); row2.addWidget(_sep3); row2.addSpacing(4)

        # 长度单位
        self._cog_unit_group = QButtonGroup(self)
        self._radio_cog_mm = QRadioButton("mm")
        self._radio_cog_m  = QRadioButton("m")
        self._radio_cog_mm.setChecked(self._cog_unit == "mm")
        self._radio_cog_m.setChecked(self._cog_unit == "m")
        self._cog_unit_group.addButton(self._radio_cog_mm)
        self._cog_unit_group.addButton(self._radio_cog_m)
        self._radio_cog_mm.toggled.connect(self._on_unit_changed)
        self._radio_cog_m.toggled.connect(self._on_unit_changed)
        row2.addWidget(QLabel("长度:"))
        row2.addWidget(self._radio_cog_mm)
        row2.addWidget(self._radio_cog_m)

        _sep4 = QFrame(); _sep4.setFrameShape(QFrame.Shape.VLine)
        _sep4.setFrameShadow(QFrame.Shadow.Sunken)
        row2.addSpacing(4); row2.addWidget(_sep4); row2.addSpacing(4)

        # 惯量单位（4 选 1，独立）
        _IU = ("g\u00b7mm\u00b2", "g\u00b7m\u00b2", "kg\u00b7mm\u00b2", "kg\u00b7m\u00b2")
        self._inertia_unit_group = QButtonGroup(self)
        self._radio_inertia: dict[str, QRadioButton] = {}
        row2.addWidget(QLabel("惯量:"))
        for iu in _IU:
            rb = QRadioButton(iu)
            rb.setChecked(self._inertia_unit == iu)
            self._inertia_unit_group.addButton(rb)
            rb.toggled.connect(self._on_inertia_unit_changed)
            row2.addWidget(rb)
            self._radio_inertia[iu] = rb

        _sep5 = QFrame(); _sep5.setFrameShape(QFrame.Shape.VLine)
        _sep5.setFrameShadow(QFrame.Shadow.Sunken)
        row2.addSpacing(4); row2.addWidget(_sep5); row2.addSpacing(4)

        # 汇总BOM专用选项（排序列）
        self._summary_opts_widget = QWidget()
        summary_opts_layout = QHBoxLayout(self._summary_opts_widget)
        summary_opts_layout.setContentsMargins(0, 0, 0, 0)
        summary_opts_layout.setSpacing(6)
        summary_opts_layout.addWidget(QLabel("排序列:"))
        self._sort_col_combo = QComboBox()
        self._sort_col_combo.addItem("（不排序）", "")
        for col in _SUMMARY_SORT_COLUMNS:
            self._sort_col_combo.addItem(MASS_PROPS_COLUMN_DISPLAY_NAMES.get(col, col), col)
        saved_sort_idx = self._sort_col_combo.findData(self._summary_sort_column)
        if saved_sort_idx >= 0:
            self._sort_col_combo.setCurrentIndex(saved_sort_idx)
        self._sort_col_combo.currentIndexChanged.connect(self._on_sort_col_changed)
        self._sort_col_combo.setMaximumHeight(24)
        summary_opts_layout.addWidget(self._sort_col_combo)
        self._summary_opts_widget.setVisible(self._summarize)
        row2.addWidget(self._summary_opts_widget)

        row2.addStretch()
        opts_main.addLayout(row2)

        layout.addWidget(opts_group)

        # ── BOM说明标签 ─────────────────────────────────────────────────────
        self._bom_desc_lbl = QLabel(self._bom_desc_text())
        self._bom_desc_lbl.setWordWrap(True)
        self._bom_desc_lbl.setStyleSheet(
            "QLabel { background-color: #EEF4FC; border: 1px solid #B8D0F0;"
            " border-radius: 4px; padding: 4px 8px; color: #2B4C7E; font-size: 11px; }"
        )
        layout.addWidget(self._bom_desc_lbl)

        # ── 树形表格 ────────────────────────────────────────────────────────
        self._table = _BomTreeWidget()
        self._table.setColumnCount(len(self._columns))
        self._table.setHeaderLabels(self._display_headers())
        hdr = self._table.header()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(False)
        hdr.setFixedHeight(28)
        self._table.setUniformRowHeights(True)
        self._table.setRootIsDecorated(True)
        self._table.setSortingEnabled(False)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setAlternatingRowColors(True)
        self._table.setIndentation(16)
        self._table.setStyleSheet("QTreeWidget::item { min-height: 24px; }")
        self._table.setItemDelegate(_MassPropsDelegate(lambda: self._columns, self._table))
        self._table.itemChanged.connect(self._on_item_changed)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_tree_context_menu)
        hdr.sectionResized.connect(self._on_section_resized)
        layout.addWidget(self._table, 1)

        # ── 汇总面板 ────────────────────────────────────────────────────────
        summary_group  = QGroupBox("汇总结果（基于根产品坐标系）")
        summary_layout = QGridLayout(summary_group)
        summary_layout.setSpacing(8)
        summary_layout.setContentsMargins(10, 8, 10, 8)

        def _lbl(text: str) -> QLabel:
            lb = QLabel(text)
            lb.setStyleSheet("font-weight: bold;")
            return lb

        def _val_lbl() -> QLabel:
            lb = QLabel("—")
            lb.setMinimumWidth(120)
            return lb

        # 第0行：总重量（独占）
        summary_layout.addWidget(_lbl("总重量："), 0, 0)
        self._lbl_weight = _val_lbl()
        summary_layout.addWidget(self._lbl_weight, 0, 1)

        # 第1行：总重心 X / Y / Z
        summary_layout.addWidget(_lbl("总重心 X："), 1, 0)
        self._lbl_cx = _val_lbl()
        summary_layout.addWidget(self._lbl_cx, 1, 1)

        summary_layout.addWidget(_lbl("总重心 Y："), 1, 2)
        self._lbl_cy = _val_lbl()
        summary_layout.addWidget(self._lbl_cy, 1, 3)

        summary_layout.addWidget(_lbl("总重心 Z："), 1, 4)
        self._lbl_cz = _val_lbl()
        summary_layout.addWidget(self._lbl_cz, 1, 5)

        # 第2行：Ixx / Iyy / Izz
        for c_i, (text, attr) in enumerate([("Ixx:", "lbl_ixx"), ("Iyy:", "lbl_iyy"), ("Izz:", "lbl_izz")]):
            summary_layout.addWidget(_lbl(text), 2, c_i * 2)
            lbl = _val_lbl()
            setattr(self, f"_{attr}", lbl)
            summary_layout.addWidget(lbl, 2, c_i * 2 + 1)

        # 第3行：Ixy / Ixz / Iyz
        for c_i, (text, attr) in enumerate([("Ixy:", "lbl_ixy"), ("Ixz:", "lbl_ixz"), ("Iyz:", "lbl_iyz")]):
            summary_layout.addWidget(_lbl(text), 3, c_i * 2)
            lbl = _val_lbl()
            setattr(self, f"_{attr}", lbl)
            summary_layout.addWidget(lbl, 3, c_i * 2 + 1)

        layout.addWidget(summary_group)

        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(line)

        # ── 底部按钮行 ──────────────────────────────────────────────────────
        btn_row = QHBoxLayout()

        autofit_btn = QPushButton("自适应列宽")
        autofit_btn.clicked.connect(self._autofit_columns)
        btn_row.addWidget(autofit_btn)

        expand_btn = QPushButton("全部展开")
        expand_btn.clicked.connect(self._table.expandAll)
        btn_row.addWidget(expand_btn)

        collapse_btn = QPushButton("全部折叠")
        collapse_btn.clicked.connect(self._table.collapseAll)
        btn_row.addWidget(collapse_btn)

        btn_row.addStretch()

        self._calc_btn = QPushButton("计算")
        self._calc_btn.setToolTip("汇总装配体总质量特性（质量 / 重心 / 转动惯量）")
        self._calc_btn.setEnabled(False)
        self._calc_btn.clicked.connect(self._calculate)
        btn_row.addWidget(self._calc_btn)

        self._save_json_btn = QPushButton("保存数据…")
        self._save_json_btn.setToolTip("将当前行数据保存为数据文件，可在不打开CATIA的情况下重新载入")
        self._save_json_btn.setEnabled(False)
        self._save_json_btn.clicked.connect(self._save_data_to_json)
        btn_row.addWidget(self._save_json_btn)

        self._export_btn = QPushButton("导出表格")
        self._export_btn.setToolTip("将当前表格（含汇总行）导出为 Excel（.xlsx）文件")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_table)
        btn_row.addWidget(self._export_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.reject)
        btn_row.addWidget(close_btn)

        layout.addLayout(btn_row)

    # ── BOM 说明文字 ───────────────────────────────────────────────────────

    def _bom_desc_text(self) -> str:
        """返回当前 BOM 模式对应的说明文字。"""
        if self._summarize:
            return (
                "【汇总BOM】按零件编号合并，仅列出零件（不含产品和部件）。"
                "Weight / CogX / CogY / CogZ / Ixx–Iyz "
                "在零件自身坐标系下显示，与装配位置无关。"
                "底部「汇总结果」在根产品坐标系下计算。"
            )
        return (
            "【层级BOM】展示零件节点和产品/部件节点。"
            "Weight / CogX / CogY / CogZ / Ixx–Iyz "
            "在根产品坐标系下显示，与零件的装配位置有关。"
            "底部「汇总结果」在根产品坐标系下计算。"
        )

    # ── 文件/活动文档切换 ──────────────────────────────────────────────────

    def _toggle_file_row(self, use_active: bool) -> None:
        self._file_edit.setEnabled(not use_active)
        self._file_browse_btn.setEnabled(not use_active)

    def _browse_file(self) -> None:
        file, _ = QFileDialog.getOpenFileName(
            self, "选择CATProduct文件",
            self._last_browse_dir,
            "*.CATProduct (*.CATProduct);;All Files (*)",
        )
        if file:
            self._file_edit.setText(file)
            self._last_browse_dir = str(Path(file).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)

    # ── 显示选项响应 ───────────────────────────────────────────────────────

    def _on_bom_type_changed(self, checked: bool) -> None:
        self._summarize = self._radio_summ.isChecked()
        self._settings.setValue("summarize", self._summarize)
        self._summary_opts_widget.setVisible(self._summarize)
        self._bom_desc_lbl.setText(self._bom_desc_text())
        self._rebuild_columns_and_table()

    def _on_read_mode_changed(self, checked: bool) -> None:
        if self._radio_read_first.isChecked():
            self._read_mode = "first"
        elif self._radio_read_last.isChecked():
            self._read_mode = "last"
        else:
            self._read_mode = "all"
        self._settings.setValue("read_mode", self._read_mode)

    def _on_skip_hidden_changed(self, checked: bool) -> None:
        self._skip_hidden = self._skip_hidden_chk.isChecked()
        self._settings.setValue("skip_hidden", self._skip_hidden)

    def _on_unit_changed(self, checked: bool) -> None:
        self._mass_unit = "g" if self._radio_mass_g.isChecked() else "kg"
        self._cog_unit  = "mm" if self._radio_cog_mm.isChecked() else "m"
        self._unit_factor, _, self._cog_unit_factor = (
            self._calc_unit_factors(self._mass_unit, self._cog_unit)
        )
        self._settings.setValue("mass_unit", self._mass_unit)
        self._settings.setValue("cog_unit", self._cog_unit)
        if self._rows:
            self._refresh_unit_display()

    def _on_inertia_unit_changed(self, checked: bool) -> None:
        for iu, rb in self._radio_inertia.items():
            if rb.isChecked():
                self._inertia_unit = iu
                break
        self._inertia_unit_factor = self._calc_inertia_factor(self._inertia_unit)
        self._settings.setValue("inertia_unit", self._inertia_unit)
        if self._rows:
            self._refresh_unit_display()

    def _on_col_visibility_changed(self, checked: bool) -> None:
        for col_name, cb in self._hid_col_checks.items():
            if cb.isChecked():
                self._visible_hideable_cols.add(col_name)
            else:
                self._visible_hideable_cols.discard(col_name)
        self._settings.setValue("visible_hideable_cols",
                                list(self._visible_hideable_cols))
        self._rebuild_columns_and_table()

    def _on_sort_col_changed(self, _index: int) -> None:
        col = self._sort_col_combo.currentData()
        self._summary_sort_column = col or ""
        self._settings.setValue("summary_sort_column", self._summary_sort_column)
        if self._summarize and self._rows:
            self._populate_table()

    def _rebuild_columns_and_table(self) -> None:
        """重建列列表并重新填充表格（保留列宽）。"""
        if self._rows:
            # 重建前保存各列宽度
            for col_idx, col_name in enumerate(self._columns):
                self._col_widths[col_name] = self._table.columnWidth(col_idx)
        self._columns = self._build_columns()
        self._populate_table()
        # 恢复各列宽度
        for col_idx, col_name in enumerate(self._columns):
            if col_name in self._col_widths:
                self._table.setColumnWidth(col_idx, self._col_widths[col_name])

    def _refresh_unit_display(self) -> None:
        """仅更新列标题和重量/惯量单元格的显示值（单位切换时调用，避免全量重建）。"""
        # 更新列标题
        self._table.setHeaderLabels(self._display_headers())

        mass_col_indices: list[tuple[str, int]] = []
        for col_name in _UNIT_SENSITIVE_COLUMNS:
            if col_name in self._columns:
                mass_col_indices.append((col_name, self._columns.index(col_name)))

        if not mass_col_indices:
            return

        self._is_updating = True
        display_rows = self._get_display_rows()
        for di, row_data in enumerate(display_rows):
            if di >= len(self._item_by_row):
                break
            item = self._item_by_row[di]
            if not any(row_data.get(c) is not None for c in _UNIT_SENSITIVE_COLUMNS):
                continue
            for col_name, col_idx in mass_col_indices:
                raw = row_data.get(col_name)
                if raw is not None:
                    if col_name == "Weight":
                        item.setText(col_idx, self._fmt_mass_val(raw))
                    elif col_name in _INERTIA_IDX:
                        item.setText(col_idx, self._fmt_inertia_val(raw))
                    else:
                        item.setText(col_idx, self._fmt_cog_val(raw))
        self._is_updating = False

        # 若已有汇总结果，更新底部汇总标签
        if self._rollup_result:
            self._update_summary_labels(self._rollup_result)

    # ── 加载数据 ───────────────────────────────────────────────────────────

    def _load_data(self) -> None:
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请先选择一个CATProduct文件。")
                return
            if not Path(file_path).exists():
                QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{file_path}")
                return

        self._load_btn.setEnabled(False)
        self._load_btn.setText("加载中…")
        QApplication.processEvents()

        progress = QProgressDialog("正在加载产品树，请稍候…", None, 0, 0, self)
        progress.setWindowTitle("加载质量特性")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)
        progress.setValue(0)

        def _on_row_collected(count: int) -> None:
            progress.setLabelText(f"正在加载产品树，请稍候… 已读取 {count} 个节点")
            progress.repaint()
            QApplication.processEvents()

        try:
            rows = collect_mass_props_rows(file_path, progress_callback=_on_row_collected,
                                           read_mode=self._read_mode,
                                           skip_hidden=self._skip_hidden)
        except Exception as e:
            progress.close()
            logger.error(f"加载质量特性失败: {e}")
            QMessageBox.critical(
                self, "加载失败",
                f"加载产品树时出错：\n{e}\n\n请确保CATIA已启动。",
            )
            self._load_btn.setEnabled(True)
            self._load_btn.setText("加载")
            return
        finally:
            progress.close()

        self._load_btn.setEnabled(True)
        self._load_btn.setText("重新加载")

        self._apply_loaded_rows(rows)

    def _apply_loaded_rows(self, rows: list[dict]) -> None:
        """将已就绪的行列表应用到对话框：重建表格、调整列宽、启用按钮并计算。

        由 :meth:`_load_data` 和 :meth:`_load_data_from_json` 共用。
        """
        # 重新填充前保存列宽
        if self._loaded:
            for col_idx, col_name in enumerate(self._columns):
                self._col_widths[col_name] = self._table.columnWidth(col_idx)

        self._rows = rows
        self._rollup_result = None
        self._clear_summary_labels()
        self._columns = self._build_columns()
        self._populate_table()

        if not self._loaded:
            # 首次加载时自适应列宽
            for _c, col_name in enumerate(self._columns):
                if col_name == "#":
                    self._table.setColumnWidth(_c, 40)
                    self._col_widths[col_name] = 40
                else:
                    self._table.resizeColumnToContents(_c)
                    self._col_widths[col_name] = self._table.columnWidth(_c)
            self._loaded = True
        else:
            for col_idx, col_name in enumerate(self._columns):
                if col_name in self._col_widths:
                    self._table.setColumnWidth(col_idx, self._col_widths[col_name])

        self._calc_btn.setEnabled(True)
        self._export_btn.setEnabled(True)
        self._save_json_btn.setEnabled(True)

        failed_count = sum(1 for r in rows if r.get("_meas_failed") and r.get("Type") == "零件")
        if failed_count:
            _read_mode_desc = {
                "first": "「惯量包络体.1」",
                "last":  f"编号最大的「惯量包络体.N」（N ≤ {MAX_INERTIA_INDEX}）",
                "all":   f"「惯量包络体.1」至「惯量包络体.{MAX_INERTIA_INDEX}」",
            }.get(self._read_mode, "惯量包络体")
            QMessageBox.information(
                self, "部分零件测量失败",
                f"有 {failed_count} 个零件节点无法完成质量特性测量（显示橙色背景）。\n\n"
                "可能原因：\n"
                "  • 零件文档未加载到CATIA会话中\n"
                f"  • 当前读取模式要求的 {_read_mode_desc} 保持测量不存在\n"
                "  • 测量是在产品环境下建立的（使用产品坐标系，不会被读取）\n"
                "  • 需单独打开零件文件，在SPA中建立惯量保持测量\n\n"
                "未能测量的零件不参与最终汇总计算。",
            )


        # 加载完成后自动计算汇总结果
        self._calculate()

    def _save_data_to_json(self) -> None:
        """将当前行数据保存为压缩二进制数据文件（不包含 _root_mp，可重新计算）。"""
        if not self._rows:
            return

        # ── 默认文件名：根产品零件编号 + "_惯量汇总" ───────────────────────
        root_pn = str(self._rows[0].get("Part Number", "")).strip()
        default_name = f"{root_pn}_惯量汇总" if root_pn else "惯量汇总"

        # ── 默认目录：上次浏览目录 → 根产品文件所在目录 → 空 ──────────────
        if self._last_browse_dir and Path(self._last_browse_dir).is_dir():
            default_dir = self._last_browse_dir
        else:
            root_fp = str(self._rows[0].get("_filepath", "")).strip()
            default_dir = str(Path(root_fp).parent) if root_fp else ""

        default_path = str(Path(default_dir) / default_name) if default_dir else default_name

        dest, _ = QFileDialog.getSaveFileName(
            self, "保存质量特性数据", default_path, "质量特性数据文件 (*.mpd)"
        )
        if not dest:
            return
        if not dest.lower().endswith(".mpd"):
            dest += ".mpd"
        try:
            save_rows(self._rows, dest)
            self._last_browse_dir = str(Path(dest).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)
        except Exception as e:
            logger.error(f"保存质量特性数据失败: {e}")
            QMessageBox.critical(self, "保存失败", f"保存数据时出错：\n{e}")

    def _load_data_from_json(self) -> None:
        """从压缩二进制数据文件载入行数据（无需 CATIA，_root_mp 由后处理重建）。"""
        src, _ = QFileDialog.getOpenFileName(
            self, "载入质量特性数据", "", "质量特性数据文件 (*.mpd)"
        )
        if not src:
            return
        if not Path(src).exists():
            QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{src}")
            return
        try:
            rows = load_rows(src)
        except Exception as e:
            logger.error(f"载入质量特性数据失败: {e}")
            QMessageBox.critical(self, "载入失败", f"载入数据时出错：\n{e}")
            return
        self._apply_loaded_rows(rows)

    # ── 构建显示行 ─────────────────────────────────────────────────────────

    def _get_display_rows(self) -> list[dict]:
        """返回当前模式下应显示的行列表。"""
        if self._summarize:
            return self._build_summary_rows()
        # 层级BOM：展示全部节点（零件、产品、部件），使用根产品坐标系下的值。
        # 每行附加 _rows_idx，指向 self._rows 中的原始索引，
        # 以确保 _make_item / _on_item_changed 能正确回写数据。
        # 对零件行，将 _root_mp 中的根坐标系 COG / 惯量值覆盖显示字段；
        # 产品/部件行的显示字段已由 _post_process_rows() 写入根坐标系汇总值。
        result = []
        for i, row in enumerate(self._rows):
            r = dict(row)
            r["_rows_idx"] = i
            if r.get("Type") == "零件":
                rmp = r.get("_root_mp")
                if rmp:
                    cog = rmp.get("cog", [None, None, None])
                    r["CogX"] = cog[0]
                    r["CogY"] = cog[1]
                    r["CogZ"] = cog[2]
                    I = rmp.get("inertia")
                    if I:
                        r["Ixx"] = I[0][0]
                        r["Iyy"] = I[1][1]
                        r["Izz"] = I[2][2]
                        r["Ixy"] = I[0][1]
                        r["Ixz"] = I[0][2]
                        r["Iyz"] = I[1][2]
            result.append(r)
        return result

    def _build_summary_rows(self) -> list[dict]:
        """汇总模式：将相同零件编号的行合并，增加 Quantity 字段。

        每个唯一 PN 保留第一次出现的行数据（含 _rows 中的索引），
        Quantity = 该 PN 在 _rows 中出现的实例数量。
        """
        seen_pn: dict[str, dict] = {}    # pn → 首次出现的规范行副本
        qty: dict[str, int] = {}
        order: list[str] = []

        for i, row in enumerate(self._rows):
            pn = str(row.get("Part Number", ""))
            if not pn:
                pn = str(row.get("Filename", "")) or "(未分组)"
            if pn not in seen_pn:
                r = dict(row)
                r["_rows_idx"] = i   # 映射回 _rows 的规范索引
                seen_pn[pn] = r
                qty[pn] = 1
                order.append(pn)
            else:
                qty[pn] += 1

        result = []
        for pn in order:
            r = dict(seen_pn[pn])
            r["Quantity"] = qty[pn]
            result.append(r)

        # 仅保留零件行（汇总BOM不显示产品和部件）
        result = [r for r in result if r.get("Type") == "零件"]

        # 按排序列排序
        if self._summary_sort_column:
            col = self._summary_sort_column
            result.sort(key=lambda r: str(r.get(col, "") or ""))

        return result

    # ── 填充表格 ───────────────────────────────────────────────────────────

    def _populate_table(self) -> None:
        self._is_updating = True
        self._table.blockSignals(True)

        self._table.clear()
        self._table.setColumnCount(len(self._columns))
        self._table.setHeaderLabels(self._display_headers())
        self._table.setRootIsDecorated(not self._summarize)
        self._item_by_row = []
        self._pn_to_items.clear()

        display_rows = self._get_display_rows()

        if self._summarize:
            self._populate_flat(display_rows)
        else:
            self._populate_tree(display_rows)

        self._table.expandAll()
        self._table.blockSignals(False)
        self._is_updating = False

    def _make_item(self, row_idx: int, row_data: dict) -> QTreeWidgetItem:
        """构建并填充一行的 QTreeWidgetItem。

        row_idx: 对应的 self._rows 索引（汇总模式用 _rows_idx 字段）。
        """
        item = QTreeWidgetItem()
        item.setData(0, _ROW_IDX_ROLE, row_idx)

        pn          = str(row_data.get("Part Number", ""))
        not_found   = bool(row_data.get("_not_found"))
        no_file     = bool(row_data.get("_no_file"))
        unreadable  = bool(row_data.get("_unreadable"))
        meas_failed = bool(row_data.get("_meas_failed"))
        node_type   = str(row_data.get("Type", ""))
        row_locked  = unreadable or not_found or meas_failed

        if pn:
            self._pn_to_items.setdefault(pn, []).append(item)

        seq_no = str(len(self._item_by_row) + 1)

        for col_idx, col_name in enumerate(self._columns):
            if col_name == "#":
                item.setText(col_idx, seq_no)
            elif col_name == "Level":
                item.setText(col_idx, str(row_data.get("Level", 0)))
            elif col_name == "Filename":
                fp = str(row_data.get("_filepath", ""))
                fn = str(row_data.get("Filename", ""))
                if no_file:
                    value = FILENAME_UNSAVED
                else:
                    value = Path(fp).name if fp else fn
                item.setText(col_idx, value)
                if no_file:
                    pass  # tooltip 由下方 no_file 块统一设置
                elif fp:
                    item.setToolTip(col_idx, fp)
            elif col_name == "Quantity":
                item.setText(col_idx, str(row_data.get("Quantity", 1)))
            elif col_name == "Weight":
                raw = row_data.get("Weight")
                if raw is None:
                    item.setText(col_idx, "—" if node_type == "零件" else "")
                else:
                    item.setText(col_idx, self._fmt_mass_val(raw))
            elif col_name in _INERTIA_IDX or col_name in ("CogX", "CogY", "CogZ"):
                raw = row_data.get(col_name)
                if raw is None:
                    item.setText(col_idx, "—" if node_type == "零件" else "")
                else:
                    if col_name in _INERTIA_IDX:
                        item.setText(col_idx, self._fmt_inertia_val(raw))
                    else:
                        item.setText(col_idx, self._fmt_cog_val(raw))
            else:
                item.setText(col_idx, str(row_data.get(col_name, "")))

        # 可编辑性：仅未锁定零件行的 Weight 列可编辑
        if node_type == "零件" and not row_locked:
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            item.setData(0, _ITEM_LOCKED_ROLE, False)
        else:
            item.setData(0, _ITEM_LOCKED_ROLE, True)

        # 行背景色设置
        if row_locked:
            grey = QColor(160, 160, 160)
            if not_found:
                bg  = QColor(255, 205, 205)
                tip = "该零件/装配体的文件未被CATIA检索到，行内容不可编辑。"
            elif meas_failed:
                bg  = QColor(255, 210, 160)
                tip = "该零件的质量特性测量失败，行内容不可编辑。"
            else:
                bg  = QColor(245, 245, 245)
                tip = "该零件/装配体处于轻量化模式，无法读取属性。"
            for ci in range(len(self._columns)):
                item.setForeground(ci, grey)
                item.setBackground(ci, bg)
                item.setToolTip(ci, tip)
        elif no_file:
            bg_unsaved = QColor(255, 245, 180)
            no_file_tip = "该零件尚未保存到磁盘，质量特性数据可能不完整。"
            for ci in range(len(self._columns)):
                item.setBackground(ci, bg_unsaved)
                item.setToolTip(ci, no_file_tip)
        elif node_type in ("产品", "部件"):
            bg = QColor(240, 242, 245)
            for ci in range(len(self._columns)):
                item.setBackground(ci, bg)

        self._item_by_row.append(item)
        return item

    def _populate_flat(self, display_rows: list[dict]) -> None:
        """汇总BOM模式：所有行为顶级项（无树形层级）。"""
        for di, row_data in enumerate(display_rows):
            rows_idx = row_data.get("_rows_idx", di)
            item = self._make_item(rows_idx, row_data)
            self._table.addTopLevelItem(item)

    def _populate_tree(self, display_rows: list[dict]) -> None:
        """层级BOM模式：按 Level 构建树形结构。"""
        parent_stack: list[tuple[int, QTreeWidgetItem | None]] = [(-1, None)]

        for di, row_data in enumerate(display_rows):
            level = int(row_data.get("Level", 0))
            # 使用 _rows_idx（若存在）映射回 self._rows，保持与 _populate_flat 一致
            rows_idx = row_data.get("_rows_idx", di)

            while len(parent_stack) > 1 and parent_stack[-1][0] >= level:
                parent_stack.pop()

            parent_item = parent_stack[-1][1]
            item = self._make_item(rows_idx, row_data)

            if parent_item is None:
                self._table.addTopLevelItem(item)
            else:
                parent_item.addChild(item)

            parent_stack.append((level, item))

    # ── 单元格编辑 ─────────────────────────────────────────────────────────

    def _on_item_changed(self, item: QTreeWidgetItem, col_idx: int) -> None:
        if self._is_updating:
            return
        row_idx = item.data(0, _ROW_IDX_ROLE)
        if row_idx is None:
            return

        col_name = self._columns[col_idx]
        if col_name != "Weight":
            return

        if item.data(0, _ITEM_LOCKED_ROLE):
            return

        row_data = self._rows[row_idx]
        if row_data.get("Type") != "零件":
            return

        new_text = item.text(col_idx).strip()
        try:
            # 输入值为当前显示单位；除以 _unit_factor 还原到内部单位（kg）
            new_display_val = float(new_text)
            new_weight_stored = new_display_val / self._unit_factor
        except (ValueError, TypeError):
            self._is_updating = True
            item.setText(col_idx, self._fmt_mass_val(row_data.get("Weight")))
            self._is_updating = False
            return

        if new_weight_stored < 0.0:
            QMessageBox.warning(
                self, "重量不合法",
                "重量不能为负数，请输入大于或等于 0 的值。",
            )
            self._is_updating = True
            item.setText(col_idx, self._fmt_mass_val(row_data.get("Weight")))
            self._is_updating = False
            return

        pn = str(row_data.get("Part Number", ""))

        # ── 更新 _rows 中所有相同 PN 的实例 ────────────────────────────────
        #
        # 同一零件的所有实例共享同一个 _mass_props dict（来自 _mass_cache）。
        # 若在循环内对每个实例各乘一次 scale，则第 n 个实例的惯量会被放大 scale^n 倍。
        # 正确做法：先从第一个匹配行计算 scale，对共享 dict 仅缩放一次，
        # 再遍历各实例分别用各自的 _placement 矩阵重新旋转到根坐标系。
        #
        # Step 1：从第一个匹配行取 scale 和共享的 _mass_props。
        scale: float = 1.0
        mp_shared: dict | None = None
        for r in self._rows:
            if str(r.get("Part Number", "")) == pn and r.get("Type") == "零件":
                try:
                    old_w_f_0 = float(r.get("Weight") or 0.0)
                except (ValueError, TypeError):
                    old_w_f_0 = 0.0
                if old_w_f_0 > 0.0:
                    scale = new_weight_stored / old_w_f_0
                mp_shared = r.get("_mass_props")
                break

        # Step 2：对共享 _mass_props 的惯量只缩放一次。
        if mp_shared is not None:
            mp_shared["weight"] = new_weight_stored
            if scale != 1.0:
                orig_i = mp_shared.get("inertia", [[0.0] * 3 for _ in range(3)])
                mp_shared["inertia"] = [[orig_i[ir][ic] * scale for ic in range(3)]
                                        for ir in range(3)]

        # Step 3：遍历所有实例，更新 Weight / 行级显示字段 / _root_mp。
        for r in self._rows:
            if str(r.get("Part Number", "")) != pn or r.get("Type") != "零件":
                continue
            r["Weight"] = new_weight_stored
            mp = r.get("_mass_props")
            if mp:
                # mp["inertia"] 已在 Step 2 缩放完毕；
                # 仅需用本实例自己的 _placement 重新旋转到根坐标系。
                if scale != 1.0:
                    # 更新行级惯量显示字段（零件自身坐标系）
                    I_local_new = mp.get("inertia", [[0.0] * 3 for _ in range(3)])
                    for ic_name, (ir2, ic2) in _INERTIA_IDX.items():
                        r[ic_name] = I_local_new[ir2][ic2]
                    # 同步更新 _root_mp 中的惯量（缩放后重新旋转到根坐标系）
                    I_root = _row_inertia_to_root(r)
                    rmp = r.get("_root_mp")
                    if rmp is not None:
                        rmp["inertia"] = I_root
                        rmp["weight"]  = new_weight_stored
                else:
                    rmp = r.get("_root_mp")
                    if rmp is not None:
                        rmp["weight"] = new_weight_stored
            else:
                # 无 _mass_props（各实例行独立存储惯量值），逐行缩放显示字段。
                if scale != 1.0:
                    for ic_name in _INERTIA_IDX:
                        cur = r.get(ic_name)
                        if cur is not None:
                            r[ic_name] = float(cur) * scale
                rmp = r.get("_root_mp")
                if rmp is not None:
                    rmp["weight"] = new_weight_stored

        # ── 更新可见树节点中同 PN 的所有行 ────────────────────────────────
        self._is_updating = True
        w_idx = self._columns.index("Weight") if "Weight" in self._columns else -1

        for vis_item in self._pn_to_items.get(pn, []):
            vis_row_idx = vis_item.data(0, _ROW_IDX_ROLE)
            if vis_row_idx is None:
                continue
            vis_row = self._rows[vis_row_idx]
            if vis_row.get("Type") != "零件":
                continue
            if w_idx >= 0:
                vis_item.setText(w_idx, self._fmt_mass_val(vis_row.get("Weight")))
            for ic_name, (ir, ic) in _INERTIA_IDX.items():
                if ic_name in self._columns:
                    ic_idx = self._columns.index(ic_name)
                    if self._summarize:
                        # 汇总BOM：显示零件自身坐标系值
                        raw_i = vis_row.get(ic_name)
                    else:
                        # 层级BOM：显示根产品坐标系值
                        rmp = vis_row.get("_root_mp")
                        raw_i = (
                            rmp["inertia"][ir][ic]
                            if rmp and rmp.get("inertia")
                            else vis_row.get(ic_name)
                        )
                    if raw_i is not None:
                        vis_item.setText(ic_idx, self._fmt_inertia_val(raw_i))

        self._is_updating = False
        self._rollup_result = None
        self._clear_summary_labels()

    # ── 计算 ───────────────────────────────────────────────────────────────

    def _calculate(self) -> None:
        if not self._rows:
            return
        # 先重新计算产品/部件行（使用更新后的 _root_mp），刷新表格
        recompute_product_rows(self._rows)
        self._refresh_product_items()
        try:
            result = rollup_mass_properties(self._rows)
        except Exception as e:
            logger.error(f"质量特性计算失败: {e}")
            QMessageBox.critical(self, "计算失败", f"计算总质量特性时出错：\n{e}")
            return
        self._rollup_result = result
        self._update_summary_labels(result)

    def _refresh_product_items(self) -> None:
        """刷新树形表格中所有产品/部件行的显示值（仅层级BOM模式有效）。

        在 _calculate() 调用 recompute_product_rows() 更新 self._rows 后，
        调用本方法将新的汇总值写回对应的 QTreeWidgetItem，以保持表格与数据同步。
        汇总BOM不含产品/部件行，故直接返回。
        """
        if self._summarize:
            return
        self._is_updating = True
        try:
            for item in self._item_by_row:
                row_idx = item.data(0, _ROW_IDX_ROLE)
                if row_idx is None:
                    continue
                row_data = self._rows[row_idx]
                if row_data.get("Type") not in ("产品", "部件"):
                    continue
                for col_idx, col_name in enumerate(self._columns):
                    if col_name == "Weight":
                        raw = row_data.get("Weight")
                        item.setText(col_idx, self._fmt_mass_val(raw) if raw is not None else "")
                    elif col_name in ("CogX", "CogY", "CogZ"):
                        raw = row_data.get(col_name)
                        item.setText(col_idx, self._fmt_cog_val(raw) if raw is not None else "")
                    elif col_name in _INERTIA_IDX:
                        raw = row_data.get(col_name)
                        item.setText(col_idx, self._fmt_inertia_val(raw) if raw is not None else "")
        finally:
            self._is_updating = False

    def _clear_summary_labels(self) -> None:
        for lbl in (self._lbl_weight, self._lbl_cx, self._lbl_cy, self._lbl_cz,
                    self._lbl_ixx, self._lbl_iyy, self._lbl_izz,
                    self._lbl_ixy, self._lbl_ixz, self._lbl_iyz):
            lbl.setText("—")

    def _update_summary_labels(self, result: dict) -> None:
        unit_lbl     = self._weight_unit_label()
        inertia_unit = self._inertia_unit_label()
        cog_unit     = self._cog_unit_label()
        w_val = result.get("total_weight", 0.0)
        self._lbl_weight.setText(f"{self._fmt_mass_val(w_val)} {unit_lbl}")
        cog = result.get("cog", [0.0, 0.0, 0.0])
        self._lbl_cx.setText(f"{self._fmt_cog_val(cog[0])} {cog_unit}")
        self._lbl_cy.setText(f"{self._fmt_cog_val(cog[1])} {cog_unit}")
        self._lbl_cz.setText(f"{self._fmt_cog_val(cog[2])} {cog_unit}")
        I = result.get("inertia", [[0.0] * 3 for _ in range(3)])
        self._lbl_ixx.setText(f"{self._fmt_inertia_val(I[0][0])} {inertia_unit}")
        self._lbl_iyy.setText(f"{self._fmt_inertia_val(I[1][1])} {inertia_unit}")
        self._lbl_izz.setText(f"{self._fmt_inertia_val(I[2][2])} {inertia_unit}")
        self._lbl_ixy.setText(f"{self._fmt_inertia_val(I[0][1])} {inertia_unit}")
        self._lbl_ixz.setText(f"{self._fmt_inertia_val(I[0][2])} {inertia_unit}")
        self._lbl_iyz.setText(f"{self._fmt_inertia_val(I[1][2])} {inertia_unit}")

    # ── 导出 ───────────────────────────────────────────────────────────────

    def _export_table(self) -> None:
        if not self._rows:
            QMessageBox.warning(self, "无数据", "请先加载产品树数据。")
            return

        # ── 默认文件名：根产品零件编号 + "_惯量汇总"（与"保存数据"对话框一致）──
        root_pn = str(self._rows[0].get("Part Number", "")).strip()
        default_stem = f"{root_pn}_惯量汇总" if root_pn else "惯量汇总"

        # ── 默认目录：上次浏览目录 → 根产品文件所在目录 → 空（与"保存数据"对话框一致）──
        if self._last_browse_dir and Path(self._last_browse_dir).is_dir():
            default_dir = self._last_browse_dir
        else:
            root_fp = str(self._rows[0].get("_filepath", "")).strip()
            default_dir = str(Path(root_fp).parent) if root_fp else ""

        default_path = str(Path(default_dir) / f"{default_stem}.xlsx") if default_dir else f"{default_stem}.xlsx"

        dest, _ = QFileDialog.getSaveFileName(
            self, "导出质量特性表格",
            default_path,
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv)",
        )
        if not dest:
            return

        dest_path = Path(dest)
        suffix = dest_path.suffix.lower()
        if suffix not in (".xlsx", ".csv"):
            dest_path = dest_path.with_suffix(".xlsx")
            suffix = ".xlsx"

        try:
            if suffix == ".csv":
                self._do_export_csv(dest_path)
            else:
                self._do_export(str(dest_path))
            self._last_browse_dir = str(dest_path.parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)
            QMessageBox.information(self, "导出成功", f"文件已保存到：\n{dest_path}")
        except Exception as e:
            logger.error(f"导出失败: {e}")
            QMessageBox.critical(self, "导出失败", f"导出时出错：\n{e}")

    def _do_export(self, dest: str) -> None:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from catia_copilot.utils import estimate_column_width

        # 导出列（排除内部序号列 "#"）
        export_cols = [c for c in self._columns if c != "#"]

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "质量特性"

        center      = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        thin_side   = Side(style="thin")
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side,
        )

        # 写入表头
        for ci, col_name in enumerate(export_cols, start=1):
            cell = ws.cell(row=1, column=ci, value=self._column_header(col_name))
            cell.font   = Font(bold=True)
            cell.fill   = header_fill
            cell.border = thin_border

        # 写入数据行
        display_rows = self._get_display_rows()
        for ri, row_data in enumerate(display_rows, start=2):
            for ci, col_name in enumerate(export_cols, start=1):
                raw = row_data.get(col_name)
                if raw is None:
                    value = ""
                elif col_name == "Weight":
                    try:
                        value = float(raw) * self._unit_factor
                    except (TypeError, ValueError):
                        value = ""
                elif col_name in _INERTIA_IDX:
                    try:
                        value = float(raw) * self._inertia_unit_factor
                    except (TypeError, ValueError):
                        value = ""
                elif col_name in ("CogX", "CogY", "CogZ"):
                    try:
                        value = float(raw) * self._cog_unit_factor
                    except (TypeError, ValueError):
                        value = ""
                else:
                    value = raw
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.border = thin_border
                if col_name == "Level":
                    cell.alignment = center

        # 汇总行（若已计算）
        if self._rollup_result:
            summary_row_idx = len(display_rows) + 2
            cog = self._rollup_result.get("cog", [0.0, 0.0, 0.0])
            I   = self._rollup_result.get("inertia", [[0.0] * 3 for _ in range(3)])
            w   = self._rollup_result.get("total_weight", 0.0)
            summary = {
                "Part Number":  "总计 (根产品)",
                "Weight":       w * self._unit_factor,
                "CogX":         cog[0] * self._cog_unit_factor,
                "CogY":         cog[1] * self._cog_unit_factor,
                "CogZ":         cog[2] * self._cog_unit_factor,
                "Ixx":          I[0][0] * self._inertia_unit_factor,
                "Iyy":          I[1][1] * self._inertia_unit_factor,
                "Izz":          I[2][2] * self._inertia_unit_factor,
                "Ixy":          I[0][1] * self._inertia_unit_factor,
                "Ixz":          I[0][2] * self._inertia_unit_factor,
                "Iyz":          I[1][2] * self._inertia_unit_factor,
            }
            summary_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
            for ci, col_name in enumerate(export_cols, start=1):
                val  = summary.get(col_name, "")
                cell = ws.cell(row=summary_row_idx, column=ci, value=val)
                cell.font   = Font(bold=True)
                cell.fill   = summary_fill
                cell.border = thin_border

        ws.freeze_panes = "A2"

        # 自适应列宽
        for ci, col_name in enumerate(export_cols, start=1):
            col_letter = ws.cell(row=1, column=ci).column_letter
            header     = self._column_header(col_name)
            max_width  = max(estimate_column_width(header), 8)
            for row_i in range(2, ws.max_row + 1):
                cv = ws.cell(row=row_i, column=ci).value
                if cv is not None:
                    max_width = max(max_width, estimate_column_width(str(cv)))
            ws.column_dimensions[col_letter].width = max_width + 2

        wb.save(dest)

    def _do_export_csv(self, dest: Path) -> None:
        """将当前表格数据（含汇总行）写入 UTF-8 with BOM 的 CSV 文件。"""
        import csv

        export_cols = [c for c in self._columns if c != "#"]
        display_rows = self._get_display_rows()

        def _cell_value(col_name: str, raw) -> str:
            if raw is None:
                return ""
            if col_name == "Weight":
                try:
                    return str(float(raw) * self._unit_factor)
                except (TypeError, ValueError):
                    return ""
            if col_name in _INERTIA_IDX:
                try:
                    return str(float(raw) * self._inertia_unit_factor)
                except (TypeError, ValueError):
                    return ""
            if col_name in ("CogX", "CogY", "CogZ"):
                try:
                    return str(float(raw) * self._cog_unit_factor)
                except (TypeError, ValueError):
                    return ""
            return str(raw)

        with open(dest, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([self._column_header(c) for c in export_cols])
            for row_data in display_rows:
                writer.writerow([
                    _cell_value(c, row_data.get(c)) for c in export_cols
                ])
            if self._rollup_result:
                cog = self._rollup_result.get("cog", [0.0, 0.0, 0.0])
                I   = self._rollup_result.get("inertia", [[0.0] * 3 for _ in range(3)])
                w   = self._rollup_result.get("total_weight", 0.0)
                summary = {
                    "Part Number":  "总计 (根产品)",
                    "Weight":       str(w * self._unit_factor),
                    "CogX":         str(cog[0] * self._cog_unit_factor),
                    "CogY":         str(cog[1] * self._cog_unit_factor),
                    "CogZ":         str(cog[2] * self._cog_unit_factor),
                    "Ixx":          str(I[0][0] * self._inertia_unit_factor),
                    "Iyy":          str(I[1][1] * self._inertia_unit_factor),
                    "Izz":          str(I[2][2] * self._inertia_unit_factor),
                    "Ixy":          str(I[0][1] * self._inertia_unit_factor),
                    "Ixz":          str(I[0][2] * self._inertia_unit_factor),
                    "Iyz":          str(I[1][2] * self._inertia_unit_factor),
                }
                writer.writerow([summary.get(c, "") for c in export_cols])
        logger.info(f"质量特性表格已导出 (csv) -> {dest}")

    def _autofit_columns(self) -> None:
        min_width = 60
        for col_idx in range(len(self._columns)):
            self._table.resizeColumnToContents(col_idx)
            if self._table.columnWidth(col_idx) < min_width:
                self._table.setColumnWidth(col_idx, min_width)
        for col_idx, col_name in enumerate(self._columns):
            self._col_widths[col_name] = self._table.columnWidth(col_idx)

    def _on_section_resized(self, logical_index: int, _old: int, new_size: int) -> None:
        if logical_index < len(self._columns):
            self._col_widths[self._columns[logical_index]] = new_size

    # ── 右键上下文菜单 ─────────────────────────────────────────────────────

    def _on_tree_context_menu(self, pos) -> None:
        """显示表格行的右键上下文菜单。"""
        item = self._table.itemAt(pos)
        if item is None:
            return
        row_idx = item.data(0, _ROW_IDX_ROLE)
        if row_idx is None:
            return

        row_data     = self._rows[row_idx]
        fp           = str(row_data.get("_filepath", ""))
        fp_path      = Path(fp) if fp else None
        is_component = row_data.get("Type") == "部件"
        is_part      = row_data.get("Type") == "零件"
        not_found    = bool(row_data.get("_not_found"))
        no_file      = bool(row_data.get("_no_file"))
        unreadable   = bool(row_data.get("_unreadable"))

        if not item.isSelected():
            self._table.clearSelection()
            item.setSelected(True)

        menu = QMenu(self)

        # ── 打开路径 ──────────────────────────────────────────────────────
        act_open_path = menu.addAction("打开路径")
        path_available = (
            bool(fp) and not no_file and fp_path is not None
            and (fp_path.exists() or fp_path.parent.exists())
        )
        act_open_path.setEnabled(path_available)

        # ── 复制路径 ──────────────────────────────────────────────────────
        act_copy_path = menu.addAction("复制路径")
        act_copy_path.setEnabled(bool(fp) and not no_file)

        # ── 在CATIA中打开 ─────────────────────────────────────────────────
        act_open_catia = menu.addAction("在CATIA中打开")
        catia_available = (
            not is_component and not not_found and not unreadable
            and fp_path is not None and fp_path.exists()
        )
        act_open_catia.setEnabled(catia_available)

        menu.addSeparator()

        # ── 重新读取质量特性 ───────────────────────────────────────────────
        act_reread = menu.addAction("重新读取质量特性")
        reread_available = (
            is_part and not not_found
            and fp_path is not None and fp_path.exists()
        )
        act_reread.setEnabled(reread_available)
        if reread_available:
            pn = str(row_data.get("Part Number", ""))
            act_reread.setToolTip(
                f"重新从 CATIA 读取零件「{pn}」的惯量包络体 Keep 测量参数，"
                "并同步更新所有相同零件编号的节点。"
            )

        action = menu.exec(self._table.viewport().mapToGlobal(pos))

        if action == act_open_path:
            self._open_path(fp)
        elif action == act_copy_path:
            QApplication.clipboard().setText(fp)
        elif action == act_open_catia:
            self._open_in_catia(fp)
        elif action == act_reread:
            self._reread_mass_props_for_row(row_idx)

    def _open_path(self, fp: str) -> None:
        """在 Windows 资源管理器中打开包含 *fp* 的文件夹，并高亮选中该文件。"""
        p = Path(fp).resolve()
        try:
            if p.exists():
                subprocess.Popen(f'explorer /select,"{p}"', shell=True)
            elif p.parent.exists():
                subprocess.Popen(f'explorer "{p.parent}"', shell=True)
        except Exception as exc:
            logger.warning(f"无法在资源管理器中打开路径: {exc}")

    def _open_in_catia(self, fp: str) -> None:
        """通过 ``documents.open`` 在CATIA中打开 *fp* 指向的文档。

        打开后，若 ``win32gui`` 可用，则将CATIA V5主窗口置于Windows前台。
        """
        try:
            from pycatia import catia as _pycatia  # noqa: PLC0415
            caa         = _pycatia()
            application = caa.application
            application.visible = True
            documents   = application.documents

            fp_resolved = Path(fp).resolve()
            documents.open(str(fp_resolved))

            try:
                import win32gui  # noqa: PLC0415
                import win32con  # noqa: PLC0415

                def _raise_catia_window(hwnd, _extra):
                    if not win32gui.IsWindowVisible(hwnd):
                        return
                    title = win32gui.GetWindowText(hwnd)
                    if title.startswith("CATIA V5"):
                        try:
                            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                            win32gui.SetForegroundWindow(hwnd)
                        except Exception:
                            pass
                        return False

                win32gui.EnumWindows(_raise_catia_window, None)
            except ImportError:
                pass
            except Exception:
                pass

        except Exception as e:
            QMessageBox.warning(self, "在CATIA中打开失败", f"无法在CATIA中打开文件：\n{e}")

    # ── 重新读取质量特性 ────────────────────────────────────────────────────

    def _reread_mass_props_for_row(self, row_idx: int) -> None:
        """重新从 CATIA 读取指定行（及所有同零件编号行）的质量特性。

        用于用户在 CATIA 中补充或更改惯量包络体后，无需重新加载整个产品树
        即可刷新单个零件的质量特性数据。若重新读取成功，同时恢复该零件所有
        节点的正常显示状态（清除橙色背景、恢复文字颜色、解除行锁定），并
        重新计算产品/部件节点的汇总质量特性。

        按零件编号检索：所有在 self._rows 中拥有相同 Part Number 且类型为
        "零件" 的行均会被同步更新，确保同一零件的多个实例数据一致。
        """
        row_data = self._rows[row_idx]
        fp = str(row_data.get("_filepath", ""))
        pn = str(row_data.get("Part Number", ""))
        if not fp:
            QMessageBox.warning(self, "无文件路径", "该零件没有有效的文件路径，无法重新读取。")
            return

        # 设置等待光标，提示用户正在操作
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            new_mp = remeasure_part_mass_props(fp, pn, self._read_mode)
        finally:
            QApplication.restoreOverrideCursor()

        if new_mp is None:
            _read_mode_desc = {
                "first": "「惯量包络体.1」",
                "last":  f"编号最大的「惯量包络体.N」（N ≤ {MAX_INERTIA_INDEX}）",
                "all":   f"「惯量包络体.1」至「惯量包络体.{MAX_INERTIA_INDEX}」",
            }.get(self._read_mode, "惯量包络体")
            QMessageBox.warning(
                self, "重新读取失败",
                f"未能从以下零件读取到有效的质量特性：\n{fp}\n\n"
                "可能原因：\n"
                "  • 该零件文档尚未在 CATIA 中打开\n"
                f"  • 当前读取模式要求的 {_read_mode_desc} 保持测量不存在\n"
                "  • 测量是在产品环境下建立的（使用产品坐标系，不会被读取）\n"
                "  • 需单独打开零件文件，在SPA中建立惯量保持测量",
            )
            return

        # ── 更新 _rows 中所有相同 PN 的零件实例 ─────────────────────────────
        #
        # 与初始加载时 _mass_cache 的设计一致：所有同 PN 实例共享同一个
        # _mass_props 对象引用，以确保后续手动修改重量时缩放逻辑正确运行。
        # 此处将所有实例的 _mass_props 统一指向同一个 new_mp 对象。
        target_rows: list[int] = [
            i for i, r in enumerate(self._rows)
            if str(r.get("Part Number", "")) == pn and r.get("Type") == "零件"
        ]

        for ri in target_rows:
            r = self._rows[ri]

            # 将所有实例的 _mass_props 指向同一个对象（与 _mass_cache 机制一致）
            r["_mass_props"]  = new_mp
            r["Weight"]       = new_mp["weight"]
            cog_local         = new_mp["cog"]
            r["CogX"]         = cog_local[0]
            r["CogY"]         = cog_local[1]
            r["CogZ"]         = cog_local[2]
            I_local           = new_mp["inertia"]
            r["Ixx"]          = I_local[0][0]
            r["Iyy"]          = I_local[1][1]
            r["Izz"]          = I_local[2][2]
            r["Ixy"]          = I_local[0][1]
            r["Ixz"]          = I_local[0][2]
            r["Iyz"]          = I_local[1][2]
            r["_meas_failed"] = False

            # 重新计算根坐标系质量特性（_placement 在初始加载时已保存，此处直接复用）
            placement = r.get("_placement")
            if placement is not None:
                R  = [[placement[i][j] for j in range(3)] for i in range(3)]
                T  = [placement[i][3] for i in range(3)]
                cog_root = [
                    sum(R[i][k] * cog_local[k] for k in range(3)) + T[i]
                    for i in range(3)
                ]
                RT = [[R[j][i] for j in range(3)] for i in range(3)]
                RI = [
                    [sum(R[i][k] * I_local[k][j] for k in range(3)) for j in range(3)]
                    for i in range(3)
                ]
                I_root = [
                    [sum(RI[i][k] * RT[k][j] for k in range(3)) for j in range(3)]
                    for i in range(3)
                ]
                r["_root_mp"] = {
                    "weight":  new_mp["weight"],
                    "cog":     cog_root,
                    "inertia": I_root,
                }
            else:
                # _placement 不存在（异常情况）：根坐标系与局部坐标系视为相同
                r["_root_mp"] = {
                    "weight":  new_mp["weight"],
                    "cog":     list(cog_local),
                    "inertia": [list(row_i) for row_i in I_local],
                }

        # ── 刷新可见树节点（_pn_to_items 中同 PN 的所有条目）─────────────────
        self._is_updating = True
        try:
            for vis_item in self._pn_to_items.get(pn, []):
                vis_row_idx = vis_item.data(0, _ROW_IDX_ROLE)
                if vis_row_idx is None:
                    continue
                vis_row = self._rows[vis_row_idx]
                if vis_row.get("Type") != "零件":
                    continue
                self._refresh_part_item_after_reread(vis_item, vis_row)
        finally:
            self._is_updating = False

        # ── 重新计算产品/部件汇总行并刷新底部计算结果 ──────────────────────
        recompute_product_rows(self._rows)
        self._refresh_product_items()
        self._rollup_result = None
        self._clear_summary_labels()
        self._calculate()

        QMessageBox.information(
            self, "重新读取成功",
            f"已成功重新读取零件「{pn}」的质量特性，\n"
            f"共更新了 {len(target_rows)} 个节点。",
        )

    def _refresh_part_item_after_reread(
        self,
        item: QTreeWidgetItem,
        row_data: dict,
    ) -> None:
        """重新读取质量特性成功后，更新零件行的视觉状态和显示值。

        清除之前因测量失败而设置的橙色背景和灰色文字，解除行锁定，
        并将新的质量特性数值写入各单元格。
        """
        default_brush = QBrush()  # 空画刷：重置为系统默认背景/前景

        # 恢复背景色、前景色和工具提示至默认状态
        for ci in range(len(self._columns)):
            item.setBackground(ci, default_brush)
            item.setForeground(ci, default_brush)
            item.setToolTip(ci, "")

        # 解除行锁定，允许编辑 Weight 列
        item.setData(0, _ITEM_LOCKED_ROLE, False)
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)

        # 更新各数值列的显示内容
        rmp = row_data.get("_root_mp")
        for col_idx, col_name in enumerate(self._columns):
            if col_name == "Weight":
                item.setText(col_idx, self._fmt_mass_val(row_data.get("Weight")))
            elif col_name in ("CogX", "CogY", "CogZ"):
                if self._summarize:
                    # 汇总BOM：显示零件自身坐标系值
                    raw = row_data.get(col_name)
                else:
                    # 层级BOM：显示根产品坐标系值（来自 _root_mp）
                    cog_idx = ("CogX", "CogY", "CogZ").index(col_name)
                    raw = rmp["cog"][cog_idx] if rmp else row_data.get(col_name)
                item.setText(col_idx, self._fmt_cog_val(raw) if raw is not None else "—")
            elif col_name in _INERTIA_IDX:
                ir, ic = _INERTIA_IDX[col_name]
                if self._summarize:
                    raw = row_data.get(col_name)
                else:
                    raw = (
                        rmp["inertia"][ir][ic]
                        if rmp and rmp.get("inertia")
                        else row_data.get(col_name)
                    )
                item.setText(col_idx, self._fmt_inertia_val(raw) if raw is not None else "—")
