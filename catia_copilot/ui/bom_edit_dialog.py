"""
BOM 编辑对话框模块。

提供：
- BomEditDialog – 可编辑表格，用于完成 BOM 属性并通过 COM 写回 CATIA。
"""

import copy
import os
import logging
import subprocess
from pathlib import Path

from PySide6.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTreeWidgetItem, QHeaderView, QAbstractItemView,
    QComboBox, QCheckBox, QGroupBox, QMessageBox, QApplication,
    QFileDialog, QProgressDialog, QRadioButton, QButtonGroup,
    QMenu, QWidgetAction, QLineEdit, QGridLayout
)
from PySide6.QtGui import QPixmap, QColor
from PySide6.QtCore import Qt, QSettings

from catia_copilot.constants import (
    PRESET_USER_REF_PROPERTIES,
    PRESET_USER_REF_PROPERTY_OPTIONS,
    BOM_EDIT_COLUMN_ORDER,
    BOM_COLUMN_DISPLAY_NAMES,
    BOM_READONLY_COLUMNS,
    BOM_HIDEABLE_COLUMNS,
    BOM_ROW_NUMBER_COLUMN,
    SOURCE_TO_DISPLAY,
    SOURCE_OPTIONS,
    PART_NUMBER_VALID_PATTERN,
    FILENAME_NOT_FOUND,
    FILENAME_UNSAVED,
    BOM_THUMBNAIL_MAX_SIZE,
)
from catia_copilot.catia.bom_collect import collect_bom_rows, flatten_bom_to_summary
from catia_copilot.catia.bom_write import write_bom_to_catia
from catia_copilot.utils import read_catia_thumbnail
from catia_copilot.ui.bom_catia_helpers import (
    _is_catia_com_error,
    _find_catia_doc_by_path,
)
from catia_copilot.ui.bom_widgets import _BomTreeDelegate, _BomTreeWidget, _ITEM_LOCKED_ROLE
from catia_copilot.ui.bom_file_rename_dialog import _FileRenameDialog

logger = logging.getLogger(__name__)



class BomEditDialog(QDialog):
    """可编辑BOM表格，用于补全产品属性并通过COM写回CATIA。

    - 文件名 / 层级 / 类型 / 数量 为只读结构属性。
    - 零件编号 可编辑，带重复检测。
    - 来源（Source）使用下拉框（未知 / 自制 / 外购）。
    - 共享相同零件编号的行联动更新。
    - "应用" 写回但不关闭对话框；"完成" 写回后关闭。
    """

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("BOM属性补全")
        self.setMinimumSize(900, 600)
        self.resize(1100, 700)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowType.WindowMaximizeButtonHint
            | Qt.WindowType.WindowMinimizeButtonHint
        )

        # ── 配置与持久化设置 ──────────────────────────────────────────────────
        # 与"导出BOM"对话框共享自定义列配置
        self._export_settings = QSettings("CATIACompanion", "ExportBOMDialog")
        self._last_browse_dir = self._export_settings.value("last_browse_dir", "")

        saved_custom = self._export_settings.value("custom_columns", [])
        if isinstance(saved_custom, str):
            saved_custom = [saved_custom]
        self._custom_columns: list[str] = list(saved_custom)

        # BomEditDialog 专用设置
        self._edit_settings  = QSettings("CATIACompanion", "BomEditDialog")
        saved_visible        = self._edit_settings.value("visible_preset_columns", [])
        if isinstance(saved_visible, str):
            saved_visible = [saved_visible]
        self._visible_preset_cols: list[str] = [
            c for c in saved_visible if c in PRESET_USER_REF_PROPERTIES
        ]

        # 可显示/隐藏的标准列（品名、版本、定义、来源）
        saved_hideable = self._edit_settings.value("visible_hideable_columns", BOM_HIDEABLE_COLUMNS)
        if isinstance(saved_hideable, str):
            saved_hideable = [saved_hideable]
        self._visible_hideable_cols: list[str] = [
            c for c in saved_hideable if c in BOM_HIDEABLE_COLUMNS
        ]

        self._summarize: bool = self._edit_settings.value("summarize", False, type=bool)
        self._summary_include_assemblies: bool = self._edit_settings.value(
            "summary_include_assemblies", False, type=bool
        )
        self._summary_sort_column: str = self._edit_settings.value(
            "summary_sort_column", "Part Number"
        )

        # 包含所有预设的完整自定义列列表；从CATIA预读时覆盖所有列，不受当前可见性限制
        self._all_custom_columns: list[str] = list(dict.fromkeys(
            self._custom_columns + list(PRESET_USER_REF_PROPERTIES)
        ))

        self._show_filepath_col: bool = self._edit_settings.value(
            "show_filepath_column", False, type=bool,
        )
        self._show_filename_col: bool = self._edit_settings.value(
            "show_filename_column", True, type=bool,
        )

        self._columns: list[str] = self._build_visible_columns()

        # ── 内部状态 ──────────────────────────────────────────────────────────
        # {原始零件编号: {列名: 值}}（规范数据，来源字段用显示标签）
        self._canonical_data: dict[str, dict[str, str]] = {}
        # 最后一次加载/应用时的快照，用于仅写回变更字段
        self._snapshot_data: dict[str, dict[str, str]] = {}
        # {原始零件编号: {列名, ...}} — 自上次写回以来已修改的字段
        self._modified_keys: dict[str, set[str]] = {}
        # 按遍历顺序排列的所有BOM行
        self._rows: list[dict] = []
        # 防止变更处理回调重入的标志
        self._is_updating: bool = False
        # 与 self._rows 平行的列表：self._item_by_row[i] 对应 self._rows[i] 的树形控件项
        self._item_by_row: list[QTreeWidgetItem] = []
        # BOM成功加载至少一次后置为True
        self._bom_loaded: bool = False
        # 原始（层级）BOM行，由 collect_bom_rows() 返回；切换显示模式时用于重建行数据
        self._raw_rows: list[dict] = []
        # 零件编号→树形项索引，用于快速联动更新（性能优化）
        self._pn_to_items: dict[str, list[QTreeWidgetItem]] = {}
        # 列名→像素宽度缓存；在列可见性切换时保留用户调整的列宽
        self._col_widths: dict[str, int] = {}

        # ── 界面布局 ──────────────────────────────────────────────────────────
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        # 数据来源选择行
        self._use_active_chk = QCheckBox("使用当前CATIA活动文档（不选择文件）")
        self._use_active_chk.toggled.connect(self._toggle_file_row)
        layout.addWidget(self._use_active_chk)

        file_row = QHBoxLayout()
        self._file_edit       = QLineEdit()
        self._file_edit.setPlaceholderText("选择一个CATProduct文件...")
        self._file_edit.setReadOnly(True)
        self._file_browse_btn = QPushButton("浏览...")
        self._file_browse_btn.clicked.connect(self._browse_file)
        self._load_btn        = QPushButton("加载BOM")
        self._load_btn.clicked.connect(self._load_bom)
        file_row.addWidget(self._file_edit)
        file_row.addWidget(self._file_browse_btn)
        file_row.addWidget(self._load_btn)
        layout.addLayout(file_row)

        # ── BOM类型与显示选项（紧凑分组）────────────────────────────────────
        display_group  = QGroupBox("BOM类型与显示选项")
        display_group.setMinimumHeight(60)  # 切换BOM类型时防止高度抖动
        display_layout = QVBoxLayout(display_group)
        display_layout.setSpacing(4)
        display_layout.setContentsMargins(8, 6, 8, 6)

        # 第一行：单选按钮 + 汇总选项
        bom_type_row = QHBoxLayout()
        self._bom_type_btn_group = QButtonGroup(self)
        self._radio_hierarchical = QRadioButton("层级BOM")
        self._radio_summary_bom  = QRadioButton("汇总BOM")
        if self._summarize:
            self._radio_summary_bom.setChecked(True)
        else:
            self._radio_hierarchical.setChecked(True)
        self._bom_type_btn_group.addButton(self._radio_hierarchical)
        self._bom_type_btn_group.addButton(self._radio_summary_bom)
        self._radio_summary_bom.toggled.connect(self._on_bom_type_changed)
        bom_type_row.addWidget(self._radio_hierarchical)
        bom_type_row.addWidget(self._radio_summary_bom)

        self._summary_opts_widget = QWidget()
        summary_opts_layout = QHBoxLayout(self._summary_opts_widget)
        summary_opts_layout.setContentsMargins(0, 0, 0, 0)
        summary_opts_layout.setSpacing(8)

        self._include_assemblies_chk = QCheckBox("包含产品和部件（子装配体）")
        self._include_assemblies_chk.setToolTip(
            "勾选后，汇总BOM中也会列出产品和部件（子装配体），而不仅限于零件。"
        )
        self._include_assemblies_chk.setChecked(self._summary_include_assemblies)
        self._include_assemblies_chk.toggled.connect(self._on_include_assemblies_toggled)
        summary_opts_layout.addWidget(self._include_assemblies_chk)
        summary_opts_layout.addSpacing(8)
        summary_opts_layout.addWidget(QLabel("排序列:"))
        self._sort_col_combo = QComboBox()
        _sort_cols = list(BOM_EDIT_COLUMN_ORDER) + [
            c for c in PRESET_USER_REF_PROPERTIES if c not in BOM_EDIT_COLUMN_ORDER
        ] + [
            c for c in self._custom_columns
            if c not in BOM_EDIT_COLUMN_ORDER and c not in PRESET_USER_REF_PROPERTIES
        ]
        for col in _sort_cols:
            self._sort_col_combo.addItem(BOM_COLUMN_DISPLAY_NAMES.get(col, col), col)
        sort_saved_idx = self._sort_col_combo.findData(self._summary_sort_column)
        if sort_saved_idx >= 0:
            self._sort_col_combo.setCurrentIndex(sort_saved_idx)
        self._sort_col_combo.currentIndexChanged.connect(self._on_sort_col_changed)
        summary_opts_layout.addWidget(self._sort_col_combo)

        self._summary_opts_widget.setVisible(self._summarize)
        bom_type_row.addWidget(self._summary_opts_widget)
        bom_type_row.addStretch()
        display_layout.addLayout(bom_type_row)

        layout.addWidget(display_group)

        hint = QLabel(
            "层级 / 类型 / 数量 为结构属性，不可编辑，"
            "零件编号可编辑但不能与其他行冲突，"
            "文件名/路径可编辑。"
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: gray; font-size: 11px;")
        layout.addWidget(hint)

        # 预设列可见性复选框（两行网格布局，对齐列）
        preset_group  = QGroupBox("属性列（勾选以显示）")
        preset_main_layout = QVBoxLayout(preset_group)
        preset_main_layout.setSpacing(8)
        preset_main_layout.setContentsMargins(8, 6, 8, 6)

        # 使用 QGridLayout 实现对齐与均匀分布
        grid_layout = QGridLayout()
        grid_layout.setSpacing(12)
        grid_layout.setColumnStretch(100, 1)  # 末尾添加弹性空间

        self._preset_checkboxes: dict[str, QCheckBox] = {}

        # 第0行：文件名复选框 + 显示完整路径 + 可隐藏标准列
        col = 0

        # "文件名"是内置列，但可像预设列一样切换可见性
        fn_cb = QCheckBox(BOM_COLUMN_DISPLAY_NAMES.get("Filename", "Filename"))
        fn_cb.setChecked(self._show_filename_col)
        fn_cb.toggled.connect(self._on_preset_col_toggled)
        grid_layout.addWidget(fn_cb, 0, col)
        self._preset_checkboxes["Filename"] = fn_cb
        col += 1

        # "显示完整路径"复选框紧跟文件名复选框之后
        self._filepath_chk = QCheckBox("显示完整路径")
        self._filepath_chk.setToolTip("勾选后文件名列将显示文件完整路径（含目录），而非仅文件名")
        self._filepath_chk.setChecked(self._show_filepath_col)
        self._filepath_chk.toggled.connect(self._on_show_filepath_toggled)
        grid_layout.addWidget(self._filepath_chk, 0, col)
        col += 1

        # 可隐藏标准列（品名、版本、定义、来源）
        for col_name in BOM_HIDEABLE_COLUMNS:
            cb = QCheckBox(BOM_COLUMN_DISPLAY_NAMES.get(col_name, col_name))
            cb.setChecked(col_name in self._visible_hideable_cols)
            cb.toggled.connect(self._on_hideable_col_toggled)
            grid_layout.addWidget(cb, 0, col)
            self._preset_checkboxes[col_name] = cb
            col += 1

        # 第1行：预设用户自定义属性（物料编码、物料名称等）
        col = 0
        for col_name in PRESET_USER_REF_PROPERTIES:
            cb = QCheckBox(col_name)
            cb.setChecked(col_name in self._visible_preset_cols)
            cb.toggled.connect(self._on_preset_col_toggled)
            grid_layout.addWidget(cb, 1, col)
            self._preset_checkboxes[col_name] = cb
            col += 1

        preset_main_layout.addLayout(grid_layout)
        layout.addWidget(preset_group)

        # BOM树形控件（替代 QTableWidget，原生支持展开/折叠）
        self._table = _BomTreeWidget()
        _init_headers = self._display_headers()
        self._table.setColumnCount(len(_init_headers))
        self._table.setHeaderLabels(_init_headers)
        hdr = self._table.header()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)
        hdr.setFixedHeight(28)
        self._table.setUniformRowHeights(True)
        self._table.setRootIsDecorated(True)
        self._table.setSortingEnabled(False)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._table.setAlternatingRowColors(True)
        self._table.setIndentation(16)
        self._table.setStyleSheet("QTreeWidget::item { min-height: 24px; }")
        self._table.itemChanged.connect(self._on_item_changed)
        hdr.sectionResized.connect(self._on_section_resized)
        _delegate = _BomTreeDelegate(lambda: self._columns, self._table)
        self._table.setItemDelegate(_delegate)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_tree_context_menu)
        layout.addWidget(self._table, 1)

        # 底部按钮行
        btn_row = QHBoxLayout()

        autofit_btn = QPushButton("自适应列宽")
        autofit_btn.setToolTip("根据内容自动调整所有列的宽度")
        autofit_btn.clicked.connect(self._autofit_columns)
        btn_row.addWidget(autofit_btn)

        expand_btn = QPushButton("全部展开")
        expand_btn.setToolTip("展开结构树中的所有节点")
        expand_btn.clicked.connect(self._table.expandAll)
        btn_row.addWidget(expand_btn)

        collapse_btn = QPushButton("全部折叠")
        collapse_btn.setToolTip("折叠结构树中的所有节点")
        collapse_btn.clicked.connect(self._table.collapseAll)
        btn_row.addWidget(collapse_btn)

        self._rename_btn = QPushButton("按零件编号修改文件名")
        self._rename_btn.setEnabled(False)
        self._rename_btn.clicked.connect(self._rename_by_part_number)
        btn_row.addWidget(self._rename_btn)

        self._rename_file_btn = QPushButton("另存为")
        self._rename_file_btn.setToolTip("对选中文件执行另存为操作（通过CATIA另存为）")
        self._rename_file_btn.setEnabled(False)
        self._rename_file_btn.clicked.connect(self._rename_selected_file)
        btn_row.addWidget(self._rename_file_btn)
        btn_row.addStretch()

        self._export_btn = QPushButton("导出表格")
        self._export_btn.setToolTip("将当前表格导出为 Excel（.xlsx）或 CSV 文件")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_table)
        btn_row.addWidget(self._export_btn)

        self._save_btn   = QPushButton("应用")
        self._save_btn.setEnabled(False)
        self._save_btn.clicked.connect(self._apply_changes)

        self._finish_btn = QPushButton("完成")
        self._finish_btn.setDefault(True)
        self._finish_btn.setEnabled(False)
        self._finish_btn.clicked.connect(self._finish_and_close)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)

        btn_row.addWidget(self._save_btn)
        btn_row.addWidget(self._finish_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

    # ── 文件/活动文档切换 ─────────────────────────────────────────────────────

    def _toggle_file_row(self, use_active: bool) -> None:
        self._file_edit.setEnabled(not use_active)
        self._file_browse_btn.setEnabled(not use_active)

    # ── BOM类型切换 ───────────────────────────────────────────────────────────

    def _on_bom_type_changed(self, summary_checked: bool) -> None:
        self._summarize = summary_checked
        self._edit_settings.setValue("summarize", summary_checked)
        self._summary_opts_widget.setVisible(summary_checked)
        # 若BOM已加载，则从原始行重新生成显示行并刷新表格
        if self._raw_rows:
            self._rows = (
                flatten_bom_to_summary(
                    self._raw_rows,
                    include_assemblies=self._summary_include_assemblies,
                    sort_column=self._summary_sort_column or None,
                )
                if summary_checked else self._raw_rows
            )
            self._rebuild_columns_and_repopulate()

    def _on_include_assemblies_toggled(self, checked: bool) -> None:
        self._summary_include_assemblies = checked
        self._edit_settings.setValue("summary_include_assemblies", checked)
        # 若BOM已加载且汇总模式激活，则重建汇总显示
        if self._summarize and self._raw_rows:
            self._rows = flatten_bom_to_summary(
                self._raw_rows,
                include_assemblies=checked,
                sort_column=self._summary_sort_column or None,
            )
            # 包含装配体时显示"类型"列；否则隐藏
            self._rebuild_columns_and_repopulate()

    def _on_sort_col_changed(self, _index: int) -> None:
        col = self._sort_col_combo.currentData()
        if col:
            self._summary_sort_column = col
            self._edit_settings.setValue("summary_sort_column", col)
            # 如有必要，对当前显示的汇总行重新排序
            if self._summarize and self._raw_rows:
                self._rows = flatten_bom_to_summary(
                    self._raw_rows,
                    include_assemblies=self._summary_include_assemblies,
                    sort_column=col,
                )
                self._populate_table()

    # ── 表格辅助方法 ──────────────────────────────────────────────────────────

    def _autofit_columns(self) -> None:
        """根据内容自动调整所有列宽，设有最小宽度下限。"""
        # QTreeWidget 使用 resizeColumnToContents(int)，而非 resizeColumnsToContents()
        min_width = 60
        for col_idx, col_name in enumerate(self._columns):
            self._table.resizeColumnToContents(col_idx)
            if self._table.columnWidth(col_idx) < min_width:
                self._table.setColumnWidth(col_idx, min_width)
            # 更新缓存，使后续列可见性切换能保留此列宽
            self._col_widths[col_name] = self._table.columnWidth(col_idx)

    def _on_section_resized(self, logical_index: int, _old_size: int, new_size: int) -> None:
        """用户或代码调整列宽时，将新宽度写入缓存。"""
        if logical_index < len(self._columns):
            self._col_widths[self._columns[logical_index]] = new_size

    def _rebuild_columns_and_repopulate(self) -> None:
        """重建可见列列表，更新表头，若已有行数据则刷新表格。"""
        # --- 不可变宽度快照（同时用于锚点计算和宽度恢复）---
        #
        # 在任何 Qt 树/表头操作之前先取快照。快照将持久缓存
        # （为此前隐藏的列提供列宽）与当前 columnWidth() 值
        # （对可见列具有权威性）合并，使锚点计算和宽度恢复都从
        # 稳定副本读取，而非从可能在重建过程中被修改的
        # self._col_widths 读取。
        width_snapshot: dict[str, int] = dict(self._col_widths)
        for col_idx, col_name in enumerate(self._columns):
            w = self._table.columnWidth(col_idx)
            width_snapshot[col_name] = w
            self._col_widths[col_name] = w  # keep persistent cache current

        # --- 锚定列水平滚动位置记录 ---
        # 目标：在添加/删除列后，保持视口最左侧可见列不变。
        # 跨列数变化的原始像素滚动值并不稳定，因为 Qt 可能重置滚动条，
        # 且删除视口左侧的列会使所有剩余可见列的像素位置发生偏移。
        #
        # 策略：在重建前确定视口左边缘所在的列名（及偏入该列的像素数），
        # 重建后再根据列宽重新计算目标水平滚动值。
        old_columns = list(self._columns)
        old_hscroll = self._table.horizontalScrollBar().value()

        anchor_col_name: str | None = None  # 视口左边缘所在列名
        anchor_offset: int = 0              # 偏入该列的像素数

        x = 0
        for col_name in old_columns:
            w = width_snapshot[col_name]    # 使用快照，而非实时 columnWidth
            if x + w > old_hscroll:
                anchor_col_name = col_name
                anchor_offset = old_hscroll - x
                break
            x += w

        vscroll = self._table.verticalScrollBar().value()

        self._columns = self._build_visible_columns()
        if self._rows:
            self._populate_table()  # 内部已设置列数和表头
            # 从不可变快照恢复列宽；对从未出现过的新列执行自适应宽度
            for col_idx, col_name in enumerate(self._columns):
                if col_name in width_snapshot:
                    self._table.setColumnWidth(col_idx, width_snapshot[col_name])
                else:
                    self._table.resizeColumnToContents(col_idx)

            # 根据锚定列的新像素位置计算新的水平滚动值
            new_hscroll = 0
            if anchor_col_name is not None:
                # 在新布局中查找锚定列
                x = 0
                found = False
                for col_idx, col_name in enumerate(self._columns):
                    if col_name == anchor_col_name:
                        new_hscroll = x + anchor_offset
                        found = True
                        break
                    x += self._table.columnWidth(col_idx)

                if not found:
                    # 锚定列已被隐藏；滚动到旧布局中锚定列右侧第一个
                    # 仍存在的列（即被删除列右边第一个幸存列）
                    old_col_order = {c: i for i, c in enumerate(old_columns)}
                    anchor_old_idx = old_col_order.get(anchor_col_name, -1)
                    x = 0
                    for col_idx, col_name in enumerate(self._columns):
                        if old_col_order.get(col_name, -1) > anchor_old_idx:
                            new_hscroll = x
                            break
                        x += self._table.columnWidth(col_idx)
                    else:
                        # 所有剩余列均在锚定列原位置左侧；滚动到末尾
                        new_hscroll = self._table.horizontalScrollBar().maximum()

            new_hscroll = max(0, min(new_hscroll, self._table.horizontalScrollBar().maximum()))
            self._table.verticalScrollBar().setValue(vscroll)
            self._table.horizontalScrollBar().setValue(new_hscroll)
        else:
            # 尚无行数据：仅更新列数和表头，以便在加载BOM前也能反映最新列选择
            _headers = self._display_headers()
            self._table.setColumnCount(len(_headers))
            self._table.setHeaderLabels(_headers)

    # ── 列可见性管理 ──────────────────────────────────────────────────────────

    def _display_headers(self) -> list[str]:
        """返回当前列列表的显示表头标签。

        当"显示完整路径"选项激活时，文件名列的表头显示为"完整路径"，
        以便用户直观区分。
        """
        result = []
        for c in self._columns:
            if c == "Filename" and self._show_filepath_col:
                result.append("完整路径")
            else:
                result.append(BOM_COLUMN_DISPLAY_NAMES.get(c, c))
        return result

    def _build_visible_columns(self) -> list[str]:
        base = list(BOM_EDIT_COLUMN_ORDER)
        # 过滤隐藏列（文件名列和可隐藏列）
        if not self._show_filename_col:
            base = [c for c in base if c != "Filename"]
        # 过滤已隐藏的可隐藏列
        base = [c for c in base if c not in BOM_HIDEABLE_COLUMNS or c in self._visible_hideable_cols]
        if self._summarize:
            # 汇总模式下层级列无意义；不含装配体时也隐藏类型列
            cols_to_hide = {"Level"}
            if not self._summary_include_assemblies:
                cols_to_hide.add("Type")
            base = [c for c in base if c not in cols_to_hide]
        visible_preset = [
            c for c in PRESET_USER_REF_PROPERTIES if c in self._visible_preset_cols
        ]
        other_custom   = [
            c for c in self._custom_columns
            if c not in BOM_EDIT_COLUMN_ORDER and c not in PRESET_USER_REF_PROPERTIES
        ]
        # 将"#"紧插在"Level"之后（逻辑索引0→Level，逻辑索引1→"#"），
        # 使 QTreeWidget 的树形装饰（分支线）保留在 Level 列（逻辑索引0）。
        # 汇总模式下 Level 被隐藏，"#"自然落到第0列，无需特殊处理。
        result = base + visible_preset + other_custom
        if "Level" in result:
            level_idx = result.index("Level")
            result.insert(level_idx + 1, BOM_ROW_NUMBER_COLUMN)
        else:
            result.insert(0, BOM_ROW_NUMBER_COLUMN)
        return result

    def _on_preset_col_toggled(self) -> None:
        # "文件名"复选框控制内置文件名列的可见性
        if "Filename" in self._preset_checkboxes:
            new_show_fn = self._preset_checkboxes["Filename"].isChecked()
            if new_show_fn != self._show_filename_col:
                self._show_filename_col = new_show_fn
                self._edit_settings.setValue("show_filename_column", self._show_filename_col)
        self._visible_preset_cols = [
            name for name, cb in self._preset_checkboxes.items()
            if name != "Filename" and name not in BOM_HIDEABLE_COLUMNS and cb.isChecked()
        ]
        self._edit_settings.setValue("visible_preset_columns", self._visible_preset_cols)
        self._rebuild_columns_and_repopulate()

    def _on_hideable_col_toggled(self) -> None:
        """处理可隐藏列复选框切换（品名、版本、定义、来源）。"""
        self._visible_hideable_cols = [
            name for name, cb in self._preset_checkboxes.items()
            if name in BOM_HIDEABLE_COLUMNS and cb.isChecked()
        ]
        self._edit_settings.setValue("visible_hideable_columns", self._visible_hideable_cols)
        self._rebuild_columns_and_repopulate()

    def _on_show_filepath_toggled(self, checked: bool) -> None:
        self._show_filepath_col = checked
        self._edit_settings.setValue("show_filepath_column", checked)
        self._rebuild_columns_and_repopulate()

    # ── 文件选择 ──────────────────────────────────────────────────────────────

    def _browse_file(self) -> None:
        file, _ = QFileDialog.getOpenFileName(
            self, "选择CATProduct文件",
            self._last_browse_dir,
            "*.CATProduct (*.CATProduct);;All Files (*)",
        )
        if file:
            self._file_edit.setText(file)
            self._last_browse_dir = str(Path(file).parent)
            self._export_settings.setValue("last_browse_dir", self._last_browse_dir)

    # ── 加载BOM ───────────────────────────────────────────────────────────────

    def _load_bom(self) -> None:
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请先选择一个CATProduct文件。")
                return
            if not Path(file_path).exists():
                QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{file_path}")
                return

        self._load_btn.setEnabled(False)
        self._load_btn.setText("加载中…")
        QApplication.processEvents()

        progress = QProgressDialog("正在加载BOM，请稍候…", None, 0, 0, self)
        progress.setWindowTitle("加载BOM")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)
        progress.setValue(0)

        def _on_row_collected(count: int) -> None:
            progress.setLabelText(f"正在加载BOM，请稍候… 已读取 {count} 个节点")
            progress.repaint()
            QApplication.processEvents()

        try:
            all_read_cols = list(dict.fromkeys(
                BOM_EDIT_COLUMN_ORDER
                + [c for c in self._all_custom_columns if c not in BOM_EDIT_COLUMN_ORDER]
            ))
            rows = collect_bom_rows(
                file_path, all_read_cols, self._all_custom_columns,
                progress_callback=_on_row_collected,
            )
        except Exception as e:
            progress.close()
            logger.error(f"Failed to load BOM for edit: {e}")
            QMessageBox.critical(
                self, "加载失败",
                f"加载BOM时出错：\n{e}\n\n请确保CATIA已启动。",
            )
            self._load_btn.setEnabled(True)
            self._load_btn.setText("加载BOM")
            return
        finally:
            progress.close()

        self._load_btn.setEnabled(True)
        self._load_btn.setText("重新加载BOM")

        # 始终保存原始层级行，以便之后切换显示模式
        self._raw_rows = rows

        # 汇总模式下将层级折叠为唯一零件
        display_rows = (
            flatten_bom_to_summary(
                rows,
                include_assemblies=self._summary_include_assemblies,
                sort_column=self._summary_sort_column or None,
            )
            if self._summarize else rows
        )

        self._rows = display_rows

        # 以零件编号为键构建规范数据（首次出现者优先），
        # 使用原始行确保所有零件被索引，不受当前显示模式影响
        all_data_cols = list(dict.fromkeys(
            BOM_EDIT_COLUMN_ORDER
            + [c for c in self._all_custom_columns if c not in BOM_EDIT_COLUMN_ORDER]
        ))
        self._canonical_data = {}
        for row in rows:
            pn = str(row.get("Part Number", ""))
            if pn and pn not in self._canonical_data:
                data: dict[str, str] = {}
                for col in all_data_cols:
                    val = str(row.get(col, ""))
                    if col == "Source":
                        val = SOURCE_TO_DISPLAY.get(val, val)
                    data[col] = val
                self._canonical_data[pn] = data

        self._snapshot_data  = copy.deepcopy(self._canonical_data)
        self._modified_keys.clear()

        # 刷新前按列名保存当前列宽
        if self._bom_loaded:
            for col_idx, col_name in enumerate(self._columns):
                self._col_widths[col_name] = self._table.columnWidth(col_idx)

        self._populate_table()
        if not self._bom_loaded:
            # 首次加载：自适应所有列宽并初始化缓存；
            # "#"行号列固定默认宽度40像素（之后可像其他列一样调整）
            for _c, col_name in enumerate(self._columns):
                if col_name == BOM_ROW_NUMBER_COLUMN:
                    self._table.setColumnWidth(_c, 40)
                    self._col_widths[col_name] = 40
                else:
                    self._table.resizeColumnToContents(_c)
                    self._col_widths[col_name] = self._table.columnWidth(_c)
            self._bom_loaded = True
        else:
            # 后续重新加载：按列名恢复已保存的列宽
            for col_idx, col_name in enumerate(self._columns):
                if col_name in self._col_widths:
                    self._table.setColumnWidth(col_idx, self._col_widths[col_name])

        self._save_btn.setEnabled(True)
        self._finish_btn.setEnabled(True)
        self._rename_btn.setEnabled(True)
        self._rename_file_btn.setEnabled(True)
        self._export_btn.setEnabled(True)

    def _populate_table(self) -> None:
        self._is_updating = True
        self._table.blockSignals(True)

        # 汇总模式：所有行均为无子项的顶层项。
        # 若保持 setRootIsDecorated(True)，每行都会预留展开箭头的空间，
        # 使第0列内容向右偏移。汇总模式下禁用，层级模式下重新启用以显示展开箭头。
        self._table.setRootIsDecorated(not self._summarize)

        self._table.clear()                          # 删除所有项；表头保留
        headers = self._display_headers()
        self._table.setColumnCount(len(headers))     # Qt 不会自动缩减列数
        self._table.setHeaderLabels(headers)
        self._item_by_row = []
        self._pn_to_items.clear()  # 重置零件编号→树形项索引

        # parent_stack：(层级, 树形项|None) 的列表
        # 索引0处的哨兵代表不可见根节点（层级为−1）
        parent_stack: list[tuple[int, QTreeWidgetItem | None]] = [(-1, None)]

        for row_idx, row_data in enumerate(self._rows):
            level = 0 if self._summarize else int(row_data.get("Level", 0))

            # 弹出栈，直到栈顶层级严格低于当前行
            while len(parent_stack) > 1 and parent_stack[-1][0] >= level:
                parent_stack.pop()

            parent_item = parent_stack[-1][1]
            item = QTreeWidgetItem()
            # 将 row_idx 存入第0列的 UserRole，用于反向查找
            item.setData(0, Qt.ItemDataRole.UserRole, row_idx)

            if parent_item is None:
                self._table.addTopLevelItem(item)
            else:
                parent_item.addChild(item)

            parent_stack.append((level, item))
            self._item_by_row.append(item)

            pn         = str(row_data.get("Part Number", ""))
            not_found  = bool(row_data.get("_not_found"))
            no_file    = bool(row_data.get("_no_file"))
            unreadable = bool(row_data.get("_unreadable"))
            row_locked = unreadable or not_found

            # 构建零件编号→树形项索引，用于快速联动更新
            if pn:
                self._pn_to_items.setdefault(pn, []).append(item)

            for col_idx, col_name in enumerate(self._columns):

                # 来源列 → QComboBox（覆盖控件；不存储为项文本）
                if col_name == "Source":
                    raw    = str(row_data.get("Source", ""))
                    pn_val = self._canonical_data.get(pn, {}).get(
                        "Source", SOURCE_TO_DISPLAY.get(raw, raw)
                    )
                    if pn_val not in SOURCE_OPTIONS:
                        pn_val = SOURCE_TO_DISPLAY.get(pn_val, SOURCE_OPTIONS[0])
                    combo = QComboBox()
                    combo.blockSignals(True)
                    combo.addItems(SOURCE_OPTIONS)
                    combo.setCurrentText(pn_val)
                    combo.blockSignals(False)
                    if row_locked:
                        combo.setEnabled(False)
                    else:
                        combo.currentTextChanged.connect(
                            lambda text, r=row_idx: self._on_source_changed(r, text)
                        )
                    self._table.setItemWidget(item, col_idx, combo)
                    continue

                # 具有受限选项的用户自定义属性列 → QComboBox
                opts = PRESET_USER_REF_PROPERTY_OPTIONS.get(col_name)
                if opts is not None:
                    pn_val = self._canonical_data.get(pn, {}).get(
                        col_name, str(row_data.get(col_name, ""))
                    )
                    # 构建有效选项列表：
                    #   • 始终在开头插入""，使未设置的属性显示为空白
                    #   • 若存储值不在允许列表中且非空，则追加以保留原始值可见性
                    display_opts = [""] + list(opts)
                    if pn_val and pn_val not in opts:
                        logger.debug(
                            "属性 '%s' 的值 '%s' 不在可选列表中，将以原始值显示（零件编号: %s）",
                            col_name, pn_val, pn,
                        )
                        display_opts.append(pn_val)
                    combo = QComboBox()
                    combo.blockSignals(True)
                    combo.addItems(display_opts)
                    combo.setCurrentText(pn_val)
                    combo.blockSignals(False)
                    if row_locked:
                        combo.setEnabled(False)
                    else:
                        combo.currentTextChanged.connect(
                            lambda text, r=row_idx, c=col_name: self._on_option_col_changed(r, c, text)
                        )
                    self._table.setItemWidget(item, col_idx, combo)
                    continue

                # 其他所有列 → 项文本
                if col_name == BOM_ROW_NUMBER_COLUMN:
                    value = str(row_idx + 1)
                elif col_name == "Quantity":
                    value = str(row_data.get("Quantity", "1"))
                elif col_name == "Filename":
                    fp = str(row_data.get("_filepath", ""))
                    fn = str(row_data.get("Filename", ""))
                    if no_file:
                        # 文件未保存到磁盘：固定显示哨兵文本
                        value = FILENAME_UNSAVED
                    elif self._show_filepath_col:
                        value = fp if fp else fn
                    else:
                        # 已知路径时显示带扩展名的文件名；
                        # 未知路径时回退到存储的文件名茎（可能等于 FILENAME_NOT_FOUND）
                        value = Path(fp).name if fp else fn
                elif col_name == "Filepath":
                    value = str(row_data.get("_filepath", ""))
                elif col_name in BOM_READONLY_COLUMNS:
                    value = str(row_data.get(col_name, ""))
                else:
                    value = str(
                        self._canonical_data.get(pn, {}).get(
                            col_name, row_data.get(col_name, "")
                        )
                    )
                item.setText(col_idx, value)

                if col_name == "Filename":
                    fp = str(row_data.get("_filepath", ""))
                    fn = str(row_data.get("Filename", ""))
                    if no_file:
                        pass  # tooltip 由下方 no_file 块统一设置
                    elif fp:
                        if self._show_filepath_col:
                            # 列显示完整路径；工具提示显示文件名+扩展名
                            name_with_ext = Path(fp).name
                            if name_with_ext and name_with_ext != FILENAME_NOT_FOUND:
                                item.setToolTip(col_idx, name_with_ext)
                        else:
                            # 列显示文件名+扩展名；工具提示显示完整路径
                            item.setToolTip(col_idx, fp)

            # 未锁定行：允许就地编辑（代理阻止只读列）
            if not row_locked:
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                item.setData(0, _ITEM_LOCKED_ROLE, False)
            else:
                grey = QColor(160, 160, 160)
                bg   = QColor(255, 205, 205) if not_found else QColor(245, 245, 245)
                item.setData(0, _ITEM_LOCKED_ROLE, True)
                for ci in range(len(self._columns)):
                    item.setForeground(ci, grey)
                    item.setBackground(ci, bg)
                fn_col = self._columns.index("Filename") if "Filename" in self._columns else -1
                if fn_col >= 0:
                    tip = (
                        "该零件/装配体的文件未被CATIA检索到，行内容不可编辑。"
                        if not_found else
                        "该零件/装配体处于轻量化模式，无法读取属性。"
                    )
                    item.setToolTip(fn_col, tip)

            # _no_file 行（文件未保存到磁盘）：不锁定，但以淡黄背景和专属提示标识
            if no_file:
                fn_col = self._columns.index("Filename") if "Filename" in self._columns else -1
                bg_unsaved = QColor(255, 245, 180)
                for ci in range(len(self._columns)):
                    item.setBackground(ci, bg_unsaved)
                if fn_col >= 0:
                    item.setToolTip(fn_col, "该零件尚未保存到磁盘，可通过右键菜单「另存为」将其保存。")

        self._table.expandAll()
        self._table.blockSignals(False)
        self._is_updating = False

    # ── 树形遍历辅助 ──────────────────────────────────────────────────────────

    def _iter_all_items(self):
        """以深度优先前序遍历方式逐个产出所有 QTreeWidgetItem。"""
        def _walk(parent: QTreeWidgetItem):
            yield parent
            for i in range(parent.childCount()):
                yield from _walk(parent.child(i))
        for i in range(self._table.topLevelItemCount()):
            yield from _walk(self._table.topLevelItem(i))

    # ── "来源"下拉框变更 ──────────────────────────────────────────────────────

    def _on_source_changed(self, row_idx: int, text: str) -> None:
        if self._is_updating:
            return
        if "Source" not in self._columns:
            return
        src_col_idx = self._columns.index("Source")

        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        direct_rows = selected_row_indices if row_idx in selected_row_indices else {row_idx}

        pns_to_update: set[str] = set()
        for r in direct_rows:
            pn = str(self._rows[r].get("Part Number", ""))
            if pn:
                pns_to_update.add(pn)

        for pn in pns_to_update:
            if pn in self._canonical_data:
                self._canonical_data[pn]["Source"] = text
                self._modified_keys.setdefault(pn, set()).add("Source")

        # 性能优化：使用零件编号→树形项索引，避免全树遍历
        self._is_updating = True
        for pn in pns_to_update:
            if pn in self._pn_to_items:
                for other_item in self._pn_to_items[pn]:
                    combo = self._table.itemWidget(other_item, src_col_idx)
                    if isinstance(combo, QComboBox) and combo.currentText() != text:
                        combo.blockSignals(True)
                        combo.setCurrentText(text)
                        combo.blockSignals(False)
        self._is_updating = False

    # ── 用户自定义选项列变更 ──────────────────────────────────────────────────

    def _on_option_col_changed(self, row_idx: int, col_name: str, text: str) -> None:
        if self._is_updating:
            return
        if col_name not in self._columns:
            return
        col_idx = self._columns.index(col_name)

        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        direct_rows = selected_row_indices if row_idx in selected_row_indices else {row_idx}

        pns_to_update: set[str] = set()
        for r in direct_rows:
            pn = str(self._rows[r].get("Part Number", ""))
            if pn:
                pns_to_update.add(pn)

        for pn in pns_to_update:
            if pn in self._canonical_data:
                self._canonical_data[pn][col_name] = text
                self._modified_keys.setdefault(pn, set()).add(col_name)

        # 性能优化：使用零件编号→树形项索引，避免全树遍历
        self._is_updating = True
        for pn in pns_to_update:
            if pn in self._pn_to_items:
                for other_item in self._pn_to_items[pn]:
                    combo = self._table.itemWidget(other_item, col_idx)
                    if isinstance(combo, QComboBox) and combo.currentText() != text:
                        combo.blockSignals(True)
                        combo.setCurrentText(text)
                        combo.blockSignals(False)
        self._is_updating = False

    # ── 普通单元格编辑 ────────────────────────────────────────────────────────

    def _on_item_changed(self, item: QTreeWidgetItem, col_idx: int) -> None:
        if self._is_updating:
            return
        row_idx = item.data(0, Qt.ItemDataRole.UserRole)
        if row_idx is None:
            return
        col_name = self._columns[col_idx]

        if col_name in BOM_READONLY_COLUMNS or col_name == "Source" or col_name in PRESET_USER_REF_PROPERTY_OPTIONS:
            return

        new_value = item.text(col_idx)
        pn        = str(self._rows[row_idx].get("Part Number", ""))

        if col_name == "Part Number":
            # ── 零件编号为空或仅含空格 ────────────────────────────────────────
            if not new_value.strip():
                QMessageBox.warning(
                    self, "零件编号不能为空",
                    "零件编号不能为空或仅含空格，请输入有效的零件编号。",
                )
                self._is_updating = True
                item.setText(col_idx, self._canonical_data.get(pn, {}).get("Part Number", pn))
                self._is_updating = False
                return

            # ── 静默去除首尾空格 ──────────────────────────────────────────────
            if new_value != new_value.strip():
                new_value = new_value.strip()
                self._is_updating = True
                item.setText(col_idx, new_value)
                self._is_updating = False

            # ── 字符合法性校验 ────────────────────────────────────────────────
            if not PART_NUMBER_VALID_PATTERN.fullmatch(new_value):
                QMessageBox.warning(
                    self, "零件编号含非法字符",
                    f"零件编号 \"{new_value}\" 含有非法字符。\n"
                    "不允许：控制字符、非ASCII字符，以及Windows文件名禁用字符"
                    "（\\ / : * ? \" < > |）。",
                )
                self._is_updating = True
                item.setText(col_idx, self._canonical_data.get(pn, {}).get("Part Number", pn))
                self._is_updating = False
                return

            # ── 与当前规范值冲突检查 ──────────────────────────────────────────
            for other_pn, data in self._canonical_data.items():
                if other_pn == pn:
                    continue
                if data.get("Part Number", other_pn) == new_value:
                    QMessageBox.warning(
                        self, "零件编号冲突",
                        f"零件编号 \"{new_value}\" 与 \"{other_pn}\" "
                        f"的当前零件编号冲突，不允许修改。",
                    )
                    self._is_updating = True
                    item.setText(col_idx, self._canonical_data.get(pn, {}).get("Part Number", pn))
                    self._is_updating = False
                    return

            # ── 与快照（CATIA当前值）冲突检查 ───────────────────────────────
            for other_pn, data in self._snapshot_data.items():
                if other_pn == pn:
                    continue
                if data.get("Part Number", other_pn) == new_value:
                    QMessageBox.warning(
                        self, "零件编号冲突",
                        f"零件编号 \"{new_value}\" 与 \"{other_pn}\" "
                        f"的原始零件编号冲突，不允许修改。",
                    )
                    self._is_updating = True
                    item.setText(col_idx, self._canonical_data.get(pn, {}).get("Part Number", pn))
                    self._is_updating = False
                    return

        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        direct_rows = selected_row_indices if row_idx in selected_row_indices else {row_idx}

        pns_to_update: set[str] = set()
        for r in direct_rows:
            r_pn = str(self._rows[r].get("Part Number", ""))
            if r_pn:
                pns_to_update.add(r_pn)
                if r_pn in self._canonical_data:
                    self._canonical_data[r_pn][col_name] = new_value
                    self._modified_keys.setdefault(r_pn, set()).add(col_name)

        # 性能优化：使用零件编号→树形项索引，避免全树遍历
        self._is_updating = True
        for pn in pns_to_update:
            if pn in self._pn_to_items:
                for other_item in self._pn_to_items[pn]:
                    if other_item.text(col_idx) != new_value:
                        other_item.setText(col_idx, new_value)
        self._is_updating = False

    # ── 写回CATIA ─────────────────────────────────────────────────────────────

    def _rename_by_part_number(self) -> None:
        """通过CATIA另存为功能，将每个CATIA文件按零件编号改名。"""
        if self._modified_keys:
            ret = QMessageBox.question(
                self, "存在未回传的修改",
                "检测到BOM属性尚未写回CATIA。\n\n"
                "必须先将修改写回CATIA，才能确保零件编号与CATIA文件一致。\n\n"
                "是否立即执行写回？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return
            self._write_back(close_on_success=False)
            # 若写回失败或仅部分成功，modified_keys 仍非空——在此停止，让用户先修复问题
            if self._modified_keys:
                return
            # 写回已清除所有修改；继续执行改名

        to_rename: list[tuple[str, str]] = []
        seen_fps:  set[str] = set()
        for row in self._rows:
            fp = str(row.get("_filepath", ""))
            if not fp or fp in seen_fps:
                continue
            seen_fps.add(fp)
            orig_pn = str(row.get("Part Number", ""))
            pn      = str(self._canonical_data.get(orig_pn, {}).get("Part Number", orig_pn))
            if pn and Path(fp).stem != pn:
                to_rename.append((fp, pn))

        if not to_rename:
            QMessageBox.information(self, "无需改名", "所有文件名已与零件编号一致。")
            return

        delete_old = (
            QMessageBox.question(
                self, "是否删除旧文件",
                "另存为完成后，是否删除旧文件？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes
        )

        QMessageBox.information(self, "请在CATIA中继续操作", "准备就绪，请在CATIA中确认后续操作。")

        renamed_count = 0

        # 性能优化：一次性构建文档缓存，避免重复扫描
        from pycatia import catia as _pycatia
        caa         = _pycatia()
        application = caa.application
        application.visible = True
        documents   = application.documents

        # 缓存：文件路径 → 文档对象
        doc_cache: dict[Path, object] = {}
        for i in range(1, documents.count + 1):
            try:
                doc = documents.item(i)
                doc_path = Path(doc.full_name).resolve()
                doc_cache[doc_path] = doc
            except Exception:
                pass

        for fp, pn in reversed(to_rename):
            if not PART_NUMBER_VALID_PATTERN.fullmatch(pn):
                QMessageBox.warning(
                    self, "零件编号含非法字符",
                    f"零件编号 「{pn}」 含有非法字符。\n"
                    "不允许：控制字符、非ASCII字符，以及Windows文件名禁用字符"
                    "（\\ / : * ? \" < > |）。\n请在表格中修改此零件编号后重试。",
                )
                continue

            if not Path(fp).exists():
                continue

            ext    = Path(fp).suffix
            new_fp = str(Path(fp).parent / (pn + ext))
            target_existed_before = Path(new_fp).exists()

            try:
                src = Path(fp).resolve()

                # 使用缓存快速查找
                target_doc = doc_cache.get(src)
                if target_doc is None:
                    documents.open(str(src))
                    candidate = documents.item(documents.count)
                    target_doc = (
                        candidate
                        if Path(candidate.full_name).resolve() == src
                        else _find_catia_doc_by_path(documents, src)
                    )
                    if target_doc:
                        doc_cache[src] = target_doc

                if target_doc is None:
                    QMessageBox.warning(
                        self, "无法找到文档",
                        f"无法在CATIA中找到或打开文档：\n{fp}",
                    )
                    continue

                target_doc.com_object.SaveAs(new_fp)

                if delete_old and Path(fp).resolve() != Path(new_fp).resolve():
                    try:
                        os.remove(fp)
                    except Exception as del_err:
                        logger.warning(f"Failed to delete old file {fp}: {del_err}")

                for row in self._rows:
                    if str(row.get("_filepath", "")) == fp:
                        row["_filepath"] = new_fp
                        row["Filename"]  = pn
                for row in self._raw_rows:
                    if str(row.get("_filepath", "")) == fp:
                        row["_filepath"] = new_fp
                        row["Filename"]  = pn
                renamed_count += 1

                # 用新路径更新缓存
                if Path(new_fp).resolve() != src:
                    doc_cache[Path(new_fp).resolve()] = target_doc
                    if src in doc_cache:
                        del doc_cache[src]

            except Exception as e:
                # 仅在以下所有条件成立时将异常视为用户主动取消：
                #   1. 异常来自CATIA COM层（pywintypes.com_error）——
                #      这是用户在CATIA自身另存为对话框中点击"取消"或"否"时
                #      CATIA抛出的异常，而非操作系统级别的失败。
                #   2. 源文件仍然完好。
                #   3. 目标文件要么在操作前就已存在，要么从未被创建。
                # 其他任何异常（OSError、PermissionError、磁盘空间不足等非COM错误）
                # 均应作为真正的失败弹出提示。
                if _is_catia_com_error(e) and Path(fp).exists() and (
                    target_existed_before or not Path(new_fp).exists()
                ):
                    logger.info(
                        f"SaveAs skipped for {Path(fp).name} "
                        "(user cancelled or declined overwrite in CATIA)"
                    )
                    continue
                QMessageBox.warning(
                    self, "另存为失败", f"文件「{Path(fp).name}」另存为失败：\n{e}"
                )

        if renamed_count > 0:
            QMessageBox.information(
                self, "改名完成",
                f"已成功将 {renamed_count} 个文件通过CATIA另存为功能改名。",
            )
            self._populate_table()

    def _rename_selected_file(self) -> None:
        """通过CATIA另存为功能，对选中的单行BOM记录执行重命名或移动操作。"""
        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        if len(selected_row_indices) != 1:
            QMessageBox.warning(
                self, "请选择单行",
                "请在表格中选中恰好一行，再执行此操作。",
            )
            return

        row_idx  = next(iter(selected_row_indices))
        row_data = self._rows[row_idx]
        fp       = str(row_data.get("_filepath", ""))

        if not fp or row_data.get("_not_found"):
            QMessageBox.warning(self, "无有效路径", "该行没有可用的文件路径，无法执行重命名/移动。")
            return
        # 注意：此处不检查 Path(fp).exists()；
        # 未保存过的零件（文件尚不在磁盘上但在CATIA内存中打开）同样允许另存为。

        # 改名前要求先写回属性，以确保文件内容与表格一致
        orig_pn = str(row_data.get("Part Number", ""))
        if orig_pn in self._modified_keys:
            ret = QMessageBox.question(
                self, "存在未写回的属性修改",
                f"零件「{orig_pn}」的属性尚未写回CATIA。\n\n"
                "必须先将修改写回CATIA，才能确保文件内容与表格一致。\n\n"
                "是否立即执行写回？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return
            self._write_back(close_on_success=False)
            # 仅当写回确实清除了修改才继续；
            # 若写回失败（弹出错误对话框），modified_keys 中仍保留该条目
            if orig_pn in self._modified_keys:
                return

        dlg = _FileRenameDialog(fp, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        new_fp                = dlg.new_path
        target_existed_before = Path(new_fp).exists()

        # 仅当旧文件实际存在于磁盘时才询问是否删除，否则无从删除
        file_on_disk = Path(fp).exists()
        delete_old = file_on_disk and (
            QMessageBox.question(
                self, "是否删除旧文件",
                f"另存为完成后，是否删除旧文件？\n\n旧文件：{fp}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes
        )

        QMessageBox.information(self, "请在CATIA中继续操作", "准备就绪，请在CATIA中确认后续操作。")

        try:
            from pycatia import catia as _pycatia
            caa         = _pycatia()
            application = caa.application
            application.visible = True
            documents   = application.documents
            src         = Path(fp).resolve()

            target_doc = _find_catia_doc_by_path(documents, src)
            # 仅当文件在磁盘上存在时才尝试打开；
            # 未保存过的零件只能通过已打开的文档缓存找到。
            if target_doc is None and file_on_disk:
                documents.open(str(src))
                candidate  = documents.item(documents.count)
                target_doc = (
                    candidate
                    if Path(candidate.full_name).resolve() == src
                    else _find_catia_doc_by_path(documents, src)
                )

            if target_doc is None:
                QMessageBox.warning(
                    self, "无法找到文档",
                    f"无法在CATIA中找到或打开文档：\n{fp}\n\n"
                    "请确认该零件已在CATIA中打开。",
                )
                return

            target_doc.com_object.SaveAs(new_fp)

            if delete_old and Path(fp).resolve() != Path(new_fp).resolve():
                try:
                    os.remove(fp)
                except Exception as del_err:
                    logger.warning(f"Failed to delete old file {fp}: {del_err}")

            new_stem = Path(new_fp).stem
            for row in self._rows:
                if str(row.get("_filepath", "")) == fp:
                    row["_filepath"] = new_fp
                    row["Filename"]  = new_stem
            for row in self._raw_rows:
                if str(row.get("_filepath", "")) == fp:
                    row["_filepath"] = new_fp
                    row["Filename"]  = new_stem
            self._populate_table()
            QMessageBox.information(
                self, "操作成功",
                f"文件已成功另存为：\n{new_fp}",
            )

        except Exception as e:
            # 仅在以下所有条件成立时将异常视为用户主动取消：
            #   1. 异常来自CATIA COM层（pywintypes.com_error）——
            #      这是用户在CATIA自身另存为对话框中点击"取消"或"否"时抛出的，
            #      而非操作系统级别的失败。
            #   2. 源文件仍然完好（或本来就未在磁盘上）。
            #   3. 目标文件要么在操作前就已存在，要么从未被创建。
            # 其他任何异常（OSError、PermissionError、磁盘空间不足等非COM错误）
            # 均应作为真正的失败弹出提示。
            source_intact = not file_on_disk or Path(fp).exists()
            if _is_catia_com_error(e) and source_intact and (
                target_existed_before or not Path(new_fp).exists()
            ):
                # 用户很可能在CATIA另存为对话框中点击了"取消"或"否"——
                # 视为主动跳过，不弹出错误对话框
                logger.info(
                    f"SaveAs skipped for {Path(fp).name} "
                    f"(user cancelled or declined overwrite in CATIA; exception: {e})"
                )
                return
            QMessageBox.warning(self, "另存为失败", f"文件操作失败：\n{e}")

    def _write_back(self, *, close_on_success: bool) -> None:
        """仅将已变更的字段写回CATIA。"""
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请选择一个CATProduct文件。")
                return

        # dirty_data 必须以 *当前* CATIA零件编号为键，
        # 该值可能与内部规范键（orig_pn）不同——当零件编号重命名
        # 已在上一次写回中完成时会出现此情况。
        # 保留 pn_remap 以便从 current_pn 反向追溯到 orig_pn，
        # 用于写回后更新快照。
        dirty_data: dict[str, dict[str, str]] = {}
        pn_remap:   dict[str, str]            = {}  # current_pn → orig_pn
        for pn, dirty_cols in self._modified_keys.items():
            if pn not in self._canonical_data:
                continue
            changed = {
                col: self._canonical_data[pn][col]
                for col in dirty_cols if col in self._canonical_data[pn]
            }
            if changed:
                # 使用快照中的零件编号（即CATIA当前保存的值）作为遍历查找键。
                # 若该零件编号从未被写回过，快照值等于 orig_pn。
                current_pn = self._snapshot_data.get(pn, {}).get(
                    "Part Number", pn
                )
                dirty_data[current_pn] = changed
                pn_remap[current_pn]   = pn

        if not dirty_data:
            if close_on_success:
                self.accept()
            else:
                QMessageBox.information(self, "无更改", "没有检测到任何修改，无需写回。")
            return

        self._save_btn.setEnabled(False)
        self._finish_btn.setEnabled(False)
        QApplication.processEvents()

        progress = QProgressDialog("正在写回CATIA，请稍候…", None, 0, 0, self)
        progress.setWindowTitle("写回CATIA")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)
        progress.setValue(0)

        def _on_node_written(count: int) -> None:
            progress.setLabelText(f"正在写回CATIA，请稍候… 已处理 {count} 个节点")
            progress.repaint()
            QApplication.processEvents()

        try:
            write_bom_to_catia(file_path, dirty_data, self._all_custom_columns,
                               _on_node_written)
        except Exception as e:
            progress.close()
            logger.error(f"Failed to write BOM back to CATIA: {e}")
            self._save_btn.setEnabled(True)
            self._finish_btn.setEnabled(True)
            QMessageBox.critical(
                self, "写回失败",
                f"写回CATIA时出错：\n{e}\n\n请确保CATIA已启动。",
            )
            return
        finally:
            progress.close()

        for current_pn, changed in dirty_data.items():
            pn = pn_remap.get(current_pn, current_pn)
            if pn in self._snapshot_data:
                self._snapshot_data[pn].update(changed)
            if pn in self._modified_keys:
                self._modified_keys[pn] -= set(changed.keys())
                if not self._modified_keys[pn]:
                    del self._modified_keys[pn]

        self._save_btn.setEnabled(True)
        self._finish_btn.setEnabled(True)

        if close_on_success:
            QMessageBox.information(
                self, "完成",
                "BOM属性已成功写回CATIA，请在CATIA中手动保存文件。",
            )
            self.accept()
        else:
            QMessageBox.information(
                self, "应用成功",
                "BOM属性已成功写回CATIA，请在CATIA中手动保存文件。",
            )

    # ── 导出表格 ──────────────────────────────────────────────────────────────

    def _export_table(self) -> None:
        """将当前显示的BOM表格导出为 Excel 或 CSV 文件。"""
        if not self._bom_loaded or not self._rows:
            QMessageBox.warning(self, "无数据", "请先加载BOM。")
            return

        # 若存在未写回的编辑，表格内容与CATIA不一致，导出前必须先写回
        if self._modified_keys:
            ret = QMessageBox.question(
                self, "存在未写回的修改",
                "检测到BOM属性尚未写回CATIA，导出前应保持表格与CATIA一致。\n\n"
                "是否立即将修改写回CATIA，再继续导出？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return
            self._write_back(close_on_success=False)
            # 若写回失败（modified_keys 仍非空），中止导出
            if self._modified_keys:
                return

        # 根据源文件建议默认文件名
        initial_name = ""
        if not self._use_active_chk.isChecked():
            fp_src = self._file_edit.text().strip()
            if fp_src:
                suffix_hint = "_汇总BOM" if self._summarize else "_BOM"
                initial_name = str(Path(fp_src).with_name(Path(fp_src).stem + suffix_hint))

        dest, selected_filter = QFileDialog.getSaveFileName(
            self,
            "导出BOM表格",
            initial_name,
            "Excel工作簿 (*.xlsx);;CSV文件 (*.csv)",
        )
        if not dest:
            return

        dest_path = Path(dest)
        # 根据扩展名推断格式；不明确时回退为 xlsx
        suffix = dest_path.suffix.lower()
        if suffix not in (".xlsx", ".csv"):
            dest_path = dest_path.with_suffix(".xlsx")
            suffix = ".xlsx"

        # 导出列：当前可见列，排除行号"#"列
        export_cols = [c for c in self._columns if c != BOM_ROW_NUMBER_COLUMN]

        # 从当前表格快照列宽（像素）
        col_px_widths: dict[str, int] = {}
        for col_idx, col_name in enumerate(self._columns):
            if col_name != BOM_ROW_NUMBER_COLUMN:
                col_px_widths[col_name] = self._table.columnWidth(col_idx)

        # 使用与 _populate_table 相同的取值逻辑收集行数据
        rows_data: list[dict] = []
        for row_data in self._rows:
            pn = str(row_data.get("Part Number", ""))
            row_out: dict = {}
            for col_name in export_cols:
                if col_name == "Source":
                    raw = str(row_data.get("Source", ""))
                    val = self._canonical_data.get(pn, {}).get(
                        "Source", SOURCE_TO_DISPLAY.get(raw, raw)
                    )
                elif col_name in PRESET_USER_REF_PROPERTY_OPTIONS:
                    val = self._canonical_data.get(pn, {}).get(
                        col_name, str(row_data.get(col_name, ""))
                    )
                elif col_name == "Filename":
                    fp_val = str(row_data.get("_filepath", ""))
                    fn_val = str(row_data.get("Filename", ""))
                    if self._show_filepath_col:
                        val = fp_val if fp_val else fn_val
                    else:
                        val = Path(fp_val).name if fp_val else fn_val
                elif col_name in BOM_READONLY_COLUMNS:
                    val = str(row_data.get(col_name, ""))
                else:
                    val = str(
                        self._canonical_data.get(pn, {}).get(
                            col_name, row_data.get(col_name, "")
                        )
                    )
                row_out[col_name] = val
            rows_data.append(row_out)

        try:
            if suffix == ".xlsx":
                self._write_xlsx(dest_path, export_cols, col_px_widths, rows_data)
            else:
                self._write_csv(dest_path, export_cols, rows_data)
        except PermissionError:
            QMessageBox.critical(
                self, "导出失败",
                f"无法写入文件（文件可能已在其他程序中打开）：\n{dest_path}",
            )
            return
        except Exception as e:
            logger.error(f"BOM table export failed: {e}")
            QMessageBox.critical(self, "导出失败", f"导出时出错：\n{e}")
            return

        QMessageBox.information(self, "导出成功", f"BOM已成功导出：\n{dest_path}")

    def _export_header(self, col_name: str) -> str:
        """返回列的显示表头字符串，与当前表格保持一致。"""
        if col_name == "Filename" and self._show_filepath_col:
            return "完整路径"
        return BOM_COLUMN_DISPLAY_NAMES.get(col_name, col_name)

    def _write_xlsx(
        self,
        dest: Path,
        cols: list[str],
        px_widths: dict[str, int],
        rows: list[dict],
    ) -> None:
        """将 *rows* 写入 *dest* 路径的 .xlsx 工作簿。"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总BOM" if self._summarize else "BOM"

        center      = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        thin_side   = Side(style="thin")
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        # 表头行
        for col_idx, col_name in enumerate(cols, start=1):
            cell        = ws.cell(row=1, column=col_idx, value=self._export_header(col_name))
            cell.font   = Font(bold=True)
            cell.fill   = header_fill
            cell.border = thin_border

        # 数据行
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(cols, start=1):
                raw_val = row.get(col_name, "")
                # 将数字存为整数，以便 Excel 排序/筛选
                if col_name in ("Level", "Quantity"):
                    try:
                        value = int(raw_val)
                    except (ValueError, TypeError):
                        logger.debug(
                            "Could not convert %r to int for column '%s'", raw_val, col_name
                        )
                        value = raw_val
                else:
                    value = raw_val
                cell        = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_name in ("Level", "Quantity", "Type"):
                    cell.alignment = center

        # 冻结表头行并启用自动筛选
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # 列宽：像素→Excel字符单位换算
        # Calibri 11pt 默认字体约为每字符7像素，作为近似换算基准
        PX_PER_CHAR = 7.0
        for col_idx, col_name in enumerate(cols, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            px = px_widths.get(col_name, 80)
            char_width = max(8.0, px / PX_PER_CHAR)
            ws.column_dimensions[col_letter].width = round(char_width, 1)

        wb.save(str(dest))
        logger.info(f"BOM table exported (xlsx) -> {dest}")

    def _write_csv(
        self,
        dest: Path,
        cols: list[str],
        rows: list[dict],
    ) -> None:
        """将 *rows* 写入 *dest* 路径的 UTF-8 CSV 文件（带BOM头）。"""
        import csv

        with open(dest, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([self._export_header(c) for c in cols])
            for row in rows:
                writer.writerow([row.get(c, "") for c in cols])
        logger.info(f"BOM table exported (csv) -> {dest}")

    def _apply_changes(self) -> None:
        """将修改写回CATIA，保持对话框不关闭。"""
        self._write_back(close_on_success=False)
    def _finish_and_close(self) -> None:
        """将修改写回CATIA，然后关闭对话框。"""
        self._write_back(close_on_success=True)

    # ── 右键上下文菜单 ────────────────────────────────────────────────────────

    def _on_tree_context_menu(self, pos) -> None:
        """显示右键点击的BOM行对应的上下文菜单。

        若关联文件内嵌了缩略图，则在菜单顶部以非交互式图片控件展示。
        """
        item = self._table.itemAt(pos)
        if item is None:
            return
        row_idx = item.data(0, Qt.ItemDataRole.UserRole)
        if row_idx is None:
            return

        row_data     = self._rows[row_idx]
        fp           = str(row_data.get("_filepath", ""))
        fp_path      = Path(fp) if fp else None
        is_component = row_data.get("Type") == "部件"
        not_found    = bool(row_data.get("_not_found"))
        unreadable   = bool(row_data.get("_unreadable"))
        pn           = str(row_data.get("Part Number", ""))

        # 确保右键点击的行已被选中，以便下游方法（_rename_selected_file 等）能找到它
        if not item.isSelected():
            self._table.clearSelection()
            item.setSelected(True)

        menu = QMenu(self)

        # ── 嵌入缩略图（可用时在菜单顶部内联显示）─────────────────────────────
        # 以下情况跳过缩略图提取：
        #   • 部件：文件路径属于父产品，而非此组件
        #   • not_found：CATIA无法解析该文件
        #   • 文件在磁盘上不存在（未保存或丢失）
        if fp and not is_component and not not_found and fp_path is not None and fp_path.exists():
            img_bytes = read_catia_thumbnail(fp)
            if img_bytes:
                pixmap = QPixmap()
                loaded = pixmap.loadFromData(img_bytes)
                if loaded and not pixmap.isNull():
                    if (pixmap.width() > BOM_THUMBNAIL_MAX_SIZE
                            or pixmap.height() > BOM_THUMBNAIL_MAX_SIZE):
                        pixmap = pixmap.scaled(
                            BOM_THUMBNAIL_MAX_SIZE,
                            BOM_THUMBNAIL_MAX_SIZE,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation,
                        )
                    thumb_label = QLabel()
                    thumb_label.setPixmap(pixmap)
                    thumb_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    thumb_label.setContentsMargins(6, 4, 6, 4)
                    thumb_action = QWidgetAction(menu)
                    thumb_action.setDefaultWidget(thumb_label)
                    menu.addAction(thumb_action)
                    menu.addSeparator()

        # ── 打开路径 ──────────────────────────────────────────────────────────
        act_open_path = menu.addAction("打开路径")
        path_available = bool(fp) and fp_path is not None and (
            fp_path.exists() or fp_path.parent.exists()
        )
        act_open_path.setEnabled(path_available)

        # ── 复制路径 ──────────────────────────────────────────────────────────
        act_copy_path = menu.addAction("复制路径")
        act_copy_path.setEnabled(bool(fp))

        # ── 在CATIA中打开 ─────────────────────────────────────────────────────
        # 仅当文件在磁盘上存在且不是损坏/轻量化引用时启用。
        # 部件行共享父产品的文件路径，因此也排除在外。
        act_open_catia = menu.addAction("在CATIA中打开")
        catia_available = (
            not is_component and not not_found and not unreadable
            and fp_path is not None and fp_path.exists()
        )
        act_open_catia.setEnabled(catia_available)

        menu.addSeparator()

        # ── 另存为 ────────────────────────────────────────────────────────────
        act_edit_path = menu.addAction("另存为")
        # 允许对未保存过的零件（文件不在磁盘上但在CATIA内存中）执行另存为；
        # 仅排除没有路径或CATIA无法找到的节点。
        act_edit_path.setEnabled(bool(fp) and not is_component and not not_found)

        action = menu.exec(self._table.viewport().mapToGlobal(pos))

        if action == act_open_path:
            self._open_path(fp)
        elif action == act_copy_path:
            QApplication.clipboard().setText(fp)
        elif action == act_open_catia:
            self._open_in_catia(fp)
        elif action == act_edit_path:
            self._rename_selected_file()

    def _open_path(self, fp: str) -> None:
        """在 Windows 资源管理器中打开包含 *fp* 的文件夹，并高亮选中该文件。"""
        p = Path(fp).resolve()
        try:
            if p.exists():
                # 对路径加引号，以确保 Explorer 能正确处理含空格的目录名
                subprocess.Popen(f'explorer /select,"{p}"', shell=True)
            elif p.parent.exists():
                subprocess.Popen(f'explorer "{p.parent}"', shell=True)
        except Exception as exc:
            logger.warning(f"Failed to open path in Explorer: {exc}")

    def _open_in_catia(self, fp: str) -> None:
        """通过 ``documents.open`` 在CATIA中打开 *fp* 指向的文档。

        打开后，若 ``win32gui`` 可用，则将CATIA V5主窗口置于Windows前台。
        """
        try:
            from pycatia import catia as _pycatia  # noqa: PLC0415
            caa         = _pycatia()
            application = caa.application
            application.visible = True
            documents   = application.documents

            fp_resolved = Path(fp).resolve()
            documents.open(str(fp_resolved))

            # ── 将CATIA V5主窗口置于Windows前台 ──────────────────────────────
            try:
                import win32gui  # noqa: PLC0415
                import win32con  # noqa: PLC0415

                def _raise_catia_window(hwnd, _extra):
                    if not win32gui.IsWindowVisible(hwnd):
                        return
                    title = win32gui.GetWindowText(hwnd)
                    # 仅匹配CATIA V5应用程序主窗口，排除其他含"CATIA"字样的窗口
                    if title.startswith("CATIA V5"):
                        try:
                            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                            win32gui.SetForegroundWindow(hwnd)
                        except Exception:
                            pass
                        # 找到第一个CATIA V5窗口后停止枚举
                        return False

                win32gui.EnumWindows(_raise_catia_window, None)
            except ImportError:
                pass
            except Exception:
                pass

        except Exception as e:
            QMessageBox.warning(self, "在CATIA中打开失败", f"无法在CATIA中打开文件：\n{e}")
