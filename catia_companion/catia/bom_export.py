"""
BOM Excel export.

Provides:
- export_bom_to_excel() – export a hierarchical or summarised BOM to an .xlsx file
"""

import logging
from collections.abc import Callable
from pathlib import Path

from catia_companion.constants import (
    BOM_DEFAULT_COLUMNS,
    BOM_COLUMN_DISPLAY_NAMES,
    BOM_COLUMN_MIN_WIDTHS,
    SOURCE_TO_DISPLAY,
)
from catia_companion.utils import estimate_column_width
from catia_companion.catia.bom_collect import collect_bom_rows, flatten_bom_to_summary

logger = logging.getLogger(__name__)


def export_bom_to_excel(
    file_paths: list[str | None],
    output_folder: str | None = None,
    columns: list[str] | None = None,
    custom_columns: list[str] | None = None,
    row_progress_callback: Callable[[int], None] | None = None,
    summarize: bool = False,
    summary_include_assemblies: bool = False,
    summary_sort_column: str | None = None,
) -> None:
    """Export a hierarchical or summarised BOM from CATProduct files to Excel (.xlsx).

    Parameters
    ----------
    file_paths:
        Paths to ``.CATProduct`` files.  A ``None`` entry means "use the
        currently active CATIA document" (no file is opened or closed).
    output_folder:
        Destination directory.  Defaults to each source file's parent.
    columns:
        Internal column names to include.  Defaults to
        :data:`~catia_companion.constants.BOM_DEFAULT_COLUMNS`.
    custom_columns:
        Column names that are user-defined properties.
    row_progress_callback:
        Optional callable invoked as ``row_progress_callback(count)`` with the
        running node count after each row is collected.  Matches the signature
        of the BOM-load progress callback so both operations can share UI code.
    summarize:
        When ``True`` the hierarchical BOM is collapsed into a flat summary
        (unique parts with cumulative quantities) before writing to Excel.
        The output filename will have the suffix ``_BOM汇总`` instead of
        ``_BOM``.
    summary_include_assemblies:
        Passed to :func:`~catia_companion.catia.bom_collect.flatten_bom_to_summary`.
        When ``True`` sub-assemblies and assemblies are included in the summary.
        Only used when *summarize* is ``True``.
    summary_sort_column:
        Column name to sort the summary by.  Defaults to ``"Part Number"``
        when ``None``.  Only used when *summarize* is ``True``.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from pycatia import catia
    from pycatia.product_structure_interfaces.product_document import ProductDocument

    if columns is None:
        columns = BOM_DEFAULT_COLUMNS
    if custom_columns is None:
        custom_columns = []

    # In summary mode the "Level" column carries no useful information.
    if summarize:
        columns = [c for c in columns if c != "Level"]

    bom_suffix = "_BOM汇总" if summarize else "_BOM"

    caa         = catia()
    application = caa.application
    application.visible = True
    documents   = application.documents

    def _write_sheet(ws, rows: list[dict]) -> None:
        center = Alignment(horizontal="center")

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell       = ws.cell(row=1, column=col_idx,
                                 value=BOM_COLUMN_DISPLAY_NAMES.get(col_name, col_name))
            cell.font  = Font(bold=True)

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            level = row.get("Level", 0)
            for col_idx, col_name in enumerate(columns, start=1):
                if col_name == "Level":
                    value = level
                elif col_name == "Quantity":
                    value = row.get("Quantity", 1)
                elif col_name == "Type":
                    value = row.get("Type", "")
                elif col_name == "Source":
                    raw   = str(row.get("Source", ""))
                    value = SOURCE_TO_DISPLAY.get(raw, raw)
                else:
                    value = row.get(col_name, "")

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_name in ("Level", "Quantity", "Type"):
                    cell.alignment = center

        # Auto-width columns
        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            header     = BOM_COLUMN_DISPLAY_NAMES.get(col_name, col_name)
            min_w      = BOM_COLUMN_MIN_WIDTHS.get(col_name, 10)
            max_width  = max(estimate_column_width(header), min_w)
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_width = max(max_width, estimate_column_width(str(cell_val)))
            ws.column_dimensions[col_letter].width = max_width + 2

    total_files = len(file_paths)
    for file_idx, path in enumerate(file_paths, start=1):
        if path is None:
            # Use the active document without opening or closing
            try:
                active_full = application.active_document.full_name
            except Exception as e:
                raise RuntimeError(
                    "无法获取当前CATIA活动文档，请确保CATIA已打开CATProduct。"
                ) from e
            src_name = Path(active_full)
            dest_dir = Path(output_folder).resolve() if output_folder else src_name.parent
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / f"{src_name.stem}{bom_suffix}.xlsx"

            rows = collect_bom_rows(None, columns, custom_columns,
                                     row_progress_callback)
            if summarize:
                rows = flatten_bom_to_summary(
                    rows,
                    include_assemblies=summary_include_assemblies,
                    sort_column=summary_sort_column,
                )
            wb   = openpyxl.Workbook()
            ws.title = "BOM汇总" if summarize else "BOM"
            _write_sheet(ws, rows)
            wb.save(str(dest))
            logger.info(f"  BOM exported -> {dest}")
            logger.info("Done: active document\n")
            continue

        src      = Path(path).resolve()
        dest_dir = Path(output_folder).resolve() if output_folder else src.parent
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest     = dest_dir / f"{src.stem}{bom_suffix}.xlsx"

        if dest.exists():
            try:
                with open(dest, "a+b"):
                    pass
            except PermissionError:
                from PySide6.QtWidgets import QMessageBox
                reply = QMessageBox.question(
                    None, "文件正在使用",
                    f"该文件当前在Excel中已打开：\n{dest}\n\n"
                    "请在Excel中关闭该文件，然后点击【重试】，或点击【取消】以中止。",
                    QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel,
                )
                if reply == QMessageBox.StandardButton.Cancel:
                    continue
                try:
                    with open(dest, "a+b"):
                        pass
                except PermissionError:
                    QMessageBox.critical(
                        None, "文件仍在使用中",
                        f"文件仍处于打开状态，请关闭后重试。\n{dest}",
                    )
                    continue

        # Track already-open documents to avoid closing files we did not open
        already_open: set[Path] = set()
        for i in range(1, documents.count + 1):
            try:
                already_open.add(Path(documents.item(i).full_name).resolve())
            except Exception:
                pass

        logger.info(f"Opening: {src}")
        rows = collect_bom_rows(str(src), columns, custom_columns,
                                row_progress_callback)
        if summarize:
            rows = flatten_bom_to_summary(
                rows,
                include_assemblies=summary_include_assemblies,
                sort_column=summary_sort_column,
            )

        wb       = openpyxl.Workbook()
        ws       = wb.active
        ws.title = "BOM汇总" if summarize else "BOM"
        _write_sheet(ws, rows)
        wb.save(str(dest))
        logger.info(f"  BOM exported -> {dest}")

        # Close the document only if we were the one who opened it
        if src not in already_open:
            for i in range(1, documents.count + 1):
                try:
                    doc = documents.item(i)
                    if Path(doc.full_name).resolve() == src:
                        ProductDocument(doc.com_object).close()
                        break
                except Exception:
                    pass

        logger.info(f"Done: {src.name}\n")
