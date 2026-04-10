import sys
import os
import re
import copy
import shutil
import subprocess
import unicodedata
import winreg
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QMessageBox, QDialog, QPushButton, QListWidget, QFileDialog,
    QAbstractItemView, QRadioButton, QButtonGroup, QLineEdit, QGroupBox,
    QListWidgetItem, QComboBox, QCheckBox, QPlainTextEdit,
    QTableWidget, QTableWidgetItem, QHeaderView,
)
from PySide6.QtGui import QAction, QColor
from PySide6.QtCore import Qt, QSettings, QObject, Signal

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

_log_dir = Path.home() / "CATIA_Companion" / "logs"
_log_dir.mkdir(parents=True, exist_ok=True)
_log_file = _log_dir / "catia_companion.log"

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        RotatingFileHandler(_log_file, maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
)
logger = logging.getLogger(__name__)


class _LogSignalEmitter(QObject):
    message_logged = Signal(str)


_log_emitter = _LogSignalEmitter()


class QtLogHandler(logging.Handler):
    def emit(self, record: logging.LogRecord):
        msg = self.format(record)
        _log_emitter.message_logged.emit(msg)


_qt_log_handler = QtLogHandler()
_qt_log_handler.setFormatter(logging.Formatter(
    "%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
))
logging.getLogger().addHandler(_qt_log_handler)


def resource_path(filename: str) -> Path:
    if hasattr(sys, "_MEIPASS"):
        return Path(sys.executable).parent / filename
    return Path(__file__).parent / filename


# ---------------------------------------------------------------------------
# App info
# ---------------------------------------------------------------------------

APP_NAME    = "CATIA Companion"
APP_VERSION = "1.0.0"
APP_DATE    = "2026-04-03"
APP_AUTHOR  = "CHEN Weibo"
APP_CONTACT = "thucwb@gmail.com"

ABOUT_TEXT = f"""{APP_NAME} v{APP_VERSION}

A CATIA V5 productivity tool for engineering teams.
Automates drawing conversion, part export, and
installation of CATIA resources.

─────────────────────────────────────────
Developer   {APP_AUTHOR}
Contact     {APP_CONTACT}
Released    {APP_DATE}
─────────────────────────────────────────

\u00a9 2026 {APP_AUTHOR}. For internal use only."""

# ---------------------------------------------------------------------------
# Part template properties
# ---------------------------------------------------------------------------

PART_TEMPLATE_PROPERTIES = ["物料编码", "物料名称", "规格型号", "物料来源", "数据状态", "存货类别", "质量", "备注"]


# ---------------------------------------------------------------------------
# Default window geometry
# ---------------------------------------------------------------------------

DEFAULT_WIDTH  = 260
DEFAULT_HEIGHT = 500


# ---------------------------------------------------------------------------
# Log Window
# ---------------------------------------------------------------------------

class LogWindow(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent, Qt.WindowType.Window)
        self.setWindowTitle("CATIA Companion – Log")
        self.resize(600, 400)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        self._log_view = QPlainTextEdit()
        self._log_view.setReadOnly(True)
        self._log_view.setStyleSheet(
            "background-color: #1e1e1e; color: #d4d4d4;"
            " font-family: Consolas, 'Courier New', monospace; font-size: 9pt;"
        )
        layout.addWidget(self._log_view)

        open_log_btn = QPushButton("打开日志文件")
        open_log_btn.clicked.connect(self._open_log_file)
        layout.addWidget(open_log_btn)

        log_path_label = QLabel(f"Log: {_log_file}")
        log_path_label.setStyleSheet("color: gray; font-size: 9pt;")
        log_path_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        layout.addWidget(log_path_label)

    def _append_log(self, message: str):
        self._log_view.appendPlainText(message)
        self._log_view.verticalScrollBar().setValue(
            self._log_view.verticalScrollBar().maximum()
        )

    def _open_log_file(self):
        try:
            if sys.platform == "win32":
                os.startfile(_log_file)
            else:
                subprocess.Popen(
                    ["xdg-open", str(_log_file)],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
        except Exception as e:
            QMessageBox.warning(self, "无法打开日志文件",
                f"无法打开日志文件：\n{_log_file}\n\n{e}")

    def closeEvent(self, event):
        # Hide instead of destroying, so log history is preserved
        event.ignore()
        self.hide()
        # Uncheck the menu action in the main window
        parent = self.parent()
        if parent and hasattr(parent, "_show_log_action"):
            parent._show_log_action.setChecked(False)


# ---------------------------------------------------------------------------
# Main Window
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CATIA Companion")
        self.resize(DEFAULT_WIDTH, DEFAULT_HEIGHT)
        self._setup_menu_bar()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        label = QLabel("欢迎使用 CATIA Companion")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label)

        # Separate log window (hidden by default)
        self._log_window = LogWindow(self)
        _log_emitter.message_logged.connect(self._log_window._append_log)

        self.statusBar().showMessage("就绪")

    def _toggle_log_window(self, checked: bool):
        if checked:
            self._log_window.show()
            self._log_window.raise_()
        else:
            self._log_window.hide()

    def _setup_menu_bar(self):
        menu_bar = self.menuBar()

        # --- File ---
        file_menu = menu_bar.addMenu("文件")
        _stub = lambda: QMessageBox.information(self, "提示", "功能尚未实现")
        new_action = QAction("新建", self)
        new_action.triggered.connect(_stub)
        file_menu.addAction(new_action)
        open_action = QAction("打开...", self)
        open_action.triggered.connect(_stub)
        file_menu.addAction(open_action)
        save_action = QAction("保存", self)
        save_action.triggered.connect(_stub)
        file_menu.addAction(save_action)
        save_as_action = QAction("另存为...", self)
        save_as_action.triggered.connect(_stub)
        file_menu.addAction(save_as_action)
        file_menu.addSeparator()
        quit_action = QAction("退出", self)
        quit_action.triggered.connect(self.close)
        file_menu.addAction(quit_action)

        # --- Export ---
        export_menu = menu_bar.addMenu("导出")
        convert_drawing_action = QAction("从CATDrawing导出pdf", self)
        convert_drawing_action.triggered.connect(self._open_convert_drawing_dialog)
        export_menu.addAction(convert_drawing_action)
        convert_part_action = QAction("从CATPart/CATProduct导出stp", self)
        convert_part_action.triggered.connect(self._open_convert_part_dialog)
        export_menu.addAction(convert_part_action)
        export_bom_action = QAction("从CATProduct导出BOM", self)
        export_bom_action.triggered.connect(self._open_export_bom_dialog)
        export_menu.addAction(export_bom_action)

        # --- Edit ---
        edit_menu = menu_bar.addMenu("编辑")
        bom_edit_action = QAction("BOM属性补全", self)
        bom_edit_action.triggered.connect(self._open_bom_edit_dialog)
        edit_menu.addAction(bom_edit_action)

        # --- Macro ---
        self._macro_menu = menu_bar.addMenu("宏")
        self._build_macro_menu()

        # --- Tools ---
        tools_menu = menu_bar.addMenu("工具")
        copy_font_action = QAction("复制字体文件到CATIA目录", self)
        copy_font_action.triggered.connect(self._copy_font_to_catia)
        tools_menu.addAction(copy_font_action)
        copy_iso_action = QAction("复制ISO.xml到CATIA目录", self)
        copy_iso_action.triggered.connect(self._copy_iso_to_catia)
        tools_menu.addAction(copy_iso_action)
        pojie_action = QAction("PoJie", self)
        pojie_action.triggered.connect(self._pojie)
        tools_menu.addAction(pojie_action)
        stamp_action = QAction("刷写零件模板", self)
        stamp_action.triggered.connect(self._open_stamp_part_template_dialog)
        tools_menu.addAction(stamp_action)
        find_deps_action = QAction("查找所有依赖项", self)
        find_deps_action.triggered.connect(self._open_find_dependencies_dialog)
        tools_menu.addAction(find_deps_action)

        # --- View ---
        view_menu = menu_bar.addMenu("视图")
        zoom_in_action = QAction("放大", self)
        zoom_in_action.triggered.connect(lambda: QMessageBox.information(self, "提示", "功能尚未实现"))
        view_menu.addAction(zoom_in_action)
        zoom_out_action = QAction("缩小", self)
        zoom_out_action.triggered.connect(lambda: QMessageBox.information(self, "提示", "功能尚未实现"))
        view_menu.addAction(zoom_out_action)
        zoom_reset_action = QAction("重置缩放", self)
        zoom_reset_action.triggered.connect(lambda: self.resize(DEFAULT_WIDTH, DEFAULT_HEIGHT))
        view_menu.addAction(zoom_reset_action)
        view_menu.addSeparator()
        self._show_log_action = QAction("显示Log", self)
        self._show_log_action.setCheckable(True)
        self._show_log_action.toggled.connect(self._toggle_log_window)
        view_menu.addAction(self._show_log_action)

        # --- Help ---
        help_menu = menu_bar.addMenu("帮助")
        doc_action = QAction("文档", self)
        doc_action.triggered.connect(lambda: QMessageBox.information(self, "提示", "功能尚未实现"))
        help_menu.addAction(doc_action)
        about_action = QAction("关于 CATIA Companion", self)
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)

    def _show_about(self):
        QMessageBox.about(self, f"About {APP_NAME}", ABOUT_TEXT)

    # ------------------------------------------------------------------
    # Macro menu helpers
    # ------------------------------------------------------------------

    _MACRO_EXTENSIONS = {".catvbs", ".catscript"}

    def _macros_dir(self) -> Path:
        return resource_path("macros")

    def _build_macro_menu(self):
        """Populate (or repopulate) the 宏 menu from the macros/ folder."""
        self._macro_menu.clear()

        macros_dir = self._macros_dir()
        macro_files: list[Path] = []
        if macros_dir.is_dir():
            macro_files = sorted(
                f for f in macros_dir.iterdir()
                if f.is_file() and f.suffix.lower() in self._MACRO_EXTENSIONS
            )

        if macro_files:
            for macro_path in macro_files:
                action = QAction(macro_path.name, self)
                action.triggered.connect(lambda checked=False, p=macro_path: self._run_macro(p))
                self._macro_menu.addAction(action)
            self._macro_menu.addSeparator()
        else:
            placeholder = QAction("（未找到宏文件）", self)
            placeholder.setEnabled(False)
            self._macro_menu.addAction(placeholder)
            self._macro_menu.addSeparator()

        open_folder_action = QAction("打开宏文件夹", self)
        open_folder_action.triggered.connect(self._open_macros_folder)
        self._macro_menu.addAction(open_folder_action)

        refresh_action = QAction("刷新宏列表", self)
        refresh_action.triggered.connect(self._build_macro_menu)
        self._macro_menu.addAction(refresh_action)

    def _open_macros_folder(self):
        macros_dir = self._macros_dir()
        macros_dir.mkdir(parents=True, exist_ok=True)
        try:
            if sys.platform == "win32":
                os.startfile(str(macros_dir))
            else:
                subprocess.Popen(
                    ["xdg-open", str(macros_dir)],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
        except Exception as e:
            QMessageBox.warning(self, "无法打开文件夹", f"无法打开宏文件夹：\n{macros_dir}\n\n{e}")

    def _run_macro(self, macro_path: Path):
        """Run a CATIA macro file (.catvbs / .CATScript) via CATIA COM."""
        if not macro_path.exists():
            QMessageBox.warning(self, "文件不存在", f"宏文件不存在：\n{macro_path}")
            return
        try:
            from pycatia import catia as _catia
            caa = _catia()
            app = caa.application
            # catScriptLibraryTypeDirectory = 1
            app.com_object.SystemService.ExecuteScript(
                str(macro_path.parent), 1, macro_path.name, "CATMain", []
            )
            logger.info(f"Macro executed: {macro_path.name}")
        except Exception as e:
            logger.error(f"Failed to run macro {macro_path.name}: {e}")
            QMessageBox.critical(self, "宏执行失败",
                f"运行宏时出错：\n{macro_path.name}\n\n{e}\n\n请确保CATIA已启动。")

    def _open_convert_part_dialog(self):
        dialog = ConvertDialog(
            parent=self,
            title="将CATPart/CATProduct导出为STP",
            file_label="已选CATPart/CATProduct文件:",
            file_filter="*.CATPart *.CATProduct (*.CATPart *.CATProduct);;All Files (*)",
            no_files_msg="请至少选择一个CATPart或CATProduct文件。",
            conversion_fn=CATPart_to_STP,
            settings_key="CATPart",
            show_prefix_option=True,
            prefix="MD_",
            note="暂时留空"
        )
        dialog.exec()

    def _open_convert_drawing_dialog(self):
        dialog = ConvertDialog(
            parent=self,
            title="将CATDrawing导出为PDF",
            file_label="已选CATDrawing文件:",
            file_filter="*.CATDrawing (*.CATDrawing);;All Files (*)",
            no_files_msg="请至少选择一个CATDrawing文件。",
            conversion_fn=CATDrawing_to_PDF,
            settings_key="CATDrawing",
            show_prefix_option=True,
            prefix="DR_",
            note="如果用于导出的CATDrawing有多页，请将CATIA设置为\u201c将多页文档保存在单向量文件中\u201d（工具->选项->常规->兼容性->图形格式->导出（另存为））"
        )
        dialog.exec()

    def _open_export_bom_dialog(self):
        dialog = ExportBOMDialog(self)
        dialog.exec()

    def _open_bom_edit_dialog(self):
        dialog = BomEditDialog(self)
        dialog.exec()

    def _copy_font_to_catia(self):
        self._copy_file_to_catia(
            file_name="Changfangsong.ttf",
            relative_dest=Path("win_b64") / "resources" / "fonts" / "TrueType"
        )

    def _copy_iso_to_catia(self):
        self._copy_file_to_catia(
            file_name="ISO.xml",
            relative_dest=Path("win_b64") / "resources" / "standard" / "drafting"
        )

    def _copy_file_to_catia(self, file_name: str, relative_dest: Path):
        src_file = resource_path(file_name)
        if not src_file.exists():
            QMessageBox.warning(self, "文件未找到",
                f"在工作目录中找不到 '{file_name}'：\n{src_file.parent}")
            return

        catia_root = detect_catia_root()
        if catia_root:
            reply = QMessageBox.question(self, "检测到CATIA安装",
                f"检测到CATIA安装路径：\n{catia_root}\n\n是否使用该目录？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                catia_root = None

        if not catia_root:
            catia_root = QFileDialog.getExistingDirectory(self,
                "选择CATIA安装目录（例如 C:\\Program Files\\Dassault Systemes\\B28）", "")
            if not catia_root:
                return

        dest_dir = Path(catia_root) / relative_dest
        if not dest_dir.exists():
            reply = QMessageBox.question(self, "文件夹未找到",
                f"目标文件夹不存在：\n{dest_dir}\n\n是否要创建该文件夹？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                dest_dir.mkdir(parents=True, exist_ok=True)
            else:
                return

        dest_file = dest_dir / file_name
        try:
            shutil.copy2(str(src_file), str(dest_file))
            QMessageBox.information(self, "成功",
                f"'{file_name}' 已成功复制到：\n{dest_file}")
        except PermissionError:
            QMessageBox.critical(self, "权限不足",
                f"无法复制文件，请以管理员身份运行程序。\n\n目标路径：\n{dest_file}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"发生意外错误：\n{e}")

    def _pojie(self):
        src_dir = resource_path("Pojie")
        if not src_dir.exists() or not src_dir.is_dir():
            QMessageBox.warning(self, "文件夹未找到",
                f"找不到 'Pojie' 文件夹：\n{src_dir.parent}")
            return

        catia_root = detect_catia_root()
        if catia_root:
            reply = QMessageBox.question(self, "检测到CATIA安装",
                f"检测到CATIA安装路径：\n{catia_root}\n\n是否使用该目录？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.No:
                catia_root = None

        if not catia_root:
            catia_root = QFileDialog.getExistingDirectory(self,
                "选择CATIA安装目录（例如 C:\\Program Files\\Dassault Systemes\\B28）", "")
            if not catia_root:
                return

        dest_dir = Path(catia_root) / "win_b64" / "code" / "bin"
        if not dest_dir.exists():
            QMessageBox.critical(self, "文件夹未找到",
                f"目标文件夹不存在：\n{dest_dir}\n\n请检查您的CATIA安装。")
            return

        files = [f for f in src_dir.iterdir() if f.is_file()]
        if not files:
            QMessageBox.warning(self, "文件夹为空", "'Pojie' 文件夹中没有文件。")
            return

        try:
            copied = []
            for src_file in files:
                dest_file = dest_dir / src_file.name
                shutil.copy2(str(src_file), str(dest_file))
                copied.append(src_file.name)
                logger.info(f"  Copied: {src_file.name} -> {dest_file}")
            QMessageBox.information(self, "成功",
                f"已成功复制 {len(copied)} 个文件到：\n{dest_dir}\n\n" + "\n".join(copied))
        except PermissionError:
            QMessageBox.critical(self, "权限不足",
                f"无法复制文件，请以管理员身份运行程序。\n\n目标路径：\n{dest_dir}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"发生意外错误：\n{e}")

    def _open_stamp_part_template_dialog(self):
        dialog = ConvertDialog(
            parent=self,
            title="刷写零件模板",
            file_label="已选CATPart文件:",
            file_filter="*.CATPart (*.CATPart);;All Files (*)",
            no_files_msg="请至少选择一个CATPart文件。",
            conversion_fn=stamp_part_template,
            settings_key="StampPartTemplate"
        )
        dialog.exec()

    def _open_find_dependencies_dialog(self):
        dialog = FindDependenciesDialog(self)
        dialog.exec()


# ---------------------------------------------------------------------------
# Generic Convert Dialog
# ---------------------------------------------------------------------------

class ConvertDialog(QDialog):
    def __init__(self, parent=None, title="Convert", file_label="Selected files:",
                 file_filter="All Files (*)", no_files_msg="Please select at least one file.",
                 conversion_fn=None, settings_key="default",
                 show_prefix_option=False, prefix="", note: str = ""):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumSize(520, 450)
        self._file_filter        = file_filter
        self._no_files_msg       = no_files_msg
        self._conversion_fn      = conversion_fn
        self._show_prefix_option = show_prefix_option
        self._prefix             = prefix

        self._settings = QSettings("CATIACompanion", f"ConvertDialog_{settings_key}")
        self._last_browse_dir = self._settings.value("last_browse_dir", "")
        self._last_output_dir = self._settings.value("last_output_dir", "")

        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        layout.addWidget(QLabel(file_label))
        self.file_list = QListWidget()
        self.file_list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)

        # Restore previously saved file list
        saved_files: list = self._settings.value("saved_files", []) or []
        if isinstance(saved_files, str):
            saved_files = [saved_files]
        for f in saved_files:
            if Path(f).exists():
                self.file_list.addItem(f)

        layout.addWidget(self.file_list)

        btn_row = QHBoxLayout()
        browse_btn = QPushButton("浏览...")
        browse_btn.clicked.connect(self._browse_files)
        remove_btn = QPushButton("移除所选")
        remove_btn.clicked.connect(self._remove_selected)
        remove_all_btn = QPushButton("全部移除")
        remove_all_btn.clicked.connect(self._remove_all)
        btn_row.addWidget(browse_btn)
        btn_row.addWidget(remove_btn)
        btn_row.addWidget(remove_all_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        # Output folder — hidden for stamp dialog
        if settings_key != "StampPartTemplate":
            output_group = QGroupBox("输出文件夹")
            output_layout = QVBoxLayout(output_group)
            self.radio_same = QRadioButton("与源文件相同目录")
            self.radio_custom = QRadioButton("自定义目录:")
            self.radio_same.setChecked(True)
            self.btn_group = QButtonGroup(self)
            self.btn_group.addButton(self.radio_same)
            self.btn_group.addButton(self.radio_custom)
            output_layout.addWidget(self.radio_same)
            output_layout.addWidget(self.radio_custom)
            folder_row = QHBoxLayout()
            self.folder_edit = QLineEdit()
            self.folder_edit.setPlaceholderText("选择输出文件夹...")
            self.folder_edit.setReadOnly(True)
            self.folder_edit.setEnabled(False)
            self.folder_browse_btn = QPushButton("浏览...")
            self.folder_browse_btn.setEnabled(False)
            self.folder_browse_btn.clicked.connect(self._browse_output_folder)
            folder_row.addWidget(self.folder_edit)
            folder_row.addWidget(self.folder_browse_btn)
            output_layout.addLayout(folder_row)
            self.radio_custom.toggled.connect(self._toggle_folder_row)
            layout.addWidget(output_group)
            if self._last_output_dir:
                self.radio_custom.setChecked(True)
                self.folder_edit.setText(self._last_output_dir)
        else:
            self.radio_same  = None
            self.folder_edit = None

        # Prefix and suffix rows — shown when show_prefix_option=True
        if show_prefix_option:
            saved_add_prefix = self._settings.value("add_prefix", True)
            if isinstance(saved_add_prefix, str):
                saved_add_prefix = saved_add_prefix.lower() == "true"
            saved_prefix_value = self._settings.value("prefix_value", prefix)

            prefix_row = QHBoxLayout()
            self.prefix_checkbox = QCheckBox("添加前缀:")
            self.prefix_checkbox.setChecked(saved_add_prefix)
            self.prefix_edit = QLineEdit(saved_prefix_value)
            self.prefix_edit.setEnabled(saved_add_prefix)
            self.prefix_checkbox.toggled.connect(self.prefix_edit.setEnabled)
            prefix_row.addWidget(self.prefix_checkbox)
            prefix_row.addWidget(self.prefix_edit)
            layout.addLayout(prefix_row)

            saved_add_suffix = self._settings.value("add_suffix", False)
            if isinstance(saved_add_suffix, str):
                saved_add_suffix = saved_add_suffix.lower() == "true"
            saved_suffix_value = self._settings.value("suffix_value", "")

            suffix_row = QHBoxLayout()
            self.suffix_checkbox = QCheckBox("添加后缀:")
            self.suffix_checkbox.setChecked(saved_add_suffix)
            self.suffix_edit = QLineEdit(saved_suffix_value)
            self.suffix_edit.setEnabled(saved_add_suffix)
            self.suffix_checkbox.toggled.connect(self.suffix_edit.setEnabled)
            suffix_row.addWidget(self.suffix_checkbox)
            suffix_row.addWidget(self.suffix_edit)
            layout.addLayout(suffix_row)
        else:
            self.prefix_checkbox = None
            self.prefix_edit     = None
            self.suffix_checkbox = None
            self.suffix_edit     = None

        if note:
            note_label = QLabel(note)
            note_label.setWordWrap(True)
            note_label.setStyleSheet("color: gray; font-size: 11px;")
            layout.addWidget(note_label)

        action_row = QHBoxLayout()
        action_row.addStretch()
        confirm_btn = QPushButton("确认")
        confirm_btn.setDefault(True)
        confirm_btn.clicked.connect(self._confirm)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        action_row.addWidget(confirm_btn)
        action_row.addWidget(cancel_btn)
        layout.addLayout(action_row)

    def _toggle_folder_row(self, checked):
        self.folder_edit.setEnabled(checked)
        self.folder_browse_btn.setEnabled(checked)

    def _browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择文件", self._last_browse_dir, self._file_filter)
        if files:
            self._last_browse_dir = str(Path(files[0]).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)
        for f in files:
            existing = [self.file_list.item(i).text() for i in range(self.file_list.count())]
            if f not in existing:
                self.file_list.addItem(f)
        self._persist_file_list()

    def _remove_selected(self):
        for item in self.file_list.selectedItems():
            self.file_list.takeItem(self.file_list.row(item))
        self._persist_file_list()

    def _remove_all(self):
        self.file_list.clear()
        self._persist_file_list()

    def _persist_file_list(self):
        """Save the current file list to QSettings so it survives dialog re-opens."""
        files = [self.file_list.item(i).text() for i in range(self.file_list.count())]
        self._settings.setValue("saved_files", files)

    def _browse_output_folder(self):
        folder = QFileDialog.getExistingDirectory(
            self, "选择输出文件夹", self._last_output_dir)
        if folder:
            self.folder_edit.setText(folder)
            self._last_output_dir = folder
            self._settings.setValue("last_output_dir", folder)

    def _confirm(self):
        files = [self.file_list.item(i).text() for i in range(self.file_list.count())]
        if not files:
            QMessageBox.warning(self, "未选择文件", self._no_files_msg)
            return

        if self.radio_same is None:
            output_folder = None
        elif self.radio_same.isChecked():
            output_folder = None
        else:
            output_folder = self.folder_edit.text().strip()
            if not output_folder:
                QMessageBox.warning(self, "未选择输出文件夹", "请选择一个输出文件夹。")
                return

        if self.prefix_checkbox is not None:
            prefix_value = self.prefix_edit.text() if self.prefix_checkbox.isChecked() else ""
            suffix_value = self.suffix_edit.text() if self.suffix_checkbox.isChecked() else ""
            self._settings.setValue("add_prefix", self.prefix_checkbox.isChecked())
            self._settings.setValue("prefix_value", self.prefix_edit.text())
            self._settings.setValue("add_suffix", self.suffix_checkbox.isChecked())
            self._settings.setValue("suffix_value", self.suffix_edit.text())
            self._conversion_fn(files, output_folder, prefix=prefix_value, suffix=suffix_value)
        else:
            self._conversion_fn(files, output_folder)
        QMessageBox.information(self, "导出成功", f"已成功导出 {len(files)} 个文件。")
        self.accept()


# ---------------------------------------------------------------------------
# CATIA installation detector
# ---------------------------------------------------------------------------

def detect_catia_root() -> str | None:
    registry_paths = [
        r"SOFTWARE\Dassault Systemes",
        r"SOFTWARE\WOW6432Node\Dassault Systemes",
    ]
    for reg_path in registry_paths:
        logger.debug(f"Trying registry path: HKEY_LOCAL_MACHINE\\{reg_path}")
        try:
            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, reg_path) as ds_key:
                i = 0
                while True:
                    try:
                        release = winreg.EnumKey(ds_key, i)
                        logger.debug(f"  Trying key: HKEY_LOCAL_MACHINE\\{reg_path}\\{release}\\0")
                        try:
                            with winreg.OpenKey(ds_key, rf"{release}\0") as release_key:
                                try:
                                    install_path, _ = winreg.QueryValueEx(release_key, "DEST_FOLDER")
                                    candidate = Path(install_path)
                                    if (candidate / "win_b64").exists():
                                        logger.debug(f"    -> Valid CATIA installation found: {candidate}")
                                        return str(candidate)
                                except FileNotFoundError:
                                    pass
                        except OSError:
                            pass
                        i += 1
                    except OSError:
                        break
        except OSError:
            pass
    logger.debug("No valid CATIA installation detected.")
    return None


# ---------------------------------------------------------------------------
# Conversion functions
# ---------------------------------------------------------------------------

def CATDrawing_to_PDF(file_paths: list[str], output_folder: str | None = None,
                      prefix: str = "DR_", suffix: str = ""):
    """
    Convert CATDrawing files to PDF using pyCATIA.
    If prefix is non-empty, prepends it to the output filename unless it
    already starts with that prefix.
    If suffix is non-empty, appends it to the output stem unless it
    already ends with that suffix.
    """
    from pycatia import catia

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    bulk_action = None  # "skip_all", "overwrite_all", or "cancel"

    for path in file_paths:
        if bulk_action == "cancel":
            break

        src = Path(path).resolve()
        dest_dir = Path(output_folder).resolve() if output_folder else src.parent.resolve()
        dest_dir.mkdir(parents=True, exist_ok=True)

        stem = src.stem
        if prefix and not stem.startswith(prefix):
            stem = f"{prefix}{stem}"
        if suffix and not stem.endswith(suffix):
            stem = f"{stem}{suffix}"
        out_stem = stem

        logger.info(f"Opening: {src}")
        dest = dest_dir / f"{out_stem}.pdf"

        if dest.exists():
            if bulk_action == "skip_all":
                logger.info(f"  Skipped (skip all): {dest}")
                continue
            if bulk_action == "overwrite_all":
                dest.unlink()
            else:
                msg = QMessageBox()
                msg.setWindowTitle("文件已存在")
                msg.setText(f'"{dest.name}" 已存在于输出文件夹中。')
                msg.setInformativeText(str(dest.parent))
                msg.setIcon(QMessageBox.Icon.Warning)
                skip_btn          = msg.addButton("跳过",     QMessageBox.ButtonRole.RejectRole)
                skip_all_btn      = msg.addButton("全部跳过", QMessageBox.ButtonRole.RejectRole)
                overwrite_btn     = msg.addButton("覆盖",     QMessageBox.ButtonRole.AcceptRole)
                overwrite_all_btn = msg.addButton("全部覆盖", QMessageBox.ButtonRole.AcceptRole)
                cancel_btn        = msg.addButton("取消",     QMessageBox.ButtonRole.DestructiveRole)
                msg.exec()
                clicked = msg.clickedButton()
                if clicked is cancel_btn:
                    bulk_action = "cancel"
                    break
                elif clicked is skip_all_btn:
                    bulk_action = "skip_all"
                    logger.info(f"  Skipped (skip all): {dest}")
                    continue
                elif clicked is skip_btn:
                    logger.info(f"  Skipped: {dest}")
                    continue
                elif clicked is overwrite_all_btn:
                    bulk_action = "overwrite_all"
                    dest.unlink()
                else:  # overwrite_btn
                    dest.unlink()

        documents.open(str(src))
        from pycatia.drafting_interfaces.drawing_document import DrawingDocument
        drawing_doc = DrawingDocument(application.active_document.com_object)
        drawing = drawing_doc.drawing_root
        sheet_count = drawing.sheets.count

        drawing_doc.export_data(str(dest), "pdf")
        if not dest.exists():
            logger.warning(f"  WARNING: export_data did not create {dest}")
        else:
            logger.info(f"  Exported {sheet_count} sheet(s) -> {dest}")

        drawing_doc.close()
        logger.info(f"Done: {src.name}\n")


def CATPart_to_STP(file_paths: list[str], output_folder: str | None = None,
                   prefix: str = "MD_", suffix: str = ""):
    """
    Convert CATPart/CATProduct files to STEP (.stp) using pyCATIA.
    If prefix is non-empty, prepends it to the output filename unless it
    already starts with that prefix.
    If suffix is non-empty, appends it to the output stem unless it
    already ends with that suffix.
    """
    from pycatia import catia

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    bulk_action = None  # "skip_all", "overwrite_all", or "cancel"

    for path in file_paths:
        if bulk_action == "cancel":
            break

        src = Path(path)
        dest_dir = Path(output_folder).resolve() if output_folder else src.parent.resolve()
        dest_dir.mkdir(parents=True, exist_ok=True)

        stem = src.stem
        if prefix and not stem.startswith(prefix):
            stem = f"{prefix}{stem}"
        if suffix and not stem.endswith(suffix):
            stem = f"{stem}{suffix}"
        out_stem = stem

        dest = dest_dir / f"{out_stem}.stp"

        logger.info(f"Opening: {src}")
        if dest.exists():
            if bulk_action == "skip_all":
                logger.info(f"  Skipped (skip all): {dest}")
                continue
            if bulk_action == "overwrite_all":
                dest.unlink()
            else:
                msg = QMessageBox()
                msg.setWindowTitle("文件已存在")
                msg.setText(f'"{dest.name}" 已存在于输出文件夹中。')
                msg.setInformativeText(str(dest.parent))
                msg.setIcon(QMessageBox.Icon.Warning)
                skip_btn          = msg.addButton("跳过",     QMessageBox.ButtonRole.RejectRole)
                skip_all_btn      = msg.addButton("全部跳过", QMessageBox.ButtonRole.RejectRole)
                overwrite_btn     = msg.addButton("覆盖",     QMessageBox.ButtonRole.AcceptRole)
                overwrite_all_btn = msg.addButton("全部覆盖", QMessageBox.ButtonRole.AcceptRole)
                cancel_btn        = msg.addButton("取消",     QMessageBox.ButtonRole.DestructiveRole)
                msg.exec()
                clicked = msg.clickedButton()
                if clicked is cancel_btn:
                    bulk_action = "cancel"
                    break
                elif clicked is skip_all_btn:
                    bulk_action = "skip_all"
                    logger.info(f"  Skipped (skip all): {dest}")
                    continue
                elif clicked is skip_btn:
                    logger.info(f"  Skipped: {dest}")
                    continue
                elif clicked is overwrite_all_btn:
                    bulk_action = "overwrite_all"
                    dest.unlink()
                else:  # overwrite_btn
                    dest.unlink()

        logger.info(f"Opening: {src}")
        documents.open(str(src))
        doc = application.active_document
        doc.export_data(str(dest), "stp")
        logger.info(f"  Exported -> {dest}")
        doc.close()
        logger.info(f"Done: {src.name}\n")


# ---------------------------------------------------------------------------
# Stamp part template function
# ---------------------------------------------------------------------------

def stamp_part_template(file_paths: list[str], output_folder: str | None = None):
    """
    For each CATPart, add the 9 standard user-defined properties if they do
    not already exist. Properties are added as strings with empty default value.
    The part is saved automatically after stamping.
    """
    from pycatia import catia
    from pycatia.mec_mod_interfaces.part_document import PartDocument

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    succeeded = []
    failed    = []

    for path in file_paths:
        src = Path(path).resolve()
        logger.info(f"Opening: {src}")
        try:
            documents.open(str(src))
            doc        = PartDocument(application.active_document.com_object)
            product    = doc.product
            user_props = product.user_ref_properties

            existing_names: set[str] = set()
            for i in range(1, user_props.count + 1):
                try:
                    existing_names.add(user_props.item(i).name)
                except Exception:
                    pass

            added = []
            for prop_name in PART_TEMPLATE_PROPERTIES:
                if prop_name not in existing_names:
                    user_props.create_string(prop_name, "")
                    added.append(prop_name)
                    logger.info(f"  Added property: '{prop_name}'")
                else:
                    logger.info(f"  Skipped (already exists): '{prop_name}'")

            doc.save()
            logger.info(f"  Saved: {src.name}")
            succeeded.append(f"{src.name} (+{len(added)} added)")

        except Exception as e:
            logger.error(f"  ERROR processing {src.name}: {e}")
            failed.append(f"{src.name}: {e}")
        finally:
            try:
                application.active_document.close()
            except Exception:
                pass

    msg = "Stamping complete.\n\n"
    if succeeded:
        msg += "✔ Succeeded:\n" + "\n".join(f"  {s}" for s in succeeded)
    if failed:
        msg += "\n\n✘ Failed:\n" + "\n".join(f"  {f}" for f in failed)

    from PySide6.QtWidgets import QMessageBox
    if failed:
        QMessageBox.warning(None, "刷写零件模板", msg)
    else:
        QMessageBox.information(None, "刷写零件模板", msg)


# ---------------------------------------------------------------------------
# Export BOM Dialog
# ---------------------------------------------------------------------------

BOM_ALL_COLUMNS       = ["Level", "Type", "Part Number", "Nomenclature", "Definition", "Revision", "Source", "Quantity"]
BOM_DEFAULT_COLUMNS   = ["Level", "Type", "Part Number", "Nomenclature", "Definition", "Revision", "Source", "Quantity"]
BOM_PRESET_CUSTOM_COLUMNS = ["物料编码", "物料名称", "规格型号", "物料来源", "数据状态", "存货类别", "质量", "备注"]


class ExportBOMDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("从CATProduct导出BOM")
        self.setMinimumSize(560, 580)

        self._settings = QSettings("CATIACompanion", "ExportBOMDialog")
        self._last_browse_dir = self._settings.value("last_browse_dir", "")
        self._last_output_dir = self._settings.value("last_output_dir", "")

        saved_custom = self._settings.value("custom_columns", [])
        if isinstance(saved_custom, str):
            saved_custom = [saved_custom]
        self._custom_columns: list[str] = list(saved_custom)

        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── Source selection (active doc vs file) ─────────────────────────
        src_group = QGroupBox("数据来源")
        src_layout = QVBoxLayout(src_group)
        self._src_btn_group = QButtonGroup(self)
        self.radio_active = QRadioButton("使用当前CATIA活动文档")
        self.radio_file   = QRadioButton("选择文件:")
        self.radio_file.setChecked(True)
        self._src_btn_group.addButton(self.radio_active)
        self._src_btn_group.addButton(self.radio_file)
        src_layout.addWidget(self.radio_active)
        file_row = QHBoxLayout()
        file_row.addWidget(self.radio_file)
        self.file_edit = QLineEdit()
        self.file_edit.setPlaceholderText("选择一个CATProduct文件...")
        self.file_edit.setReadOnly(True)
        file_browse_btn = QPushButton("浏览...")
        file_browse_btn.clicked.connect(self._browse_file)
        self._file_browse_btn = file_browse_btn
        file_row.addWidget(self.file_edit)
        file_row.addWidget(file_browse_btn)
        src_layout.addLayout(file_row)
        self.radio_active.toggled.connect(self._toggle_source_row)
        layout.addWidget(src_group)

        output_group = QGroupBox("输出文件夹")
        output_layout = QVBoxLayout(output_group)
        self.radio_same = QRadioButton("与源文件相同目录")
        self.radio_custom = QRadioButton("自定义目录:")
        self.radio_same.setChecked(True)
        self.btn_group = QButtonGroup(self)
        self.btn_group.addButton(self.radio_same)
        self.btn_group.addButton(self.radio_custom)
        output_layout.addWidget(self.radio_same)
        output_layout.addWidget(self.radio_custom)
        folder_row = QHBoxLayout()
        self.folder_edit = QLineEdit()
        self.folder_edit.setPlaceholderText("选择输出文件夹...")
        self.folder_edit.setReadOnly(True)
        self.folder_edit.setEnabled(False)
        self.folder_browse_btn = QPushButton("浏览...")
        self.folder_browse_btn.setEnabled(False)
        self.folder_browse_btn.clicked.connect(self._browse_output_folder)
        folder_row.addWidget(self.folder_edit)
        folder_row.addWidget(self.folder_browse_btn)
        output_layout.addLayout(folder_row)
        self.radio_custom.toggled.connect(self._toggle_folder_row)
        layout.addWidget(output_group)

        if self._last_output_dir:
            self.radio_custom.setChecked(True)
            self.folder_edit.setText(self._last_output_dir)

        col_group = QGroupBox("导出列（拖动以排序）")
        col_outer = QVBoxLayout(col_group)
        col_layout = QHBoxLayout()

        avail_layout = QVBoxLayout()
        avail_layout.addWidget(QLabel("可用列:"))
        self.avail_list = QListWidget()
        self.avail_list.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.avail_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        avail_layout.addWidget(self.avail_list)
        col_layout.addLayout(avail_layout)

        arrow_layout = QVBoxLayout()
        arrow_layout.addStretch()
        add_btn = QPushButton("→")
        add_btn.setFixedWidth(36)
        add_btn.clicked.connect(self._add_column)
        remove_btn = QPushButton("←")
        remove_btn.setFixedWidth(36)
        remove_btn.clicked.connect(self._remove_column)
        up_btn = QPushButton("↑")
        up_btn.setFixedWidth(36)
        up_btn.clicked.connect(self._move_up)
        down_btn = QPushButton("↓")
        down_btn.setFixedWidth(36)
        down_btn.clicked.connect(self._move_down)
        arrow_layout.addWidget(add_btn)
        arrow_layout.addWidget(remove_btn)
        arrow_layout.addSpacing(10)
        arrow_layout.addWidget(up_btn)
        arrow_layout.addWidget(down_btn)
        arrow_layout.addStretch()
        col_layout.addLayout(arrow_layout)

        selected_layout = QVBoxLayout()
        selected_layout.addWidget(QLabel("已选列:"))
        self.selected_list = QListWidget()
        self.selected_list.setDragDropMode(QAbstractItemView.DragDropMode.DragDrop)
        self.selected_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        selected_layout.addWidget(self.selected_list)
        col_layout.addLayout(selected_layout)
        col_outer.addLayout(col_layout)
        layout.addWidget(col_group)

        saved = self._settings.value("selected_columns", BOM_DEFAULT_COLUMNS)
        if isinstance(saved, str):
            saved = [saved]
        all_known = BOM_ALL_COLUMNS + list(BOM_PRESET_CUSTOM_COLUMNS) + [
            c for c in self._custom_columns
            if c not in BOM_ALL_COLUMNS and c not in BOM_PRESET_CUSTOM_COLUMNS
        ]
        for col in saved:
            if col in all_known:
                self.selected_list.addItem(self._make_col_item(col))
        for col in all_known:
            if col not in saved:
                self.avail_list.addItem(self._make_col_item(col))

        action_row = QHBoxLayout()
        action_row.addStretch()
        confirm_btn = QPushButton("导出")
        confirm_btn.setDefault(True)
        confirm_btn.clicked.connect(self._confirm)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        action_row.addWidget(confirm_btn)
        action_row.addWidget(cancel_btn)
        layout.addLayout(action_row)

    # ── Helpers ────────────────────────────────────────────────────────────

    @staticmethod
    def _make_col_item(internal_name: str) -> QListWidgetItem:
        """Create a QListWidgetItem showing the Chinese display name but
        storing the internal name in UserRole."""
        item = QListWidgetItem(_BOM_COL_DISPLAY.get(internal_name, internal_name))
        item.setData(Qt.ItemDataRole.UserRole, internal_name)
        return item

    @staticmethod
    def _item_internal(item: QListWidgetItem) -> str:
        """Return the internal column name stored in a list item."""
        data = item.data(Qt.ItemDataRole.UserRole)
        return data if data else item.text()

    def _toggle_folder_row(self, checked):
        self.folder_edit.setEnabled(checked)
        self.folder_browse_btn.setEnabled(checked)

    def _toggle_source_row(self, active_checked: bool):
        self.file_edit.setEnabled(not active_checked)
        self._file_browse_btn.setEnabled(not active_checked)
        # When using active doc, "same directory" is meaningless unless
        # a custom output folder is set
        if active_checked and self.radio_same.isChecked():
            self.radio_custom.setChecked(True)

    def _browse_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "选择CATProduct文件",
            self._last_browse_dir, "*.CATProduct (*.CATProduct);;All Files (*)")
        if file:
            self.file_edit.setText(file)
            self._last_browse_dir = str(Path(file).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)

    def _browse_output_folder(self):
        folder = QFileDialog.getExistingDirectory(
            self, "选择输出文件夹", self._last_output_dir)
        if folder:
            self.folder_edit.setText(folder)
            self._last_output_dir = folder
            self._settings.setValue("last_output_dir", folder)

    def _add_column(self):
        for item in self.avail_list.selectedItems():
            internal = self._item_internal(item)
            self.avail_list.takeItem(self.avail_list.row(item))
            self.selected_list.addItem(self._make_col_item(internal))

    def _remove_column(self):
        for item in self.selected_list.selectedItems():
            internal = self._item_internal(item)
            self.selected_list.takeItem(self.selected_list.row(item))
            self.avail_list.addItem(self._make_col_item(internal))

    def _move_up(self):
        row = self.selected_list.currentRow()
        if row > 0:
            item = self.selected_list.takeItem(row)
            self.selected_list.insertItem(row - 1, item)
            self.selected_list.setCurrentRow(row - 1)

    def _move_down(self):
        row = self.selected_list.currentRow()
        if row < self.selected_list.count() - 1:
            item = self.selected_list.takeItem(row)
            self.selected_list.insertItem(row + 1, item)
            self.selected_list.setCurrentRow(row + 1)

    def _confirm(self):
        use_active = self.radio_active.isChecked()
        if use_active:
            file_path = None
        else:
            file_path = self.file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请选择一个CATProduct文件。")
                return

        # Collect selected columns as internal names; save internal names to QSettings
        selected_cols = [self._item_internal(self.selected_list.item(i))
                         for i in range(self.selected_list.count())]
        if not selected_cols:
            QMessageBox.warning(self, "未选择列", "请至少选择一列进行导出。")
            return
        self._settings.setValue("selected_columns", selected_cols)

        if self.radio_same.isChecked() and not use_active:
            output_folder = None
        else:
            output_folder = self.folder_edit.text().strip()
            if not output_folder:
                QMessageBox.warning(self, "未选择输出文件夹",
                                    "请选择一个输出文件夹（使用活动文档时需指定）。")
                return
        export_bom_to_excel([file_path], output_folder, columns=selected_cols,
                            custom_columns=self._custom_columns)
        QMessageBox.information(self, "导出成功", "BOM已成功导出为Excel文件。")
        self.accept()


# ---------------------------------------------------------------------------
# BOM export function
# ---------------------------------------------------------------------------

def export_bom_to_excel(file_paths: list[str | None], output_folder: str | None = None,
                        columns: list[str] | None = None,
                        custom_columns: list[str] | None = None):
    """
    Export a hierarchical BOM from CATProduct files to Excel (.xlsx).
    Custom columns are read from CATIA user-defined properties (UserRefProperties).
    Each product is switched to DESIGN_MODE before reading properties.

    A *None* entry in *file_paths* means "use the currently active CATIA document"
    without opening or closing anything.

    Traversal is delegated to _collect_bom_rows to avoid code duplication.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from pycatia import catia
    from pycatia.product_structure_interfaces.product_document import ProductDocument

    if columns is None:
        columns = BOM_DEFAULT_COLUMNS
    if custom_columns is None:
        custom_columns = []

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    def _write_sheet(ws, rows):
        center = Alignment(horizontal="center")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx,
                           value=_BOM_COL_DISPLAY.get(col_name, col_name))
            cell.font = Font(bold=True)

        for row_idx, row in enumerate(rows, start=2):
            level = row.get("Level", 0)
            for col_idx, col_name in enumerate(columns, start=1):
                if col_name == "Level":
                    value = level
                elif col_name == "Quantity":
                    value = row.get("Quantity", 1)
                elif col_name == "Type":
                    value = row.get("Type", "")
                elif col_name == "Source":
                    raw = str(row.get("Source", ""))
                    value = _SOURCE_TO_DISPLAY.get(raw, raw)
                else:
                    value = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_name in ("Level", "Quantity", "Type"):
                    cell.alignment = center

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            header = _BOM_COL_DISPLAY.get(col_name, col_name)
            min_w = _BOM_COL_MIN_WIDTH.get(col_name, 10)
            max_width = max(_col_display_width(header), min_w)
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_width = max(max_width, _col_display_width(str(cell_val)))
            ws.column_dimensions[col_letter].width = max_width + 2

    for path in file_paths:
        if path is None:
            # Use the active document without opening or closing
            try:
                active_full = application.active_document.full_name
            except Exception as e:
                raise RuntimeError("无法获取当前CATIA活动文档，请确保CATIA已打开CATProduct。") from e
            src_name = Path(active_full)
            dest_dir = Path(output_folder).resolve() if output_folder else src_name.parent
            dest_dir.mkdir(parents=True, exist_ok=True)
            dest = dest_dir / f"{src_name.stem}_BOM.xlsx"
            rows = _collect_bom_rows(None, columns, custom_columns)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM"
            _write_sheet(ws, rows)
            wb.save(str(dest))
            logger.info(f"  BOM exported -> {dest}")
            logger.info("Done: active document\n")
            continue

        src = Path(path).resolve()
        dest_dir = Path(output_folder).resolve() if output_folder else src.parent
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / f"{src.stem}_BOM.xlsx"

        if dest.exists():
            try:
                with open(dest, "a+b"):
                    pass
            except PermissionError:
                from PySide6.QtWidgets import QMessageBox
                reply = QMessageBox.question(None, "文件正在使用",
                    f"该文件当前在Excel中已打开：\n{dest}\n\n"
                    f"请在Excel中关闭该文件，然后点击【重试】，或点击【取消】以中止。",
                    QMessageBox.StandardButton.Retry | QMessageBox.StandardButton.Cancel)
                if reply == QMessageBox.StandardButton.Cancel:
                    continue
                try:
                    with open(dest, "a+b"):
                        pass
                except PermissionError:
                    QMessageBox.critical(None, "文件仍在使用中",
                        f"文件仍处于打开状态，请关闭后重试。\n{dest}")
                    continue

        # Snapshot of open documents before calling _collect_bom_rows so we
        # can close any file that we opened ourselves.
        already_open: set[Path] = set()
        for i in range(1, documents.count + 1):
            try:
                already_open.add(Path(documents.item(i).full_name).resolve())
            except Exception:
                pass

        logger.info(f"Opening: {src}")
        rows = _collect_bom_rows(str(src), columns, custom_columns)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        _write_sheet(ws, rows)

        wb.save(str(dest))
        logger.info(f"  BOM exported -> {dest}")

        # Close the document only if we were the one who opened it
        if src not in already_open:
            for i in range(1, documents.count + 1):
                try:
                    doc = documents.item(i)
                    if Path(doc.full_name).resolve() == src:
                        ProductDocument(doc.com_object).close()
                        break
                except Exception:
                    pass

        logger.info(f"Done: {src.name}\n")


# ---------------------------------------------------------------------------
# Find Dependencies function
# ---------------------------------------------------------------------------

def find_dependencies(target_path: str, progress_callback=None) -> list[str]:
    """
    Find all files that target_path depends on via CATIA COM.

    Opens the target file in a running CATIA instance; CATIA automatically
    loads all referenced documents.  The function collects the full paths of
    every newly-opened document, then closes them before returning.
    """
    from pycatia import catia

    target = Path(target_path).resolve()

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    # Snapshot of documents already open before we do anything
    already_open: set[Path] = set()
    for i in range(1, documents.count + 1):
        try:
            already_open.add(Path(documents.item(i).full_name).resolve())
        except Exception:
            pass

    logger.info(f"Opening target for dependency scan: {target}")
    if progress_callback:
        progress_callback("正在打开文件，请稍候…")

    documents.open(str(target))

    # Collect every document CATIA opened as a side-effect
    results: list[str] = []
    newly_opened: set[Path] = set()

    for i in range(1, documents.count + 1):
        try:
            doc = documents.item(i)
            doc_path = Path(doc.full_name).resolve()
            if doc_path == target or doc_path in already_open:
                continue
            newly_opened.add(doc_path)
            results.append(str(doc_path))
            logger.info(f"  Dependency: {doc_path}")
        except Exception as e:
            logger.debug(f"  Could not read document {i}: {e}")

    # Close all documents we opened (target last)
    for i in range(documents.count, 0, -1):
        try:
            doc = documents.item(i)
            doc_path = Path(doc.full_name).resolve()
            if doc_path in newly_opened or doc_path == target:
                doc.close()
        except Exception:
            pass

    logger.info(f"Dependency scan complete: {len(results)} found for {target.name}")
    return results


# ---------------------------------------------------------------------------
# Find Dependencies Dialog
# ---------------------------------------------------------------------------

class FindDependenciesDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("查找所有依赖项")
        self.setMinimumSize(540, 420)

        self._settings = QSettings("CATIACompanion", "FindDependenciesDialog")

        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        # Target file
        layout.addWidget(QLabel("目标CATIA文件（CATPart / CATProduct / CATDrawing）:"))
        target_row = QHBoxLayout()
        self._target_edit = QLineEdit()
        self._target_edit.setReadOnly(True)
        self._target_edit.setPlaceholderText("选择目标CATIA文件...")
        self._target_edit.setText(self._settings.value("last_target", ""))
        target_browse_btn = QPushButton("浏览...")
        target_browse_btn.clicked.connect(self._browse_target)
        target_row.addWidget(self._target_edit)
        target_row.addWidget(target_browse_btn)
        layout.addLayout(target_row)

        note = QLabel("通过CATIA COM打开文件并自动收集所有引用文档，请确保CATIA已运行。")
        note.setWordWrap(True)
        note.setStyleSheet("color: gray; font-size: 11px;")
        layout.addWidget(note)

        # Action buttons
        action_row = QHBoxLayout()
        self._search_btn = QPushButton("开始搜索")
        self._search_btn.setDefault(True)
        self._search_btn.clicked.connect(self._start_search)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        action_row.addWidget(self._search_btn)
        action_row.addWidget(cancel_btn)
        action_row.addStretch()
        layout.addLayout(action_row)

        # Results area
        layout.addWidget(QLabel("找到的依赖项："))
        self._result_view = QPlainTextEdit()
        self._result_view.setReadOnly(True)
        self._result_view.setMinimumHeight(150)
        layout.addWidget(self._result_view)

        copy_btn = QPushButton("复制结果")
        copy_btn.clicked.connect(self._copy_results)
        layout.addWidget(copy_btn)

    def _browse_target(self):
        last = self._settings.value("last_target", "")
        start_dir = str(Path(last).parent) if last else ""
        file, _ = QFileDialog.getOpenFileName(
            self, "选择目标CATIA文件", start_dir,
            "*.CATPart *.CATProduct *.CATDrawing (*.CATPart *.CATProduct *.CATDrawing);;All Files (*)")
        if file:
            self._target_edit.setText(file)
            self._settings.setValue("last_target", file)

    def _start_search(self):
        target = self._target_edit.text().strip()

        if not target:
            QMessageBox.warning(self, "未选择目标文件", "请先选择一个目标CATIA文件。")
            return
        if not Path(target).exists():
            QMessageBox.warning(self, "文件不存在", f"目标文件不存在：\n{target}")
            return

        self._search_btn.setEnabled(False)
        self._result_view.setPlainText("正在通过CATIA COM搜索依赖项，请稍候…")
        QApplication.processEvents()

        def progress_callback(msg: str):
            self._result_view.setPlainText(msg)
            QApplication.processEvents()

        try:
            results = find_dependencies(target, progress_callback=progress_callback)
        except Exception as e:
            self._search_btn.setEnabled(True)
            QMessageBox.critical(self, "搜索失败",
                f"通过CATIA COM搜索依赖项时出错：\n{e}\n\n请确保CATIA已启动。")
            self._result_view.setPlainText(f"搜索失败：{e}")
            return

        self._search_btn.setEnabled(True)

        if results:
            summary = f"搜索完成，共找到 {len(results)} 个依赖项：\n\n"
            summary += "\n".join(results)
        else:
            summary = "搜索完成，未找到任何依赖项。"
        self._result_view.setPlainText(summary)

    def _copy_results(self):
        text = self._result_view.toPlainText()
        if text:
            QApplication.clipboard().setText(text)


# ---------------------------------------------------------------------------
# BOM Edit (BOM属性补全) – constants, helpers, dialog
# ---------------------------------------------------------------------------

# Columns that are structural / derived – shown read-only in the edit table
_BOM_READONLY_COLS = {"Level", "Type", "Filename", "Quantity"}

# Allowed characters in a Part Number (used for edit validation and file rename)
_PN_VALID_RE = re.compile(r'^[A-Za-z0-9_ -]*$')

# Internal column name → Chinese display name
_BOM_COL_DISPLAY: dict[str, str] = {
    "Level":        "层级",
    "Type":         "类型",
    "Filename":     "文件名",
    "Part Number":  "零件编号",
    "Nomenclature": "术语（中文名称）",
    "Definition":   "定义",
    "Revision":     "版本",
    "Source":       "源",
    "Quantity":     "数量",
}

# Source: CATIA integer string ↔ Chinese label
_SOURCE_TO_DISPLAY: dict[str, str] = {"0": "未知", "1": "自制", "2": "外购"}
_SOURCE_FROM_DISPLAY: dict[str, str] = {"未知": "0", "自制": "1", "外购": "2"}
_SOURCE_OPTIONS: list[str] = ["未知", "自制", "外购"]

# Reverse display map (Chinese → internal) for ExportBOMDialog list widgets
_BOM_COL_DISPLAY_REVERSE: dict[str, str] = {v: k for k, v in _BOM_COL_DISPLAY.items()}

# Minimum column widths (Excel character units) for standard BOM columns
_BOM_COL_MIN_WIDTH: dict[str, int] = {
    "Level":        6,
    "Type":         10,
    "Filename":     30,
    "Part Number":  20,
    "Nomenclature": 20,
    "Definition":   20,
    "Revision":     10,
    "Source":       8,
    "Quantity":     8,
}


def _col_display_width(text: str) -> int:
    """Return the approximate display width of *text* in Excel column-width units.
    CJK/wide characters count as 2; all others count as 1."""
    return sum(2 if unicodedata.east_asian_width(c) in ("W", "F") else 1
               for c in str(text))

# Column order used in the BOM edit dialog (internal names)
_BOM_EDIT_COLUMN_ORDER: list[str] = [
    "Level", "Type", "Filename", "Part Number", "Quantity",
    "Nomenclature", "Revision", "Definition", "Source",
]


def _get_product_filepath(product) -> str:
    """Return the full file path of the document backing *product*."""
    try:
        return product.reference_product.com_object.Parent.FullName
    except Exception:
        pass
    try:
        return product.com_object.ReferenceProduct.Parent.FullName
    except Exception:
        pass
    try:
        return product.com_object.Parent.FullName
    except Exception:
        pass
    return ""


def _collect_bom_rows(file_path: str | None, columns: list[str],
                      custom_columns: list[str]) -> list[dict]:
    """
    Return a list of row dicts representing the hierarchical BOM.

    *file_path* may be None, in which case the currently active CATIA document
    is used without opening or closing any file.
    """
    from pycatia import catia
    from pycatia.product_structure_interfaces.product_document import ProductDocument
    from pycatia import CatWorkModeType

    DIRECT_ATTR_MAP = {
        "Nomenclature": "nomenclature",
        "Revision":     "revision",
        "Definition":   "definition",
        "Source":       "source",
    }

    def get_prop(product, name: str) -> str:
        attr = DIRECT_ATTR_MAP.get(name)
        if not attr:
            return ""
        targets = [product]
        try:
            targets.insert(0, product.reference_product)
        except Exception:
            pass
        for target in targets:
            try:
                value = getattr(target, attr)
                if value is not None:
                    return str(value)
            except Exception:
                pass
        return ""

    def get_user_prop(product, name: str) -> str:
        targets = [product]
        try:
            targets.insert(0, product.reference_product)
        except Exception:
            pass
        for target in targets:
            try:
                prop = target.user_ref_properties.item(name)
                value = prop.value
                if value is not None and str(value).strip():
                    return str(value)
            except Exception:
                pass
        return ""

    def traverse(product, rows: list, level: int, parent_filepath: str = ""):
        try:
            pn = product.part_number
        except Exception:
            name = product.name
            pn = name.rsplit(".", 1)[0] if "." in name else name

        filepath = _get_product_filepath(product)

        _readable = True
        try:
            product.apply_work_mode(CatWorkModeType.DESIGN_MODE)
        except Exception:
            _readable = False

        row: dict = {
            "Level": level,
            "Part Number": pn,
            "Filename": Path(filepath).stem if filepath else pn,
            "_filepath": filepath,
            "_unreadable": not _readable,
        }
        try:
            child_count = product.products.count
            if filepath and filepath == parent_filepath:
                # Shares the parent's file → embedded component (部件)
                row["Type"] = "部件"
            elif child_count > 0:
                row["Type"] = "装配体"
            else:
                row["Type"] = "零件"
        except Exception:
            row["Type"] = ""

        for col in columns:
            if col in DIRECT_ATTR_MAP:
                row[col] = get_prop(product, col)
            elif col in custom_columns:
                row[col] = get_user_prop(product, col)

        rows.append(row)

        try:
            products = product.products
            count = products.count
            if count == 0:
                return
            children: dict = {}
            for i in range(1, count + 1):
                try:
                    child = products.item(i)
                    try:
                        cpn = child.part_number
                    except Exception:
                        try:
                            cpn = child.reference_product.part_number
                        except Exception:
                            n = child.name
                            cpn = n.rsplit(".", 1)[0] if "." in n else n
                except Exception:
                    continue
                if cpn not in children:
                    children[cpn] = {"product": child, "qty": 0}
                children[cpn]["qty"] += 1
            for cpn, data in children.items():
                child_rows: list = []
                traverse(data["product"], child_rows, level + 1, parent_filepath=filepath)
                if child_rows:
                    child_rows[0]["Quantity"] = data["qty"]
                rows.extend(child_rows)
        except Exception:
            pass

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    if file_path is None:
        # Use the currently active CATIA document without opening/closing
        product_doc = ProductDocument(application.active_document.com_object)
        root_product = product_doc.product
        rows: list[dict] = []
        traverse(root_product, rows, level=0)
        return rows

    src = Path(file_path).resolve()
    already_open: set[Path] = set()
    for i in range(1, documents.count + 1):
        try:
            already_open.add(Path(documents.item(i).full_name).resolve())
        except Exception:
            pass

    if src not in already_open:
        documents.open(str(src))

    # Find the target document
    target_doc = None
    for i in range(1, documents.count + 1):
        try:
            doc = documents.item(i)
            if Path(doc.full_name).resolve() == src:
                target_doc = doc
                break
        except Exception:
            pass
    if target_doc is None:
        raise RuntimeError(f"无法在CATIA中找到文档：{src}")

    product_doc = ProductDocument(target_doc.com_object)
    root_product = product_doc.product
    rows: list[dict] = []
    traverse(root_product, rows, level=0)
    return rows


def _write_bom_to_catia(file_path: str | None, pn_data: dict[str, dict[str, str]],
                        custom_columns: list[str]):
    """
    Traverse the product tree and write back every editable property stored in
    *pn_data* (keyed by original Part Number) to CATIA via COM.

    When *file_path* is None the currently active CATIA document is used and
    nothing is saved (the caller is responsible for any saving).
    Otherwise the file is opened if necessary and all modified documents are saved.
    """
    from pycatia import catia
    from pycatia.product_structure_interfaces.product_document import ProductDocument
    from pycatia import CatWorkModeType

    WRITABLE_DIRECT = {
        "Nomenclature": "nomenclature",
        "Revision":     "revision",
        "Definition":   "definition",
        "Source":       "source",
    }

    def set_prop(product, name: str, value: str):
        attr = WRITABLE_DIRECT.get(name)
        if not attr:
            return
        targets: list = []
        try:
            targets.append(product.reference_product)
        except Exception:
            pass
        targets.append(product)
        for target in targets:
            try:
                setattr(target, attr, int(_SOURCE_FROM_DISPLAY.get(value, value))
                        if name == "Source" else value)
                return
            except Exception:
                continue

    def set_user_prop(product, name: str, value: str):
        targets: list = []
        try:
            targets.append(product.reference_product)
        except Exception:
            pass
        targets.append(product)
        for target in targets:
            try:
                target.user_ref_properties.item(name).value = value
                return
            except Exception:
                pass
        # Property does not exist on any target – create it on the first
        # available target (reference_product preferred).
        for target in targets:
            try:
                target.user_ref_properties.create_string(name, value)
                return
            except Exception:
                continue

    def traverse_write(product):
        try:
            pn = product.part_number
        except Exception:
            name = product.name
            pn = name.rsplit(".", 1)[0] if "." in name else name

        if pn in pn_data:
            try:
                product.apply_work_mode(CatWorkModeType.DESIGN_MODE)
            except Exception:
                pass
            for col, value in pn_data[pn].items():
                if col in _BOM_READONLY_COLS:
                    continue
                if col == "Part Number":
                    try:
                        product.part_number = value
                    except Exception:
                        try:
                            product.com_object.PartNumber = value
                        except Exception:
                            pass
                elif col in WRITABLE_DIRECT:
                    set_prop(product, col, value)
                elif col in custom_columns:
                    set_user_prop(product, col, value)

        try:
            count = product.products.count
            for i in range(1, count + 1):
                try:
                    traverse_write(product.products.item(i))
                except Exception:
                    pass
        except Exception:
            pass

    caa = catia()
    application = caa.application
    application.visible = True
    documents = application.documents

    if file_path is None:
        # Use the active document; do not open or save anything
        product_doc = ProductDocument(application.active_document.com_object)
        root_product = product_doc.product
        traverse_write(root_product)
        logger.info("Write-back complete for active document (not saved)")
        return

    src = Path(file_path).resolve()
    already_open: set[Path] = set()
    for i in range(1, documents.count + 1):
        try:
            already_open.add(Path(documents.item(i).full_name).resolve())
        except Exception:
            pass

    if src not in already_open:
        documents.open(str(src))

    target_doc = None
    for i in range(1, documents.count + 1):
        try:
            doc = documents.item(i)
            if Path(doc.full_name).resolve() == src:
                target_doc = doc
                break
        except Exception:
            pass
    if target_doc is None:
        raise RuntimeError(f"无法在CATIA中找到文档：{src}")

    product_doc = ProductDocument(target_doc.com_object)
    root_product = product_doc.product
    traverse_write(root_product)

    logger.info(f"Write-back complete for {src.name} (not saved; user must save manually in CATIA)")


class BomEditDialog(QDialog):
    """
    Displays the BOM of a CATProduct in an editable table.

    - 文件名 (Filename), 层级 (Level), 类型 (Type), 数量 (Quantity) are read-only.
    - 零件编号 (Part Number) is editable with conflict checking.
    - All other columns are editable.  源 (Source) is a QComboBox (未知/自制/外购).
    - Level cells are indented with (level) leading spaces for hierarchy clarity.
    - Rows sharing the same Part Number are linked: editing one cell propagates
      the new value to every other row with the same PN.
    - Clicking "完成" writes the changes back to CATIA via COM.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("BOM属性补全")
        self.setMinimumSize(900, 600)
        self.resize(1100, 700)
        self.setWindowFlags(
            self.windowFlags() |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowMinimizeButtonHint
        )

        # Share custom-column config with ExportBOMDialog
        self._settings = QSettings("CATIACompanion", "ExportBOMDialog")
        self._last_browse_dir = self._settings.value("last_browse_dir", "")

        saved_custom = self._settings.value("custom_columns", [])
        if isinstance(saved_custom, str):
            saved_custom = [saved_custom]
        self._custom_columns: list[str] = list(saved_custom)

        # BomEditDialog-specific settings (visible preset columns)
        self._edit_settings = QSettings("CATIACompanion", "BomEditDialog")
        saved_visible = self._edit_settings.value("visible_preset_columns", [])
        if isinstance(saved_visible, str):
            saved_visible = [saved_visible]
        self._visible_preset_cols: list[str] = [
            c for c in saved_visible if c in BOM_PRESET_CUSTOM_COLUMNS
        ]

        # All custom columns (for data reading) = shared custom + all presets
        self._all_custom_columns: list[str] = list(dict.fromkeys(
            self._custom_columns + list(BOM_PRESET_CUSTOM_COLUMNS)
        ))

        # Internal column names in the desired display order (visible only)
        self._columns: list[str] = self._build_visible_columns()

        # PN-keyed canonical data: {original_pn: {internal_col: value}}
        # Source is stored as display label (未知/自制/外购).
        self._pn_data: dict[str, dict[str, str]] = {}
        # Snapshot of values at last load/apply (for dirty-only write-back)
        self._original_pn_data: dict[str, dict[str, str]] = {}
        # Modified fields per original PN: {original_pn: {col_name, ...}}
        self._dirty: dict[str, set[str]] = {}
        # All BOM rows in traversal order
        self._rows: list[dict] = []
        # Guard against re-entrant change handling
        self._updating = False
        # Row indices of collapsed assembly rows
        self._collapsed_rows: set[int] = set()

        # ── Layout ──────────────────────────────────────────────────────────
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── Source selection (active doc vs file) ─────────────────────────
        self._use_active_chk = QCheckBox("使用当前CATIA活动文档（不选择文件）")
        self._use_active_chk.toggled.connect(self._toggle_file_row)
        layout.addWidget(self._use_active_chk)

        # File picker row
        file_row = QHBoxLayout()
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("选择一个CATProduct文件...")
        self._file_edit.setReadOnly(True)
        self._file_browse_btn = QPushButton("浏览...")
        self._file_browse_btn.clicked.connect(self._browse_file)
        self._load_btn = QPushButton("加载BOM")
        self._load_btn.clicked.connect(self._load_bom)
        file_row.addWidget(self._file_edit)
        file_row.addWidget(self._file_browse_btn)
        file_row.addWidget(self._load_btn)
        layout.addLayout(file_row)

        note = QLabel(
            "文件名 / 层级 / 类型 / 数量 为结构属性，不可编辑。"
            "零件编号可编辑但不能与其他行冲突。"
            "相同零件编号的行会联动更新。请确保 CATIA 已启动。"
        )
        note.setWordWrap(True)
        note.setStyleSheet("color: gray; font-size: 11px;")
        layout.addWidget(note)

        # ── Preset custom column checkboxes ──────────────────────────────
        preset_group = QGroupBox("自定义属性列（勾选以显示）")
        preset_layout = QHBoxLayout(preset_group)
        preset_layout.setSpacing(12)
        self._preset_checkboxes: dict[str, QCheckBox] = {}
        for col_name in BOM_PRESET_CUSTOM_COLUMNS:
            cb = QCheckBox(col_name)
            cb.setChecked(col_name in self._visible_preset_cols)
            cb.toggled.connect(self._on_preset_col_toggled)
            preset_layout.addWidget(cb)
            self._preset_checkboxes[col_name] = cb
        layout.addWidget(preset_group)

        # Editable table
        display_headers = [_BOM_COL_DISPLAY.get(c, c) for c in self._columns]
        self._table = QTableWidget(0, len(self._columns))
        self._table.setHorizontalHeaderLabels(display_headers)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)
        self._table.verticalHeader().setDefaultSectionSize(24)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._table.setAlternatingRowColors(True)
        self._table.itemChanged.connect(self._on_item_changed)
        self._table.cellClicked.connect(self._on_cell_clicked)
        layout.addWidget(self._table)

        # Bottom buttons
        btn_row = QHBoxLayout()
        self._rename_btn = QPushButton("按零件编号将文件改名")
        self._rename_btn.setEnabled(False)
        self._rename_btn.clicked.connect(self._rename_by_part_number)
        btn_row.addWidget(self._rename_btn)
        btn_row.addStretch()
        self._save_btn = QPushButton("应用（写回CATIA）")
        self._save_btn.setEnabled(False)
        self._save_btn.clicked.connect(self._apply_changes)
        self._finish_btn = QPushButton("完成（写回CATIA）")
        self._finish_btn.setDefault(True)
        self._finish_btn.setEnabled(False)
        self._finish_btn.clicked.connect(self._finish_and_close)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(self._save_btn)
        btn_row.addWidget(self._finish_btn)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

    # ── Source toggle ───────────────────────────────────────────────────────

    def _toggle_file_row(self, use_active: bool):
        self._file_edit.setEnabled(not use_active)
        self._file_browse_btn.setEnabled(not use_active)

    # ── Preset column helpers ──────────────────────────────────────────────

    def _build_visible_columns(self) -> list[str]:
        """Return the ordered list of visible columns for the table."""
        # Base columns + visible preset columns + other non-preset custom columns
        visible_preset = [c for c in BOM_PRESET_CUSTOM_COLUMNS
                          if c in self._visible_preset_cols]
        other_custom = [c for c in self._custom_columns
                        if c not in _BOM_EDIT_COLUMN_ORDER
                        and c not in BOM_PRESET_CUSTOM_COLUMNS]
        return _BOM_EDIT_COLUMN_ORDER + visible_preset + other_custom

    def _on_preset_col_toggled(self):
        """Handle toggling of a preset custom column checkbox."""
        self._visible_preset_cols = [
            name for name, cb in self._preset_checkboxes.items()
            if cb.isChecked()
        ]
        self._edit_settings.setValue(
            "visible_preset_columns", self._visible_preset_cols
        )

        # Rebuild visible columns and update table
        self._columns = self._build_visible_columns()
        display_headers = [_BOM_COL_DISPLAY.get(c, c) for c in self._columns]
        self._table.setColumnCount(len(self._columns))
        self._table.setHorizontalHeaderLabels(display_headers)

        if self._rows:
            self._populate_table()
            self._table.resizeColumnsToContents()

    # ── File picker ────────────────────────────────────────────────────────

    def _browse_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "选择CATProduct文件",
            self._last_browse_dir,
            "*.CATProduct (*.CATProduct);;All Files (*)"
        )
        if file:
            self._file_edit.setText(file)
            self._last_browse_dir = str(Path(file).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)

    # ── Load BOM ───────────────────────────────────────────────────────────

    def _load_bom(self):
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请先选择一个CATProduct文件。")
                return
            if not Path(file_path).exists():
                QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{file_path}")
                return

        self._load_btn.setEnabled(False)
        self._load_btn.setText("加载中…")
        QApplication.processEvents()

        try:
            # Read ALL custom columns (including all presets) so toggling
            # visibility later doesn't require a CATIA reload.
            all_read_cols = list(dict.fromkeys(
                _BOM_EDIT_COLUMN_ORDER +
                [c for c in self._all_custom_columns
                 if c not in _BOM_EDIT_COLUMN_ORDER]
            ))
            rows = _collect_bom_rows(file_path, all_read_cols,
                                     self._all_custom_columns)
        except Exception as e:
            logger.error(f"Failed to load BOM for edit: {e}")
            QMessageBox.critical(
                self, "加载失败",
                f"加载BOM时出错：\n{e}\n\n请确保CATIA已启动。"
            )
            self._load_btn.setEnabled(True)
            self._load_btn.setText("加载BOM")
            return

        self._load_btn.setEnabled(True)
        self._load_btn.setText("重新加载BOM")

        self._rows = rows
        self._collapsed_rows.clear()

        # Build PN-keyed canonical data (first occurrence wins).
        # Source is converted from CATIA integer string to display label.
        # Store values for ALL columns (including hidden preset columns)
        # so that toggling visibility later is instant.
        all_data_cols = list(dict.fromkeys(
            _BOM_EDIT_COLUMN_ORDER +
            [c for c in self._all_custom_columns
             if c not in _BOM_EDIT_COLUMN_ORDER]
        ))
        self._pn_data = {}
        for row in rows:
            pn = str(row.get("Part Number", ""))
            if pn and pn not in self._pn_data:
                data: dict[str, str] = {}
                for col in all_data_cols:
                    val = str(row.get(col, ""))
                    if col == "Source":
                        val = _SOURCE_TO_DISPLAY.get(val, val)
                    data[col] = val
                self._pn_data[pn] = data

        # Snapshot original values for dirty-only write-back; clear stale dirty state
        self._original_pn_data = copy.deepcopy(self._pn_data)
        self._dirty.clear()

        self._populate_table()
        # Auto-size columns to content after initial load; user can resize manually after
        self._table.resizeColumnsToContents()
        self._save_btn.setEnabled(True)
        self._finish_btn.setEnabled(True)
        self._rename_btn.setEnabled(True)

    def _populate_table(self):
        self._updating = True

        # Sync table columns with current visible column list
        display_headers = [_BOM_COL_DISPLAY.get(c, c) for c in self._columns]
        self._table.setColumnCount(len(self._columns))
        self._table.setHorizontalHeaderLabels(display_headers)

        self._table.setRowCount(0)
        self._table.setRowCount(len(self._rows))

        src_col_idx = (
            self._columns.index("Source") if "Source" in self._columns else -1
        )

        for row_idx, row_data in enumerate(self._rows):
            pn = str(row_data.get("Part Number", ""))
            level = row_data.get("Level", 0)
            unreadable = bool(row_data.get("_unreadable"))

            for col_idx, col_name in enumerate(self._columns):

                # ── Source column → QComboBox ────────────────────────────
                if col_name == "Source":
                    raw = str(row_data.get("Source", ""))
                    display_val = _SOURCE_TO_DISPLAY.get(raw, raw)
                    # Canonical value may already be a display label
                    pn_val = self._pn_data.get(pn, {}).get("Source", display_val)
                    if pn_val not in _SOURCE_OPTIONS:
                        pn_val = _SOURCE_TO_DISPLAY.get(pn_val, _SOURCE_OPTIONS[0])

                    combo = QComboBox()
                    combo.addItems(_SOURCE_OPTIONS)
                    combo.setCurrentText(pn_val)
                    combo.currentTextChanged.connect(
                        lambda text, r=row_idx: self._on_source_changed(r, text)
                    )
                    self._table.setCellWidget(row_idx, col_idx, combo)
                    continue

                # ── All other columns → QTableWidgetItem ─────────────────
                if col_name == "Level":
                    value = self._level_cell_text(row_idx)
                elif col_name == "Quantity":
                    value = str(row_data.get("Quantity", "1"))
                elif col_name in _BOM_READONLY_COLS:
                    value = str(row_data.get(col_name, ""))
                else:
                    value = str(
                        self._pn_data.get(pn, {}).get(col_name, row_data.get(col_name, ""))
                    )

                item = QTableWidgetItem(value)
                if col_name in _BOM_READONLY_COLS:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                # Show full path as tooltip on the Filename cell
                if col_name == "Filename":
                    fp = str(row_data.get("_filepath", ""))
                    if fp:
                        item.setToolTip(fp)
                self._table.setItem(row_idx, col_idx, item)

            # Grey out and lock rows whose document could not be loaded
            if unreadable:
                grey = QColor(160, 160, 160)
                bg = QColor(245, 245, 245)
                for ci in range(len(self._columns)):
                    it = self._table.item(row_idx, ci)
                    if it:
                        it.setForeground(grey)
                        it.setBackground(bg)
                        it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    w = self._table.cellWidget(row_idx, ci)
                    if isinstance(w, QComboBox):
                        w.setEnabled(False)

        self._updating = False

    # ── Collapse / expand helpers ───────────────────────────────────────────

    def _row_has_children(self, row_idx: int) -> bool:
        """Return True if the next row exists and has a deeper level."""
        if row_idx + 1 >= len(self._rows):
            return False
        return (self._rows[row_idx + 1].get("Level", 0) >
                self._rows[row_idx].get("Level", 0))

    def _level_cell_text(self, row_idx: int) -> str:
        """Return the text to display in the Level cell, including fold indicator."""
        level = self._rows[row_idx].get("Level", 0)
        if self._row_has_children(row_idx):
            indicator = "▶ " if row_idx in self._collapsed_rows else "▼ "
        else:
            indicator = "  "
        return "  " * level + indicator + str(level)

    def _update_row_visibility(self):
        """Show/hide rows based on which ancestor assembly rows are collapsed."""
        hide_depth_stack: list[int] = []
        for r, row_data in enumerate(self._rows):
            level = row_data.get("Level", 0)
            # Pop ancestors that are no longer parents of the current row
            while hide_depth_stack and hide_depth_stack[-1] >= level:
                hide_depth_stack.pop()
            should_hide = bool(hide_depth_stack)
            self._table.setRowHidden(r, should_hide)
            if not should_hide and r in self._collapsed_rows:
                hide_depth_stack.append(level)

    def _on_cell_clicked(self, row: int, col: int):
        """Toggle collapse/expand when the Level cell of an assembly row is clicked."""
        if "Level" not in self._columns:
            return
        level_col = self._columns.index("Level")
        if col != level_col or not self._row_has_children(row):
            return
        if row in self._collapsed_rows:
            self._collapsed_rows.discard(row)
        else:
            self._collapsed_rows.add(row)
        # Refresh the indicator text on the clicked cell
        item = self._table.item(row, level_col)
        if item:
            item.setText(self._level_cell_text(row))
        self._update_row_visibility()


    def _on_source_changed(self, row_idx: int, text: str):
        if self._updating:
            return

        src_col_idx = (
            self._columns.index("Source") if "Source" in self._columns else -1
        )
        if src_col_idx < 0:
            return

        # Determine the set of rows to propagate to: all selected rows (if the
        # triggering row is among the selection) + all rows with the same PN.
        selected_rows = {idx.row() for idx in self._table.selectedIndexes()}
        if row_idx in selected_rows:
            direct_rows = selected_rows
        else:
            direct_rows = {row_idx}

        # Collect the original PNs of all directly affected rows
        pns_to_update: set[str] = set()
        for r in direct_rows:
            pn = str(self._rows[r].get("Part Number", ""))
            if pn:
                pns_to_update.add(pn)

        # Update canonical data for every affected PN
        for pn in pns_to_update:
            if pn in self._pn_data:
                self._pn_data[pn]["Source"] = text
                self._dirty.setdefault(pn, set()).add("Source")

        # Sync all rows whose original PN is in the affected set
        self._updating = True
        for r in range(self._table.rowCount()):
            if r == row_idx:
                continue
            other_pn = str(self._rows[r].get("Part Number", ""))
            if other_pn in pns_to_update:
                combo = self._table.cellWidget(r, src_col_idx)
                if isinstance(combo, QComboBox) and combo.currentText() != text:
                    combo.blockSignals(True)
                    combo.setCurrentText(text)
                    combo.blockSignals(False)
        self._updating = False

    # ── Regular cell edit sync ─────────────────────────────────────────────

    def _on_item_changed(self, item: QTableWidgetItem):
        if self._updating:
            return
        col_idx = item.column()
        row_idx = item.row()
        col_name = self._columns[col_idx]

        # Source is handled by _on_source_changed; read-only cols are skipped
        if col_name in _BOM_READONLY_COLS or col_name == "Source":
            return

        new_value = item.text()
        pn = str(self._rows[row_idx].get("Part Number", ""))

        # ── Part Number conflict checking ────────────────────────────────
        if col_name == "Part Number":
            # Check against current Part Numbers of other PNs
            for other_pn, data in self._pn_data.items():
                if other_pn == pn:
                    continue
                if data.get("Part Number", other_pn) == new_value:
                    QMessageBox.warning(
                        self, "零件编号冲突",
                        f"零件编号 \"{new_value}\" 与 "
                        f"\"{other_pn}\" 的当前零件编号冲突，"
                        f"不允许修改。")
                    self._updating = True
                    item.setText(
                        self._pn_data.get(pn, {}).get("Part Number", pn))
                    self._updating = False
                    return
            # Check against original Part Numbers of other PNs
            for other_pn, data in self._original_pn_data.items():
                if other_pn == pn:
                    continue
                if data.get("Part Number", other_pn) == new_value:
                    QMessageBox.warning(
                        self, "零件编号冲突",
                        f"零件编号 \"{new_value}\" 与 "
                        f"\"{other_pn}\" 的原始零件编号冲突，"
                        f"不允许修改。")
                    self._updating = True
                    item.setText(
                        self._pn_data.get(pn, {}).get("Part Number", pn))
                    self._updating = False
                    return

        # ── Part Number character validity ───────────────────────────────
        if col_name == "Part Number" and new_value:
            if not _PN_VALID_RE.fullmatch(new_value):
                QMessageBox.warning(
                    self, "零件编号含非法字符",
                    f"零件编号 \"{new_value}\" 含有非法字符。\n"
                    "只允许：字母(A-Z, a-z)、数字(0-9)、下划线(_)、连字符(-)、空格( )。")
                self._updating = True
                item.setText(self._pn_data.get(pn, {}).get("Part Number", pn))
                self._updating = False
                return

        # Determine the set of rows to propagate to: all selected rows (if the
        # edited row is among the selection) + all rows with the same PN.
        selected_rows = {idx.row() for idx in self._table.selectedIndexes()}
        if row_idx in selected_rows:
            direct_rows = selected_rows
        else:
            direct_rows = {row_idx}

        # Collect original PNs and update canonical data
        pns_to_update: set[str] = set()
        for r in direct_rows:
            r_pn = str(self._rows[r].get("Part Number", ""))
            if r_pn:
                pns_to_update.add(r_pn)
                if r_pn in self._pn_data:
                    self._pn_data[r_pn][col_name] = new_value
                    self._dirty.setdefault(r_pn, set()).add(col_name)

        # Propagate to all rows whose original PN is in the affected set
        self._updating = True
        for r in range(self._table.rowCount()):
            if r == row_idx:
                continue
            other_pn = str(self._rows[r].get("Part Number", ""))
            if other_pn in pns_to_update:
                other_item = self._table.item(r, col_idx)
                if other_item and other_item.text() != new_value:
                    other_item.setText(new_value)
        self._updating = False

    # ── Write back ─────────────────────────────────────────────────────────

    def _rename_by_part_number(self):
        """Save each CATIA file under a new name matching its Part Number.

        Uses CATIA's SaveAs COM method so that CATIA remains aware of the new
        file location.  If the target file already exists CATIA itself will
        prompt the user whether to overwrite.  Invalid Part Numbers are
        reported and skipped.

        Before renaming the user is asked whether to delete the original file
        after each successful SaveAs.

        Aborts early if there are unsaved BOM edits, prompting the user to
        write them back first so that the Part Number in CATIA matches what
        the table shows.
        """
        # ── Pre-flight: ensure dirty changes have been written back ──────────
        if self._dirty:
            ret = QMessageBox.question(
                self, "存在未回传的修改",
                "检测到BOM属性尚未写回CATIA。\n\n"
                "必须先将修改写回CATIA，才能确保零件编号与CATIA文件一致。\n\n"
                "是否立即执行写回？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return
            self._write_back(close_on_success=False)
            # Always stop here: if write-back succeeded the user should
            # re-trigger the rename; if it failed _dirty is still set.
            return

        # ── Build deduplicated list of (filepath, target_pn) pairs ──────────
        to_rename: list[tuple[str, str]] = []
        seen_fps: set[str] = set()
        for row in self._rows:
            fp = str(row.get("_filepath", ""))
            if not fp or fp in seen_fps:
                continue
            seen_fps.add(fp)
            orig_pn = str(row.get("Part Number", ""))
            pn = str(self._pn_data.get(orig_pn, {}).get("Part Number", orig_pn))
            if pn and Path(fp).stem != pn:
                to_rename.append((fp, pn))

        if not to_rename:
            QMessageBox.information(self, "无需改名", "所有文件名已与零件编号一致。")
            return

        # ── Ask whether to delete original files after SaveAs ───────────────
        delete_old = QMessageBox.question(
            self, "是否删除旧文件",
            "另存为完成后，是否删除旧文件？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        ) == QMessageBox.StandardButton.Yes

        QMessageBox.information(
            self, "请在CATIA中继续操作",
            "准备就绪，请在CATIA中确认后续操作。"
        )

        renamed_count = 0

        for fp, pn in reversed(to_rename):
            # Validity check – skip and warn if PN contains illegal characters
            if not _PN_VALID_RE.fullmatch(pn):
                QMessageBox.warning(
                    self, "零件编号含非法字符",
                    f"零件编号 「{pn}」 含有非法字符。\n"
                    "只允许：字母(A-Z, a-z)、数字(0-9)、下划线(_)、连字符(-)、空格( )。\n"
                    "请在表格中修改此零件编号后重试。"
                )
                continue

            if not Path(fp).exists():
                continue

            ext = Path(fp).suffix
            new_fp = str(Path(fp).parent / (pn + ext))

            # ── Perform SaveAs via CATIA COM ─────────────────────────────────
            try:
                from pycatia import catia as _pycatia
                caa = _pycatia()
                application = caa.application
                application.visible = True
                documents = application.documents

                src = Path(fp).resolve()

                def _find_doc(docs, path: Path):
                    """Return the first open document whose full_name resolves to path."""
                    for i in range(1, docs.count + 1):
                        try:
                            d = docs.item(i)
                            if Path(d.full_name).resolve() == path:
                                return d
                        except Exception:
                            pass
                    return None

                # Open the document in CATIA if it is not already open
                target_doc = _find_doc(documents, src)
                if target_doc is None:
                    documents.open(str(src))
                    target_doc = _find_doc(documents, src)

                if target_doc is None:
                    QMessageBox.warning(
                        self, "无法找到文档",
                        f"无法在CATIA中找到或打开文档：\n{fp}"
                    )
                    continue

                target_existed_before = Path(new_fp).exists()
                target_doc.com_object.SaveAs(new_fp)

                # Delete the original file if requested and SaveAs produced a
                # different path (avoid deleting when old == new).
                if delete_old and Path(fp).resolve() != Path(new_fp).resolve():
                    try:
                        os.remove(fp)
                    except Exception as del_err:
                        logger.warning(f"Failed to delete old file {fp}: {del_err}")

                # Update in-memory rows to reflect the new path and filename
                for row in self._rows:
                    if str(row.get("_filepath", "")) == fp:
                        row["_filepath"] = new_fp
                        row["Filename"] = pn
                renamed_count += 1
            except Exception as e:
                # If the target already existed before the call and the source
                # file is still intact, the user most likely clicked "No" when
                # CATIA asked whether to overwrite – treat this as a silent skip.
                if target_existed_before and Path(fp).exists():
                    continue
                QMessageBox.warning(self, "另存为失败", f"文件「{Path(fp).name}」另存为失败：\n{e}")

        if renamed_count > 0:
            QMessageBox.information(
                self, "改名完成", f"已成功将 {renamed_count} 个文件通过CATIA另存为功能改名。"
            )
            self._populate_table()

    def _write_back(self, *, close_on_success: bool):
        """Build a dirty-only data set, write it to CATIA, then optionally close.

        Only fields that were actually changed since the last load/apply are
        written, which avoids redundant COM round-trips and makes write-back
        significantly faster for large assemblies.
        """
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请选择一个CATProduct文件。")
                return

        # Build filtered dict: only changed fields per original PN
        dirty_data: dict[str, dict[str, str]] = {}
        for pn, dirty_cols in self._dirty.items():
            if pn not in self._pn_data:
                continue
            changed = {col: self._pn_data[pn][col]
                       for col in dirty_cols if col in self._pn_data[pn]}
            if changed:
                dirty_data[pn] = changed

        if not dirty_data:
            if close_on_success:
                self.accept()
            else:
                QMessageBox.information(self, "无更改", "没有检测到任何修改，无需写回。")
            return

        self._save_btn.setEnabled(False)
        self._finish_btn.setEnabled(False)
        QApplication.processEvents()

        try:
            _write_bom_to_catia(file_path, dirty_data, self._all_custom_columns)
        except Exception as e:
            logger.error(f"Failed to write BOM back to CATIA: {e}")
            self._save_btn.setEnabled(True)
            self._finish_btn.setEnabled(True)
            QMessageBox.critical(
                self, "写回失败",
                f"写回CATIA时出错：\n{e}\n\n请确保CATIA已启动。"
            )
            return

        # Sync original snapshot and clear dirty tracking for written fields
        for pn, changed in dirty_data.items():
            if pn in self._original_pn_data:
                self._original_pn_data[pn].update(changed)
            if pn in self._dirty:
                self._dirty[pn] -= set(changed.keys())
                if not self._dirty[pn]:
                    del self._dirty[pn]

        self._save_btn.setEnabled(True)
        self._finish_btn.setEnabled(True)

        if close_on_success:
            QMessageBox.information(self, "完成",
                "BOM属性已成功写回CATIA，请在CATIA中手动保存文件。")
            self.accept()
        else:
            QMessageBox.information(self, "应用成功",
                "BOM属性已成功写回CATIA，请在CATIA中手动保存文件。")

    def _apply_changes(self):
        """Write modified properties back to CATIA and keep the dialog open."""
        self._write_back(close_on_success=False)

    def _finish_and_close(self):
        """Write modified properties back to CATIA and close the dialog."""
        self._write_back(close_on_success=True)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("CATIA Companion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
